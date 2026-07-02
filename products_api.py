"""
products_api.py — AI 商品搬運中心 Blueprint
從 app.py 拆出的所有商品相關路由與邏輯。
包含：/admin/products、/admin/products/store-scan、/admin/brand-settings、
      /api/products/*、/api/store-scan/*、/api/brand-profiles/*
LINE Bot / FB Messenger / CRM 完全不在此檔案中。
"""
import os, json, time, threading, urllib.request, io, base64
from flask import Blueprint, request, jsonify, render_template_string, Response

try:
    from zhconv import convert as _zhconv
except Exception:
    _zhconv = None

def _to_tw(text):
    """簡體轉繁體（顯示用，不改 DB 原始資料）"""
    if not text or not _zhconv:
        return text
    try:
        return _zhconv(text, "zh-tw")
    except Exception:
        return text

# ── 設定（從環境變數，與 app.py 一致）────────────────────────────
DATABASE_URL         = os.getenv("DATABASE_URL", "")
ANTHROPIC_API_KEY    = os.getenv("ANTHROPIC_API_KEY", "")
OPENAI_API_KEY       = os.getenv("OPENAI_API_KEY", "")
SUPABASE_URL         = os.getenv("SUPABASE_URL", "")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")
SUPABASE_BUCKET      = "chat-images"
GITHUB_TOKEN         = os.getenv("GITHUB_TOKEN", "")
GITHUB_REPO          = "true555132-svg/jsimple-linebot"
ADMIN_PASSWORD       = os.getenv("ADMIN_PASSWORD", "")

_db_lock = threading.Lock()

products_bp = Blueprint("products", __name__)

# ── Pipeline DB migration ─────────────────────────────────────────
def _migrate_pipeline_columns():
    """Add pipeline columns to product_jobs / brand_profiles if not present."""
    if not DATABASE_URL:
        return
    pj_cols = [
        "analysis_json TEXT",
        "search_intent_json TEXT",
        "competitor_json TEXT",
        "pipeline_used BOOLEAN DEFAULT FALSE",
        "pipeline_log TEXT",
    ]
    bp_cols = [
        "positioning TEXT DEFAULT ''",
        "target_audience TEXT DEFAULT ''",
        "forbidden_words TEXT DEFAULT ''",
        "copy_length TEXT DEFAULT ''",
        "faq_strategy TEXT DEFAULT ''",
        "aeo_rules TEXT DEFAULT ''",
        "bg_composite BOOLEAN DEFAULT FALSE",
    ]
    try:
        import sys
        conn = _pg_conn(); cur = conn.cursor()
        for col_def in pj_cols:
            cur.execute(f"ALTER TABLE product_jobs ADD COLUMN IF NOT EXISTS {col_def}")
        for col_def in bp_cols:
            cur.execute(f"ALTER TABLE brand_profiles ADD COLUMN IF NOT EXISTS {col_def}")
        conn.commit(); cur.close(); conn.close()
        print("[Pipeline] DB migration 完成", file=sys.stderr)
    except Exception as e:
        import sys
        print(f"[Pipeline] DB migration 失敗：{e}", file=sys.stderr)

# ── DB 連線 ──────────────────────────────────────────────────────
def _pg_conn():
    import psycopg2
    return psycopg2.connect(DATABASE_URL)

# ── Auth（與 app.py 保持一致，避免 circular import）──────────────
def auth_required():
    candidates = [
        request.args.get("admin_key", ""),
        request.args.get("key", ""),
    ]
    try:
        body = request.get_json(silent=True, force=True) or {}
        candidates += [body.get("admin_key", ""), body.get("key", "")]
    except Exception:
        pass
    for k in candidates:
        if k and k == ADMIN_PASSWORD:
            return True, k
    return False, ""

def check_auth():
    key = request.args.get("key", "")
    return key == ADMIN_PASSWORD, key

# ── Login HTML（複製自 app.py，避免 circular import）─────────────
LOGIN_HTML = """<!DOCTYPE html>
<html lang="zh-TW"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>J SIMPLE Bot 後台</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,sans-serif;background:#f5f5f5}
.wrap{max-width:340px;margin:80px auto;background:#fff;border-radius:16px;padding:36px;box-shadow:0 2px 16px rgba(0,0,0,.1);text-align:center}
h2{margin-bottom:22px;font-size:19px;color:#333}
input{width:100%;padding:11px;border:1px solid #ddd;border-radius:8px;font-size:15px;margin-bottom:14px}
button{width:100%;background:#00c300;color:#fff;border:none;border-radius:8px;padding:12px;font-size:15px;font-weight:600;cursor:pointer}
.err{color:red;margin-top:10px;font-size:13px}
</style></head><body>
<div class="wrap">
  <h2>🔐 J SIMPLE 後台</h2>
  <form method="POST" action="/admin/login">
    <input type="hidden" name="next" value="{{ next }}">
    <input type="password" name="password" placeholder="請輸入密碼" autofocus>
    <button type="submit">登入</button>
  </form>
  {% if error %}<p class="err">{{ error }}</p>{% endif %}
</div>
</body></html>"""

# ── Upload helpers（複製自 app.py）───────────────────────────────
def upload_image_to_supabase(filename: str, data: bytes, content_type: str = "image/jpeg") -> tuple:
    if not SUPABASE_SERVICE_KEY:
        return "", "SUPABASE_SERVICE_KEY 未設定"
    try:
        upload_url = f"{SUPABASE_URL}/storage/v1/object/{SUPABASE_BUCKET}/{filename}"
        req = urllib.request.Request(
            upload_url, data=data, method="POST",
            headers={
                "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                "Content-Type": content_type,
                "x-upsert": "true",
            }
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            r.read()
        public_url = f"{SUPABASE_URL}/storage/v1/object/public/{SUPABASE_BUCKET}/{filename}"
        return public_url, ""
    except Exception as e:
        return "", str(e)

def upload_image_to_github(filename: str, data: bytes) -> tuple:
    import sys
    if not GITHUB_TOKEN:
        return "", "GITHUB_TOKEN 未設定"
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/images/{filename}"
    try:
        req = urllib.request.Request(url, headers={
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github.v3+json",
        })
        sha = None
        try:
            with urllib.request.urlopen(req, timeout=15) as r:
                sha = json.loads(r.read()).get("sha")
        except Exception:
            pass
        body = {"message": f"upload image: {filename}", "content": base64.b64encode(data).decode()}
        if sha:
            body["sha"] = sha
        req2 = urllib.request.Request(url, data=json.dumps(body).encode(), method="PUT", headers={
            "Authorization": f"token {GITHUB_TOKEN}",
            "Content-Type": "application/json",
            "Accept": "application/vnd.github.v3+json",
        })
        with urllib.request.urlopen(req2, timeout=30) as r:
            json.loads(r.read())
        raw_url = f"https://raw.githubusercontent.com/{GITHUB_REPO}/main/images/{filename}"
        return raw_url, ""
    except urllib.error.HTTPError as e:
        body_err = e.read().decode(errors="ignore")
        msg = f"GitHub HTTP {e.code}: {body_err[:200]}"
        print(f"[GitHub Upload Error] {msg}", file=sys.stderr)
        return "", msg
    except Exception as e:
        msg = str(e)
        print(f"[GitHub Upload Error] {msg}", file=sys.stderr)
        return "", msg

# ── 品牌設定 ──────────────────────────────────────────────────────
BRAND_PROFILES = {
    "jsimple": {
        "name": "JSIMPLE",
        "category": "高架床、系統家具",
        "style": "簡潔、功能導向、現代風格。強調空間利用、承重規格、材質安全。",
        "tone": "直接說明功能與規格，像設計師推薦，不像業務推銷。",
    },
    "lander": {
        "name": "朗德燈具",
        "category": "燈具、照明",
        "style": "質感、設計感、氛圍營造。強調光線效果、設計美感、節能規格（W數、流明）。",
        "tone": "有畫面感，讓人想像裝上後的居家氛圍。",
    },
    "filterbreath": {
        "name": "濾呼吸",
        "category": "空氣濾網、淨化設備",
        "style": "健康、數據導向、信任感。強調過濾效率（等級）、適用機型、更換週期。",
        "tone": "用具體數字說話，像健康產品的專業建議，不誇大。",
    },
}

# ── Brand profile DB helpers ──────────────────────────────────────
def _bp_all():
    if not DATABASE_URL:
        return [{"brand_key": k, **v, "custom_prompt": "", "image_style": "", "seo_direction": "", "enabled": True} for k, v in BRAND_PROFILES.items()]
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute("SELECT brand_key,name,category,style,tone,custom_prompt,COALESCE(image_style,''),COALESCE(seo_direction,''),COALESCE(enabled,TRUE),COALESCE(bg_composite,FALSE) FROM brand_profiles ORDER BY brand_key")
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"brand_key": r[0], "name": r[1], "category": r[2],
                 "style": r[3], "tone": r[4], "custom_prompt": r[5] or "",
                 "image_style": r[6] or "", "seo_direction": r[7] or "", "enabled": bool(r[8]),
                 "bg_composite": bool(r[9])} for r in rows]
    except Exception:
        return [{"brand_key": k, **v, "custom_prompt": "", "image_style": "", "seo_direction": "", "enabled": True} for k, v in BRAND_PROFILES.items()]

def _bp_get(brand_key):
    if not DATABASE_URL:
        return BRAND_PROFILES.get(brand_key, {})
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute("SELECT name,category,style,tone,custom_prompt,COALESCE(image_style,''),COALESCE(seo_direction,''),COALESCE(enabled,TRUE),COALESCE(bg_composite,FALSE) FROM brand_profiles WHERE brand_key=%s", (brand_key,))
        row = cur.fetchone(); cur.close(); conn.close()
        if row:
            return {"name": row[0], "category": row[1], "style": row[2],
                    "tone": row[3], "custom_prompt": row[4] or "",
                    "image_style": row[5] or "", "seo_direction": row[6] or "",
                    "enabled": bool(row[7]), "bg_composite": bool(row[8])}
    except Exception:
        pass
    return BRAND_PROFILES.get(brand_key, {})

def _bp_save(brand_key, name, category, style, tone, custom_prompt="", image_style="", seo_direction="", enabled=True, bg_composite=False):
    if not DATABASE_URL:
        return False
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO brand_profiles(brand_key,name,category,style,tone,custom_prompt,image_style,seo_direction,enabled,bg_composite,updated_at)
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT(brand_key) DO UPDATE
            SET name=%s,category=%s,style=%s,tone=%s,custom_prompt=%s,image_style=%s,seo_direction=%s,enabled=%s,bg_composite=%s,updated_at=%s
        """, (brand_key, name, category, style, tone, custom_prompt, image_style, seo_direction, enabled, bg_composite, time.time(),
              name, category, style, tone, custom_prompt, image_style, seo_direction, enabled, bg_composite, time.time()))
        conn.commit(); cur.close(); conn.close()
        return True
    except Exception as e:
        import sys; print(f"[BP Save] {e}", file=sys.stderr)
        return False

# ── Product job DB helpers ────────────────────────────────────────
def _pj_insert(url, platform, brand=""):
    if not DATABASE_URL:
        return None
    try:
        with _db_lock:
            conn = _pg_conn(); cur = conn.cursor()
            cur.execute(
                "INSERT INTO product_jobs (url,platform,brand,status,created_at,updated_at) VALUES (%s,%s,%s,'pending',%s,%s) RETURNING id",
                (url, platform, brand, time.time(), time.time())
            )
            job_id = cur.fetchone()[0]
            conn.commit(); cur.close(); conn.close()
            return job_id
    except Exception as e:
        import sys; print(f"[PJ Insert] {e}", file=sys.stderr)
        return None

def _pj_update(job_id, **fields):
    if not DATABASE_URL or not fields:
        return
    try:
        fields["updated_at"] = time.time()
        set_clause = ", ".join(f"{k}=%s" for k in fields)
        with _db_lock:
            conn = _pg_conn(); cur = conn.cursor()
            cur.execute(f"UPDATE product_jobs SET {set_clause} WHERE id=%s", list(fields.values()) + [job_id])
            conn.commit(); cur.close(); conn.close()
    except Exception as e:
        import sys; print(f"[PJ Update] {e}", file=sys.stderr)

def _pj_list(limit=50):
    if not DATABASE_URL:
        return []
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id,url,platform,status,raw_title,ai_name,ai_desc,ai_keywords,error_msg,created_at,brand,COALESCE(category,''),COALESCE(listing_status,'草稿'),COALESCE(img_status,''),COALESCE(main_image,''),raw_images FROM product_jobs ORDER BY created_at DESC LIMIT %s",
            (limit,)
        )
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id":r[0],"url":r[1],"platform":r[2],"status":r[3],"raw_title":_to_tw(r[4]),"ai_name":r[5],"ai_desc":r[6],"ai_keywords":r[7],"error_msg":r[8],"created_at":r[9],"brand":r[10] or "","category":r[11] or "","listing_status":r[12] or "草稿","img_status":r[13] or "","main_image":r[14] or "","raw_images":json.loads(r[15] or "[]")} for r in rows]
    except Exception:
        return []

def _pj_get(job_id):
    if not DATABASE_URL:
        return None
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id,url,platform,status,raw_title,raw_desc,raw_images,raw_price,ai_name,ai_desc,ai_keywords,error_msg,created_at,processed_images,img_status,raw_extra,brand,COALESCE(translated_images,'[]'),COALESCE(translate_status,''),"
            "COALESCE(category,''),COALESCE(price_min,''),COALESCE(price_max,''),COALESCE(shopee_title,''),COALESCE(website_name,''),COALESCE(features,''),COALESCE(seo_desc,''),COALESCE(faq,'[]'),COALESCE(main_image,''),COALESCE(listing_status,'草稿'),"
            "COALESCE(analysis_json,''),COALESCE(search_intent_json,''),COALESCE(competitor_json,''),COALESCE(pipeline_used,FALSE),COALESCE(pipeline_log,'') "
            "FROM product_jobs WHERE id=%s",
            (job_id,)
        )
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close(); return None
        raw_extra = json.loads(row[15] or "{}")
        if isinstance(raw_extra.get("specs"), list):
            raw_extra["specs"] = [
                {"name": _to_tw(s.get("name","")) if isinstance(s, dict) else _to_tw(str(s)),
                 "value": _to_tw(s.get("value","")) if isinstance(s, dict) else ""}
                for s in raw_extra["specs"]
            ]
        if isinstance(raw_extra.get("sku_prices"), list):
            raw_extra["sku_prices"] = [
                {**p, "label": _to_tw(p.get("label",""))} if isinstance(p, dict) else p
                for p in raw_extra["sku_prices"]
            ]
        result = {"id":row[0],"url":row[1],"platform":row[2],"status":row[3],"raw_title":_to_tw(row[4]),"raw_desc":_to_tw(row[5]),"raw_images":json.loads(row[6] or "[]"),"raw_price":row[7],"ai_name":row[8],"ai_desc":row[9],"ai_keywords":row[10],"error_msg":row[11],"created_at":row[12],"processed_images":json.loads(row[13] or "[]"),"img_status":row[14] or "","raw_extra":raw_extra,"brand":row[16] or "","translated_images":json.loads(row[17] or "[]"),"translate_status":row[18] or "",
                  "category":row[19] or "","price_min":row[20] or "","price_max":row[21] or "","shopee_title":row[22] or "","website_name":row[23] or "","features":row[24] or "","seo_desc":row[25] or "","faq":json.loads(row[26] or "[]"),"main_image":row[27] or "","listing_status":row[28] or "草稿",
                  "analysis_json":json.loads(row[29] or "{}") if row[29] else {},"search_intent_json":json.loads(row[30] or "{}") if row[30] else {},"competitor_json":json.loads(row[31] or "{}") if row[31] else {},"pipeline_used":bool(row[32]),"pipeline_log":row[33] or ""}
        try:
            cur.execute("SELECT product_images FROM product_jobs WHERE id=%s", (job_id,))
            pi_row = cur.fetchone()
            pi = json.loads((pi_row[0] if pi_row else None) or "{}")
            for cat in ("sku_images",):
                if isinstance(pi.get(cat), list):
                    pi[cat] = [
                        {**i, "label": _to_tw(i.get("label",""))} if isinstance(i, dict) else i
                        for i in pi[cat]
                    ]
            result["product_images"] = pi
        except Exception:
            result["product_images"] = {}
        cur.close(); conn.close()
        return result
    except Exception:
        return None

def _pj_delete(job_id):
    if not DATABASE_URL:
        return
    try:
        with _db_lock:
            conn = _pg_conn(); cur = conn.cursor()
            cur.execute("DELETE FROM product_jobs WHERE id=%s", (job_id,))
            conn.commit(); cur.close(); conn.close()
    except Exception:
        pass

# ── Store Scan DB helpers ─────────────────────────────────────────
def _ss_insert(url, platform):
    if not DATABASE_URL:
        return None
    try:
        with _db_lock:
            conn = _pg_conn(); cur = conn.cursor()
            cur.execute(
                "INSERT INTO store_scan_jobs (url,platform,status,created_at,updated_at) VALUES (%s,%s,'pending',%s,%s) RETURNING id",
                (url, platform, time.time(), time.time())
            )
            job_id = cur.fetchone()[0]
            conn.commit(); cur.close(); conn.close()
            return job_id
    except Exception as e:
        import sys; print(f"[SS Insert] {e}", file=sys.stderr)
        return None

def _ss_update(job_id, **fields):
    if not DATABASE_URL or not fields:
        return
    try:
        fields["updated_at"] = time.time()
        set_clause = ", ".join(f"{k}=%s" for k in fields)
        with _db_lock:
            conn = _pg_conn(); cur = conn.cursor()
            cur.execute(f"UPDATE store_scan_jobs SET {set_clause} WHERE id=%s", list(fields.values()) + [job_id])
            conn.commit(); cur.close(); conn.close()
    except Exception as e:
        import sys; print(f"[SS Update] {e}", file=sys.stderr)

def _ss_get(job_id):
    if not DATABASE_URL:
        return None
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id,url,platform,status,item_count,error_msg,created_at FROM store_scan_jobs WHERE id=%s",
            (job_id,)
        )
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close(); return None
        result = {"id":row[0],"url":row[1],"platform":row[2],"status":row[3],"item_count":row[4],"error_msg":row[5],"created_at":row[6]}
        cur.execute(
            "SELECT id,title,url,image,price,shop_name,platform,scraped_at,added_to_queue FROM store_scan_items WHERE scan_job_id=%s ORDER BY id ASC",
            (job_id,)
        )
        result["items"] = [{"id":r[0],"title":r[1],"url":r[2],"image":r[3],"price":r[4],"shop_name":r[5],"platform":r[6],"scraped_at":r[7],"added_to_queue":bool(r[8])} for r in cur.fetchall()]
        cur.close(); conn.close()
        return result
    except Exception:
        return None

def _ss_list(limit=20):
    if not DATABASE_URL:
        return []
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id,url,platform,status,item_count,created_at FROM store_scan_jobs ORDER BY created_at DESC LIMIT %s",
            (limit,)
        )
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id":r[0],"url":r[1],"platform":r[2],"status":r[3],"item_count":r[4],"created_at":r[5]} for r in rows]
    except Exception:
        return []

def _ss_insert_items(scan_job_id, items):
    if not DATABASE_URL or not items:
        return 0
    try:
        with _db_lock:
            conn = _pg_conn(); cur = conn.cursor()
            now = time.time()
            for it in items:
                cur.execute(
                    "INSERT INTO store_scan_items (scan_job_id,title,url,image,price,shop_name,platform,scraped_at,created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (scan_job_id, it.get("title","")[:200], it.get("url",""), it.get("image",""), it.get("price","")[:50], it.get("shop_name","")[:100], it.get("platform",""), it.get("scraped_at",""), now)
                )
            conn.commit(); cur.close(); conn.close()
            return len(items)
    except Exception as e:
        import sys; print(f"[SS InsertItems] {e}", file=sys.stderr)
        return 0

def _ss_mark_added(item_ids):
    if not DATABASE_URL or not item_ids:
        return
    try:
        with _db_lock:
            conn = _pg_conn(); cur = conn.cursor()
            cur.execute("UPDATE store_scan_items SET added_to_queue=TRUE WHERE id=ANY(%s)", (item_ids,))
            conn.commit(); cur.close(); conn.close()
    except Exception as e:
        import sys; print(f"[SS MarkAdded] {e}", file=sys.stderr)

# ── 圖片處理工具 ──────────────────────────────────────────────────
def _clean_images(urls):
    import re
    seen = set()
    skip_patterns = re.compile(r'icon|logo|_\d{2}x\d{2}[_.]|_30x|_50x|_60x|_80x|\.ico$', re.IGNORECASE)
    priority_patterns = re.compile(r'_800x|_600x|_790x|_750x|_700x|_[89]\d{2}x|imgextra|mainimg', re.IGNORECASE)
    cleaned = []
    for u in urls:
        u = u.strip()
        if not u: continue
        if u.startswith("//"): u = "https:" + u
        if not u.startswith("http"): continue
        if skip_patterns.search(u): continue
        if u not in seen:
            seen.add(u); cleaned.append(u)
    priority = [u for u in cleaned if priority_patterns.search(u)]
    others   = [u for u in cleaned if not priority_patterns.search(u)]
    return (priority + others)[:10]

def _extract_meta_img(html):
    import re
    results = []
    for prop in ("og:image", "twitter:image"):
        for pat in [
            rf'<meta[^>]+(?:property|name)=["\']?{re.escape(prop)}["\']?[^>]+content=["\']([^"\']+)["\']',
            rf'<meta[^>]+content=["\']([^"\']+)["\'][^>]+(?:property|name)=["\']?{re.escape(prop)}["\']?',
        ]:
            m = re.search(pat, html, re.IGNORECASE)
            if m:
                results.append(m.group(1).strip())
                break
    return results

def _extract_alicdn_imgs(html):
    import re
    pattern = re.compile(
        r'(?:https?:)?//[^"\'<>\s]*?\.alicdn\.com/[^"\'<>\s]+?\.(?:jpg|jpeg|png|webp)(?:\?[^"\'<>\s]*)?',
        re.IGNORECASE
    )
    return list(pattern.findall(html))

def _scrape_images(html):
    urls = []
    urls += _extract_meta_img(html)
    urls += _extract_alicdn_imgs(html)
    return _clean_images(urls)

def _detect_platform(url):
    if "1688.com" in url: return "1688"
    if "taobao.com" in url or "tmall.com" in url: return "taobao"
    return "unknown"

def _extract_meta_text(html, prop):
    import re
    for pat in [
        rf'<meta[^>]+(?:property|name)=["\']?{re.escape(prop)}["\']?[^>]+content=["\']([^"\']+)["\']',
        rf'<meta[^>]+content=["\']([^"\']+)["\'][^>]+(?:property|name)=["\']?{re.escape(prop)}["\']?',
    ]:
        m = re.search(pat, html, re.IGNORECASE)
        if m: return m.group(1).strip()
    return ""

def _extract_title_tag(html):
    import re
    m = re.search(r"<title[^>]*>([^<]+)</title>", html, re.IGNORECASE)
    return m.group(1).strip() if m else ""

def _fetch_html(url):
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
        "Referer": "https://www.google.com/",
    })
    with urllib.request.urlopen(req, timeout=20) as r:
        charset = "utf-8"
        ct = r.headers.get("Content-Type", "")
        if "charset=" in ct:
            charset = ct.split("charset=")[-1].strip().split(";")[0].strip()
        return r.read().decode(charset, errors="ignore")

def _scrape_product(url, platform):
    import re
    result = {"title": "", "desc": "", "images": [], "price": "", "error": ""}
    try:
        html = _fetch_html(url)
        result["title"] = _extract_meta_text(html, "og:title") or _extract_title_tag(html)
        for suffix in ["-1688.com", "- 1688", "- 淘寶網", "-淘寶網", "- Taobao", "_淘寶搜索"]:
            if result["title"].endswith(suffix):
                result["title"] = result["title"][:-len(suffix)].strip()
        result["desc"] = _extract_meta_text(html, "og:description") or _extract_meta_text(html, "description")
        result["images"] = _scrape_images(html)
        price_m = re.search(r'["\']price["\']\s*:\s*["\']?([\d.,]+)["\']?', html)
        if price_m: result["price"] = price_m.group(1)
    except Exception as e:
        result["error"] = str(e)
    return result

# ── Claude AI 改寫 ────────────────────────────────────────────────
def _ai_rewrite(raw_title, raw_desc, price="", brand=""):
    if not ANTHROPIC_API_KEY:
        return {"error": "ANTHROPIC_API_KEY 未設定"}
    bp = _bp_get(brand) if brand else {}
    brand_name     = bp.get("name", "台灣電商品牌")
    brand_category = bp.get("category", "商品")
    brand_style    = bp.get("style", "簡潔、專業、官網風格。")
    brand_tone     = bp.get("tone", "直接說明功能，不像業務推銷。")
    brand_seo_dir  = bp.get("seo_direction", "")
    custom_prompt  = bp.get("custom_prompt", "")
    parts = []
    if raw_title: parts.append(f"原始標題：{raw_title}")
    if price:     parts.append(f"參考價格：{price}")
    if raw_desc:  parts.append(f"原始描述：{raw_desc[:1500]}")
    product_block = chr(10).join(parts)
    if custom_prompt and custom_prompt.strip():
        prompt = custom_prompt.replace("{product}", product_block)
        if "{product}" not in custom_prompt:
            prompt = custom_prompt + "\n\n" + product_block
    else:
        seo_dir_line = f"\nSEO 關鍵字方向：{brand_seo_dir}" if brand_seo_dir.strip() else ""
        prompt = f"""你是「{brand_name}」品牌的文案編輯，負責{brand_category}類商品。
品牌文案風格：{brand_style}
語氣要求：{brand_tone}{seo_dir_line}

請將以下中國電商商品資料改寫成台灣官網風格。
不要直接翻譯淘寶標題，要重新定位商品。
用具體數字代替模糊描述。不用感嘆號堆疊。
禁止用詞：喔、恩、那個、就是說、其實、基本上、保證、一定、絕對。

{product_block}

輸出格式（只輸出 JSON，不要其他文字，不要在值的內容加大括號）：
{{
  "name": "商品名稱（30字以內，繁體中文，簡潔有官網感，不堆砌關鍵字）",
  "desc": "商品描述（200-400字，條列式，繁體中文，用具體數字，不感嘆號堆疊）",
  "keywords": "關鍵字1,關鍵字2,關鍵字3,關鍵字4,關鍵字5",
  "shopee_title": "蝦皮標題（關鍵字堆疊風格，40字以內）",
  "website_name": "官網商品名稱（簡潔版，20字以內）",
  "features": "・特色第一點\\n・特色第二點\\n・特色第三點\\n・特色第四點",
  "price_min": "建議售價下限（純數字，台幣）",
  "price_max": "建議售價上限（純數字，台幣）",
  "seo_desc": "SEO 描述（80字以內，含品牌與主要關鍵字）",
  "faq": [{{"q":"購買前問題1","a":"具體回答1"}},{{"q":"問題2","a":"回答2"}},{{"q":"問題3","a":"回答3"}}]
}}"""
    try:
        req_data = json.dumps({
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 2048,
            "messages": [{"role": "user", "content": prompt}]
        }).encode()
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=req_data, method="POST",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            }
        )
        with urllib.request.urlopen(req, timeout=60) as r:
            resp = json.loads(r.read().decode())
        text = resp["content"][0]["text"].strip()
        import re
        m = re.search(r'\{[\s\S]*\}', text)
        if m:
            result = json.loads(m.group())
            return {
                "name": result.get("name",""), "desc": result.get("desc",""), "keywords": result.get("keywords",""),
                "shopee_title": result.get("shopee_title",""), "website_name": result.get("website_name",""),
                "features": result.get("features",""), "price_min": str(result.get("price_min","") or ""),
                "price_max": str(result.get("price_max","") or ""), "seo_desc": result.get("seo_desc",""),
                "faq": result.get("faq", []) if isinstance(result.get("faq"), list) else [],
            }
        return {"error": f"AI 回傳格式錯誤: {text[:200]}"}
    except Exception as e:
        return {"error": str(e)}

# ── 背景任務 ──────────────────────────────────────────────────────
def _process_product_job(job_id, url, platform):
    _pj_update(job_id, status="scraping")
    scraped = _scrape_product(url, platform)
    if scraped.get("error") and not scraped.get("title"):
        _pj_update(job_id, status="error", error_msg=f"爬取失敗：{scraped['error']}")
        return
    raw_title  = scraped.get("title", "")
    raw_desc   = scraped.get("desc", "")
    raw_images = scraped.get("images", [])
    raw_price  = scraped.get("price", "")
    _pj_update(job_id, status="rewriting", raw_title=raw_title, raw_desc=raw_desc,
               raw_images=json.dumps(raw_images, ensure_ascii=False), raw_price=raw_price)
    if not raw_title and not raw_desc:
        _pj_update(job_id, status="error", error_msg="無法取得商品資料（頁面可能需要登入）")
        return
    ai = _ai_rewrite(raw_title, raw_desc, raw_price)
    if "error" in ai:
        _pj_update(job_id, status="error", error_msg=f"AI 改寫失敗：{ai['error']}")
        return
    _pj_update(job_id, status="done", ai_name=ai.get("name",""), ai_desc=ai.get("desc",""), ai_keywords=ai.get("keywords",""),
               shopee_title=ai.get("shopee_title",""), website_name=ai.get("website_name",""), features=ai.get("features",""),
               price_min=ai.get("price_min",""), price_max=ai.get("price_max",""), seo_desc=ai.get("seo_desc",""),
               faq=json.dumps(ai.get("faq",[]), ensure_ascii=False))

def _run_ai_rewrite_for_job(job_id, use_pipeline=None):
    """
    use_pipeline=None  → 依環境變數 USE_AI_PIPELINE 決定
    use_pipeline=True  → 強制使用 pipeline
    use_pipeline=False → 強制使用舊版 _ai_rewrite
    """
    import sys
    job = _pj_get(job_id)
    if not job: return
    raw_title = job.get("raw_title", "")
    raw_desc  = job.get("raw_desc", "")
    raw_price = job.get("raw_price", "")
    if not raw_title and not raw_desc:
        _pj_update(job_id, status="error", error_msg="無法取得商品資料（頁面可能需要登入）")
        return
    brand = job.get("brand", "")
    _pj_update(job_id, status="rewriting")

    try:
        from ai_pipeline import run_pipeline, USE_AI_PIPELINE as _PIPE_ENV
        _should_pipeline = use_pipeline if use_pipeline is not None else _PIPE_ENV
    except ImportError:
        _should_pipeline = False

    if _should_pipeline:
        print(f"[Pipeline] job {job_id} 使用 AI Pipeline...", file=sys.stderr)
        pr = run_pipeline(raw_title, raw_desc, raw_price, brand)
        log_str = "\n".join(pr.get("pipeline_log", []))

        if pr.get("error"):
            print(f"[Pipeline] job {job_id} 失敗，fallback: {pr['error']}", file=sys.stderr)
            _pj_update(job_id, pipeline_used=False, pipeline_log=f"PIPELINE_FAIL: {pr['error']}\n{log_str}")
            ai = _ai_rewrite(raw_title, raw_desc, raw_price, brand)
            if "error" in ai:
                _pj_update(job_id, status="error", error_msg=f"AI 改寫失敗（fallback）：{ai['error']}")
                return
            _pj_update(job_id, status="done",
                       ai_name=ai.get("name",""), ai_desc=ai.get("desc",""), ai_keywords=ai.get("keywords",""),
                       shopee_title=ai.get("shopee_title",""), website_name=ai.get("website_name",""),
                       features=ai.get("features",""), price_min=ai.get("price_min",""),
                       price_max=ai.get("price_max",""), seo_desc=ai.get("seo_desc",""),
                       faq=json.dumps(ai.get("faq",[]), ensure_ascii=False))
            return

        ai   = pr["copy"]
        astr = json.dumps(pr.get("analysis_json") or {},      ensure_ascii=False)
        sstr = json.dumps(pr.get("search_intent_json") or {}, ensure_ascii=False)
        cstr = json.dumps(pr.get("competitor_json") or {},    ensure_ascii=False)
        _pj_update(job_id, status="done",
                   ai_name=ai.get("name",""), ai_desc=ai.get("desc",""), ai_keywords=ai.get("keywords",""),
                   shopee_title=ai.get("shopee_title",""), website_name=ai.get("website_name",""),
                   features=ai.get("features",""), price_min=ai.get("price_min",""),
                   price_max=ai.get("price_max",""), seo_desc=ai.get("seo_desc",""),
                   faq=json.dumps(ai.get("faq",[]), ensure_ascii=False),
                   analysis_json=astr, search_intent_json=sstr, competitor_json=cstr,
                   pipeline_used=True, pipeline_log=log_str)
        return

    # 舊版路徑
    ai = _ai_rewrite(raw_title, raw_desc, raw_price, brand)
    if "error" in ai:
        _pj_update(job_id, status="error", error_msg=f"AI 改寫失敗：{ai['error']}")
        return
    _pj_update(job_id, status="done",
               ai_name=ai.get("name",""), ai_desc=ai.get("desc",""), ai_keywords=ai.get("keywords",""),
               shopee_title=ai.get("shopee_title",""), website_name=ai.get("website_name",""), features=ai.get("features",""),
               price_min=ai.get("price_min",""), price_max=ai.get("price_max",""), seo_desc=ai.get("seo_desc",""),
               faq=json.dumps(ai.get("faq",[]), ensure_ascii=False),
               pipeline_used=False)

# ── 圖片處理（Phase 2A）──────────────────────────────────────────
def _download_image(url):
    try:
        if url.startswith("//"): url = "https:" + url
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://www.1688.com/",
        })
        with urllib.request.urlopen(req, timeout=15) as r:
            return r.read()
    except Exception:
        return None

def _removebg_api(img_bytes):
    """Call remove.bg API; returns PNG bytes with transparent bg, or None on failure."""
    import sys
    api_key = os.getenv("REMOVEBG_API_KEY", "")
    if not api_key:
        return None
    try:
        import requests as _req
        r = _req.post(
            "https://api.remove.bg/v1.0/removebg",
            files={"image_file": ("image.jpg", img_bytes, "image/jpeg")},
            data={"size": "auto"},
            headers={"X-Api-Key": api_key},
            timeout=30,
        )
        if r.status_code == 200:
            return r.content  # PNG with transparent background
        print(f"[RemoveBG] API error {r.status_code}: {r.text[:200]}", file=sys.stderr)
        return None
    except Exception as e:
        print(f"[RemoveBG] exception: {e}", file=sys.stderr)
        return None


def _paste_on_white(img_bytes, size=800):
    """Paste image (supports transparency) onto white canvas, output JPEG bytes."""
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
        canvas = Image.new("RGB", (size, size), (255, 255, 255))
        img.thumbnail((size, size), Image.LANCZOS)
        offset = ((size - img.width) // 2, (size - img.height) // 2)
        canvas.paste(img, offset, mask=img.split()[3])
        out = io.BytesIO()
        canvas.save(out, format="JPEG", quality=88, optimize=True)
        return out.getvalue()
    except Exception:
        return None


def _has_chinese(text):
    return any('一' <= c <= '鿿' for c in (text or ""))


def _translate_prompt_to_en(prompt):
    """中文提示詞 → 英文（用 Claude Haiku，失敗回傳原文）。"""
    if not ANTHROPIC_API_KEY or not _has_chinese(prompt):
        return prompt
    try:
        req_data = json.dumps({
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 200,
            "messages": [{"role": "user", "content":
                f'Translate this image background description to English for an AI image generation prompt. '
                f'Output only the translated text, no explanation:\n{prompt}'}]
        }).encode()
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=req_data, method="POST",
            headers={"x-api-key": ANTHROPIC_API_KEY,
                     "anthropic-version": "2023-06-01",
                     "content-type": "application/json"}
        )
        with urllib.request.urlopen(req, timeout=20) as r:
            en = json.loads(r.read().decode())["content"][0]["text"].strip()
        print(f"[GPT-Image-2] 提示詞翻譯: {prompt[:30]} → {en[:60]}", flush=True)
        return en
    except Exception as e:
        print(f"[GPT-Image-2] 翻譯失敗，使用原文: {e}", flush=True)
        return prompt


def _auto_bg_prompt(product_name, category):
    """商品名稱 + 類別 → 自動生成英文背景提示詞（Claude Haiku）。"""
    if not ANTHROPIC_API_KEY or not product_name:
        return None
    try:
        req_data = json.dumps({
            "model": "claude-haiku-4-5-20251001",
            "max_tokens": 120,
            "messages": [{"role": "user", "content":
                f'Write a short English background scene description for a product photo of: "{product_name}" '
                f'(category: {category or "general"}). '
                f'The background should suit e-commerce use (clean, professional, lifestyle). '
                f'Output only the scene description (1-2 sentences, no quotes).'}]
        }).encode()
        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=req_data, method="POST",
            headers={"x-api-key": ANTHROPIC_API_KEY,
                     "anthropic-version": "2023-06-01",
                     "content-type": "application/json"}
        )
        with urllib.request.urlopen(req, timeout=20) as r:
            prompt = json.loads(r.read().decode())["content"][0]["text"].strip()
        print(f"[GPT-Image-2] 自動提示詞: {prompt[:80]}", flush=True)
        return prompt
    except Exception as e:
        print(f"[GPT-Image-2] 自動提示詞失敗: {e}", flush=True)
        return None


def _gpt_image2_bg(transparent_png, scene_prompt):
    """用 gpt-image-2 在去背後的透明 PNG 周圍生成場景背景。
    transparent_png: remove.bg 回傳的 RGBA PNG bytes（產品不透明、背景透明）。
    回傳 JPEG bytes，失敗回傳 None。
    """
    import sys
    if not OPENAI_API_KEY:
        return None
    try:
        import requests as _req
        from PIL import Image as _Im
        # 縮放至 1024x1024（置中，透明填補），符合 API 限制
        img = _Im.open(io.BytesIO(transparent_png)).convert("RGBA")
        img.thumbnail((1024, 1024), _Im.LANCZOS)
        canvas = _Im.new("RGBA", (1024, 1024), (0, 0, 0, 0))
        offset = ((1024 - img.width) // 2, (1024 - img.height) // 2)
        canvas.paste(img, offset)
        buf = io.BytesIO()
        canvas.save(buf, format="PNG")
        png_bytes = buf.getvalue()

        r = _req.post(
            "https://api.openai.com/v1/images/edits",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
            files={"image": ("product.png", png_bytes, "image/png")},
            data={
                "model": "gpt-image-2",
                "prompt": scene_prompt,
                "n": "1",
                "size": "1024x1024",
            },
            timeout=90,
        )
        if r.status_code == 200:
            item = r.json()["data"][0]
            if item.get("b64_json"):
                img_bytes_out = base64.b64decode(item["b64_json"])
            else:
                img_bytes_out = _req.get(item["url"], timeout=30).content
            result_img = _Im.open(io.BytesIO(img_bytes_out)).convert("RGB")
            out = io.BytesIO()
            result_img.save(out, format="JPEG", quality=92, optimize=True)
            print("[GPT-Image-2] ✅ 背景生成成功", flush=True)
            return out.getvalue()
        print(f"[GPT-Image-2] ❌ {r.status_code}: {r.text[:300]}", file=sys.stderr)
        return None
    except Exception as e:
        import sys as _sys
        print(f"[GPT-Image-2] exception: {e}", file=_sys.stderr)
        return None


def _gpt_image2_composite_bg(transparent_png, scene_prompt, size=1024):
    """安全合成模式：gpt-image-2 純文字生成背景 → PIL 貼上去背產品。
    產品像素 100% 保留，不經過 AI edit，背景由 prompt 生成。
    """
    import sys
    if not OPENAI_API_KEY:
        return None
    try:
        import requests as _req
        from PIL import Image as _Im
        # Step 1: 純文字生成背景
        r = _req.post(
            "https://api.openai.com/v1/images/generations",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}",
                     "Content-Type": "application/json"},
            json={
                "model": "gpt-image-2",
                "prompt": scene_prompt + ", no products, empty scene, photorealistic",
                "n": 1,
                "size": f"{size}x{size}",
            },
            timeout=90,
        )
        if r.status_code != 200:
            print(f"[Composite] ❌ 背景生成失敗 {r.status_code}: {r.text[:200]}", file=sys.stderr)
            return None
        item = r.json()["data"][0]
        if item.get("b64_json"):
            bg_bytes = base64.b64decode(item["b64_json"])
        else:
            bg_bytes = _req.get(item["url"], timeout=30).content
        bg = _Im.open(io.BytesIO(bg_bytes)).convert("RGBA").resize((size, size))

        # Step 2: 去背產品置中貼上（產品像素 100% 保留）
        product = _Im.open(io.BytesIO(transparent_png)).convert("RGBA")
        product.thumbnail((size, size), _Im.LANCZOS)
        offset = ((size - product.width) // 2, (size - product.height) // 2)
        bg.paste(product, offset, mask=product.split()[3])

        out = io.BytesIO()
        bg.convert("RGB").save(out, format="JPEG", quality=92, optimize=True)
        print("[Composite] ✅ 安全合成完成（產品像素未經 AI 修改）", flush=True)
        return out.getvalue()
    except Exception as e:
        print(f"[Composite] exception: {e}", file=sys.stderr)
        return None


def _process_to_white_bg(img_bytes, size=800, scene_prompt=None, bg_composite=False):
    """去背 → AI 背景（edit 或 composite 模式）；否則貼白底。"""
    removed = _removebg_api(img_bytes)
    if removed:
        if scene_prompt and OPENAI_API_KEY:
            if bg_composite:
                result = _gpt_image2_composite_bg(removed, scene_prompt)
            else:
                result = _gpt_image2_bg(removed, scene_prompt)
            if result:
                return result
        result = _paste_on_white(removed, size)
        if result:
            return result
    # fallback: no bg removal, just center on white
    return _paste_on_white(img_bytes, size)

def _process_images_for_job(job_id):
    job = _pj_get(job_id)
    if not job: return
    raw_imgs = job.get("raw_images", [])
    if not raw_imgs:
        _pj_update(job_id, img_status="no_images"); return

    # 取品牌的 AI 背景提示詞（支援中文自動翻英）與合成模式
    scene_prompt = None
    bg_composite = False
    brand_key = job.get("brand", "")
    bp = {}
    if brand_key:
        bp = _bp_get(brand_key)
        raw_prompt = (bp.get("image_style") or "").strip()
        if raw_prompt:
            scene_prompt = _translate_prompt_to_en(raw_prompt)
        bg_composite = bool(bp.get("bg_composite", False))

    # 提示詞為空 → 用商品名稱自動分析生成
    if not scene_prompt and OPENAI_API_KEY:
        product_name = job.get("ai_name") or job.get("raw_title") or ""
        category     = job.get("category") or bp.get("category", "") or ""
        scene_prompt = _auto_bg_prompt(product_name, category)

    if scene_prompt:
        mode_label = "安全合成" if bg_composite else "AI Edit"
        print(f"[GPT-Image-2] 品牌={brand_key} 模式={mode_label} 提示詞: {scene_prompt[:80]}", flush=True)

    _pj_update(job_id, img_status="processing")
    processed = []
    for i, url in enumerate(raw_imgs[:10]):
        img_bytes = _download_image(url)
        if not img_bytes: continue
        result = _process_to_white_bg(img_bytes, scene_prompt=scene_prompt, bg_composite=bg_composite)
        if not result: continue
        filename = f"products/{job_id}_{i+1}.jpg"
        pub_url, _ = upload_image_to_supabase(filename, result, "image/jpeg")
        if pub_url: processed.append(pub_url)
    _pj_update(job_id,
               processed_images=json.dumps(processed, ensure_ascii=False),
               img_status="done" if processed else "failed")

# ── 圖片翻譯 helpers ──────────────────────────────────────────────
def _get_cjk_font(size=22):
    from PIL import ImageFont
    paths = [
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/arphic/uming.ttc",
        "/tmp/NotoSansCJK.ttc",
    ]
    for p in paths:
        if os.path.exists(p):
            try: return ImageFont.truetype(p, size)
            except Exception: pass
    cache = "/tmp/NotoSansCJK.ttc"
    if not os.path.exists(cache):
        try:
            urllib.request.urlretrieve(
                "https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTF/TraditionalChinese/NotoSansCJKtc-Regular.otf",
                cache
            )
            return ImageFont.truetype(cache, size)
        except Exception:
            pass
    return ImageFont.load_default()

def _sample_bg_info(img_pil, x1, y1, x2, y2):
    from PIL import ImageStat
    W, H = img_pil.size
    region = img_pil.crop((x1, y1, x2, y2))
    rw, rh = region.size
    if rw <= 0 or rh <= 0:
        return 0.0, 200.0, (255, 255, 255)
    stat_rgb = ImageStat.Stat(region)
    variance = (sum(v * v for v in stat_rgb.stddev) / len(stat_rgb.stddev)) ** 0.5
    stat_l = ImageStat.Stat(region.convert("L"))
    avg_brightness = stat_l.mean[0]
    ox1, oy1 = max(0, x1 - 3), max(0, y1 - 3)
    ox2, oy2 = min(W, x2 + 3), min(H, y2 + 3)
    stat_outer = ImageStat.Stat(img_pil.crop((ox1, oy1, ox2, oy2)))
    edge_color = tuple(int(v) for v in stat_outer.mean[:3])
    return variance, avg_brightness, edge_color

def _translate_images_job(job_id, img_urls):
    import re as _re3, traceback
    try:
        import requests as _req
        from PIL import Image, ImageDraw
    except ImportError as e:
        print(f"[translate] import error: {e}")
        _pj_update(job_id, translate_status="failed")
        return

    STABILITY_KEY = os.getenv("STABILITY_API_KEY", "")
    if not STABILITY_KEY:
        print("[translate] STABILITY_API_KEY not set — complex blocks will fallback to Pillow")

    OCR_PROMPT = (
        '請偵測圖片中所有文字區塊，直接輸出 JSON（不要 markdown）。\n'
        '格式：{"blocks":[{"text":"原文","language":"zh-CN/zh-TW/ja/en/number/mixed",'
        '"should_translate":true/false,"translated_text":"繁體（false時留空字串）",'
        '"bbox":[x1,y1,x2,y2],"font_role":"title/subtitle/body/label/number",'
        '"background_type":"white/solid/complex"}]}\n'
        '規則（嚴格遵守）：\n'
        '1. language分類：zh-CN=簡體中文漢字，zh-TW=繁體中文，ja=日文（含ひらがな/カタカナ），en=英文，number=純數字百分比\n'
        '2. should_translate只有純zh-CN才true，其他全部false\n'
        '3. translated_text：簡體→台灣繁體，混合文字只翻中文部分，英文/數字原樣保留\n'
        '4. bbox：[左%,上%,右%,下%]，圖片寬高各為100\n'
        '5. font_role：title=主標題大字，subtitle=副標題，body=說明文字，label=小標籤，number=數字\n'
        '6. background_type：white=白色背景，solid=純色背景，complex=照片/漸層/複雜背景\n'
        '範例（此圖色卡）：\n'
        '- 「海盐蓝」→zh-CN,true,「海鹽藍」,title,white\n'
        '- 「クリームホワイト」→ja,false,\"\",label,white\n'
        '- 「Shading：80%」→en,false,\"\",number,white\n'
        '- 「遮光率：80%」→zh-CN,true,「遮光率：80%」,label,white（數字保留不翻）\n'
        '只輸出JSON，不要說明文字。'
    )

    def _pct2px(pct, dim):
        return max(0, min(dim, int(pct / 100 * dim)))

    def _sample_bg(base_img, x1, y1, x2, y2):
        W, H = base_img.size
        samples = []
        for px, py in [(max(0,x1-8),max(0,y1-8)), (min(W-1,x2+8),max(0,y1-8)),
                       (max(0,x1-8),min(H-1,y2+8)), (min(W-1,x2+8),min(H-1,y2+8))]:
            try:
                s = base_img.getpixel((px, py))
                samples.append(s[:3] if len(s) > 3 else s)
            except Exception:
                pass
        if not samples: return (255, 255, 255)
        return tuple(sum(s[i] for s in samples) // len(samples) for i in range(3))

    def _fit_and_draw(draw, text, x1, y1, x2, y2, role, text_color=(15, 15, 15)):
        from PIL import ImageFont
        box_w, box_h = x2 - x1, y2 - y1
        start_sz = int(box_h * (0.80 if role == "title" else 0.75))
        start_sz = max(10, min(start_sz, 120))
        font, chosen_sz = None, start_sz
        for sz in range(start_sz, 7, -2):
            try:
                f = _get_cjk_font(sz)
                try:
                    bb = f.getbbox(text)
                    tw = bb[2] - bb[0]
                except Exception:
                    tw = len(text) * sz * 0.65
                if tw <= box_w * 1.05:
                    font, chosen_sz = f, sz
                    break
            except Exception:
                pass
        if font is None:
            font = _get_cjk_font(10)
        try:
            bb = font.getbbox(text)
            tw, th = bb[2]-bb[0], bb[3]-bb[1]
        except Exception:
            tw, th = len(text)*chosen_sz, chosen_sz
        if role in ("title", "subtitle"):
            tx = x1 + max(0, (box_w - tw) // 2)
        else:
            tx = x1 + 4
        ty = y1 + max(0, (box_h - th) // 2)
        draw.text((tx, ty), text, fill=text_color, font=font)

    translated_urls = []
    stats = dict(total=0, translated=0, skip_ja=0, skip_en=0, skip_num=0,
                 skip_tw=0, pillow=0, stab=0, stab_fail=0)

    for img_idx, url in enumerate(img_urls):
        print(f"[translate] {img_idx+1}/{len(img_urls)} {url[:70]}")
        try:
            req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0","Referer":"https://www.1688.com/"})
            with urllib.request.urlopen(req, timeout=20) as r:
                img_bytes = r.read()
            media_type = "image/jpeg"
            if img_bytes[:8] == b'\x89PNG\r\n\x1a\n': media_type = "image/png"
            elif img_bytes[:4] == b'RIFF': media_type = "image/webp"

            img_b64 = base64.standard_b64encode(img_bytes).decode()
            claude_resp = urllib.request.urlopen(
                urllib.request.Request(
                    "https://api.anthropic.com/v1/messages",
                    data=json.dumps({
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 3000,
                        "messages": [{"role":"user","content":[
                            {"type":"image","source":{"type":"base64","media_type":media_type,"data":img_b64}},
                            {"type":"text","text":OCR_PROMPT}
                        ]}]
                    }).encode(),
                    headers={"x-api-key":ANTHROPIC_API_KEY,"anthropic-version":"2023-06-01","content-type":"application/json"},
                    method="POST"
                ), timeout=45
            )
            raw_text = json.loads(claude_resp.read())["content"][0]["text"].strip()
            m = _re3.search(r'\{[\s\S]*\}', raw_text)
            if not m:
                print("  [OCR] no JSON"); translated_urls.append(url); continue

            blocks = json.loads(m.group()).get("blocks", [])
            stats['total'] += len(blocks)
            print(f"  [OCR] {len(blocks)} 區塊")

            to_do = []
            for b in blocks:
                lang = b.get("language","")
                if not b.get("should_translate") or not b.get("translated_text","") or lang != "zh-CN":
                    if lang=="ja": stats['skip_ja']+=1
                    elif lang=="en": stats['skip_en']+=1
                    elif lang=="number": stats['skip_num']+=1
                    elif lang=="zh-TW": stats['skip_tw']+=1
                    continue
                stats['translated']+=1
                to_do.append(b)

            if not to_do:
                print("  [OCR] 無需翻譯"); translated_urls.append(url); continue

            from PIL import Image
            img_pil = Image.open(io.BytesIO(img_bytes)).convert("RGB")
            W, H = img_pil.size
            for b in to_do:
                bx = b.get("bbox", [0, 0, 100, 100])
                b['_px'] = (_pct2px(bx[0],W), _pct2px(bx[1],H), _pct2px(bx[2],W), _pct2px(bx[3],H))

            simple, complex_ = [], []
            for b in to_do:
                x1, y1, x2, y2 = b['_px']
                variance, brightness, edge_color = _sample_bg_info(img_pil, x1, y1, x2, y2)
                b['_variance'] = variance
                b['_brightness'] = brightness
                b['_edge_color'] = edge_color
                b['_text_color'] = (255, 255, 255) if brightness < 128 else (15, 15, 15)
                color_label = "white" if brightness < 128 else "dark"
                bg_hint = b.get("background_type", "white")
                if variance < 20:
                    b['_pad'] = 2; simple.append(b)
                    print(f"    [{b.get('text','')[:12]}] var={variance:.1f} bright={brightness:.0f} ({bg_hint}) → pillow | {color_label}")
                elif variance < 50:
                    b['_pad'] = 6; simple.append(b)
                    print(f"    [{b.get('text','')[:12]}] var={variance:.1f} bright={brightness:.0f} ({bg_hint}) → pillow+pad | {color_label}")
                else:
                    complex_.append(b)
                    print(f"    [{b.get('text','')[:12]}] var={variance:.1f} bright={brightness:.0f} ({bg_hint}) → stability | {color_label}")

            stability_saved = len(simple)
            print(f"  [classify] pillow={len(simple)} stability={len(complex_)} | 本張省下 {stability_saved}/{len(to_do)} 次 Stability 呼叫")

            result_img = img_pil.copy()

            if simple:
                draw = ImageDraw.Draw(result_img)
                for b in simple:
                    x1,y1,x2,y2 = b['_px']
                    role = b.get("font_role","body")
                    pad = b.get('_pad', 4)
                    rx1,ry1 = max(0,x1-pad), max(0,y1-pad)
                    rx2,ry2 = min(W,x2+pad), min(H,y2+pad)
                    bg = b.get('_edge_color') or _sample_bg(img_pil, rx1,ry1,rx2,ry2)
                    draw.rectangle([rx1,ry1,rx2,ry2], fill=bg)
                    _fit_and_draw(draw, b.get("translated_text",""), rx1,ry1,rx2,ry2, role, b['_text_color'])
                    stats['pillow']+=1
                    clabel = "white" if b.get('_brightness', 200) < 128 else "dark"
                    print(f"    Pillow [{role}]: '{b.get('text','')}' → '{b.get('translated_text','')}' | {clabel}")

            if complex_ and not STABILITY_KEY:
                draw_fb = ImageDraw.Draw(result_img)
                for b in complex_:
                    x1,y1,x2,y2 = b['_px']
                    role = b.get("font_role","body")
                    pad = 6
                    rx1,ry1 = max(0,x1-pad),max(0,y1-pad)
                    rx2,ry2 = min(W,x2+pad),min(H,y2+pad)
                    bg = b.get('_edge_color') or _sample_bg(img_pil,rx1,ry1,rx2,ry2)
                    draw_fb.rectangle([rx1,ry1,rx2,ry2],fill=bg)
                    _fit_and_draw(draw_fb, b.get("translated_text",""), rx1,ry1,rx2,ry2, role, b['_text_color'])
                    stats['pillow']+=1
                complex_ = []
                print(f"  [Stability] skipped (no key) — fell back to Pillow for all blocks")

            if complex_:
                from PIL import ImageDraw as _ID2
                mask = Image.new("L",(W,H),0)
                dm = _ID2.Draw(mask)
                for b in complex_:
                    x1,y1,x2,y2 = b['_px']
                    mp = 8
                    dm.rectangle([max(0,x1-mp),max(0,y1-mp),min(W,x2+mp),min(H,y2+mp)],fill=255)
                ib = io.BytesIO(); result_img.save(ib,"PNG"); ib.seek(0)
                mb = io.BytesIO(); mask.convert("RGB").save(mb,"PNG"); mb.seek(0)
                inpaint_ok = False
                try:
                    sr = _req.post(
                        "https://api.stability.ai/v2beta/stable-image/edit/erase",
                        headers={"Authorization":f"Bearer {STABILITY_KEY}","Accept":"image/*"},
                        files={"image":("i.png",ib.getvalue(),"image/png"),
                               "mask":("m.png",mb.getvalue(),"image/png")},
                        data={"output_format":"png"}, timeout=60
                    )
                    if sr.status_code == 200:
                        result_img = Image.open(io.BytesIO(sr.content)).convert("RGB")
                        stats['stab']+=len(complex_)
                        inpaint_ok = True
                        print(f"    Stability: {len(complex_)} 區塊")
                    else:
                        print(f"    Stability {sr.status_code} → Pillow fallback")
                        stats['stab_fail']+=len(complex_)
                except Exception as se:
                    print(f"    Stability error: {se} → Pillow fallback")
                    stats['stab_fail']+=len(complex_)

                if not inpaint_ok:
                    draw_fb = ImageDraw.Draw(result_img)
                    for b in complex_:
                        x1,y1,x2,y2 = b['_px']
                        bg = b.get('_edge_color') or _sample_bg(img_pil,x1,y1,x2,y2)
                        draw_fb.rectangle([x1,y1,x2,y2],fill=bg)
                    stats['pillow']+=len(complex_)

                draw_c = ImageDraw.Draw(result_img)
                for b in complex_:
                    x1,y1,x2,y2 = b['_px']
                    role = b.get("font_role","body")
                    _fit_and_draw(draw_c, b.get("translated_text",""),
                                  x1,y1,x2,y2, role, b.get('_text_color', (15,15,15)))

            print(f"  [stats] OCR:{stats['total']} 翻:{stats['translated']} 跳日:{stats['skip_ja']} 跳英:{stats['skip_en']} 跳數:{stats['skip_num']} 跳繁:{stats['skip_tw']} Pillow:{stats['pillow']} Stab:{stats['stab']} Stab失敗:{stats['stab_fail']}")

            out = io.BytesIO()
            result_img.save(out,"JPEG",quality=93)
            fname = f"translated_{job_id}_{img_idx+1}.jpg"
            turl, _ = upload_image_to_supabase(fname, out.getvalue(), "image/jpeg")
            if not turl:
                turl, _ = upload_image_to_github(fname, out.getvalue())
            translated_urls.append(turl or url)
            print(f"  [done] {(turl or url)[:70]}")

        except Exception as e:
            print(f"  [ERROR] {e}")
            traceback.print_exc()
            translated_urls.append(url)

    _pj_update(job_id,
               translated_images=json.dumps(translated_urls, ensure_ascii=False),
               translate_status="done")
    print(f"[translate] 完成 {len(translated_urls)}/{len(img_urls)} 張，stats={stats}")

# ── 匯出 helpers ──────────────────────────────────────────────────
def _pj_list_done():
    if not DATABASE_URL: return []
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id,url,platform,raw_title,raw_price,ai_name,ai_desc,ai_keywords,raw_images,processed_images,created_at FROM product_jobs WHERE status='done' ORDER BY created_at DESC"
        )
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id":r[0],"url":r[1],"platform":r[2],"raw_title":r[3],"raw_price":r[4],
                 "ai_name":r[5],"ai_desc":r[6],"ai_keywords":r[7],
                 "raw_images":json.loads(r[8] or "[]"),
                 "processed_images":json.loads(r[9] or "[]"),
                 "created_at":r[10]} for r in rows]
    except Exception:
        return []

def _export_csv(jobs):
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["平台","AI商品名稱","AI商品描述","SEO關鍵字","原始標題","原始價格","來源URL","圖片URL_1","圖片URL_2","圖片URL_3"])
    for j in jobs:
        imgs = j.get("processed_images") or j.get("raw_images",[])
        w.writerow([
            j["platform"], j["ai_name"], j["ai_desc"], j["ai_keywords"],
            j["raw_title"], j["raw_price"], j["url"],
            imgs[0] if len(imgs)>0 else "",
            imgs[1] if len(imgs)>1 else "",
            imgs[2] if len(imgs)>2 else "",
        ])
    output = buf.getvalue().encode("utf-8-sig")
    return Response(output, mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=products.csv"})

def _export_xlsx(jobs):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import openpyxl.utils
    except ImportError:
        return jsonify({"error": "openpyxl 未安裝"}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品資料"
    headers = ["平台","AI商品名稱","AI商品描述","SEO關鍵字","原始標題","原始價格","來源URL","圖片URL_1","圖片URL_2","圖片URL_3"]
    hfill = PatternFill("solid", fgColor="1a1a1a")
    hfont = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    col_widths = [10, 35, 60, 30, 35, 12, 55, 55, 55, 55]
    for i, cw in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = cw
    for ri, j in enumerate(jobs, 2):
        imgs = j.get("processed_images") or j.get("raw_images",[])
        values = [
            j["platform"], j["ai_name"], j["ai_desc"], j["ai_keywords"],
            j["raw_title"], j["raw_price"], j["url"],
            imgs[0] if len(imgs)>0 else "",
            imgs[1] if len(imgs)>1 else "",
            imgs[2] if len(imgs)>2 else "",
        ]
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ri].height = 60
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products.xlsx"})


STORE_SCAN_HTML = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>店鋪選品掃描</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,sans-serif;background:#f5f5f5;color:#333}
.header{background:#1a1a1a;color:#fff;padding:14px 20px;display:flex;align-items:center;gap:14px}
.header a{color:#888;text-decoration:none;font-size:14px}
.header a:hover{color:#fff}
.header-title{font-size:17px;font-weight:700;flex:1}
.wrap{max-width:960px;margin:0 auto;padding:20px 16px}
.card{background:#fff;border-radius:14px;padding:20px;margin-bottom:16px;box-shadow:0 1px 4px rgba(0,0,0,.08)}
.card h3{font-size:14px;font-weight:700;color:#666;margin-bottom:12px}
.url-row{display:flex;gap:10px}
.url-row input{flex:1;border:1.5px solid #ddd;border-radius:10px;padding:10px 14px;font-size:14px;outline:none;font-family:inherit}
.url-row input:focus{border-color:#1a1a1a}
.btn-primary{background:#1a1a1a;color:#fff;border:none;border-radius:10px;padding:10px 22px;font-size:14px;font-weight:700;cursor:pointer;white-space:nowrap;font-family:inherit}
.btn-primary:hover{background:#333}
.btn-primary:disabled{background:#aaa;cursor:default}
.status-bar{margin-top:12px;font-size:13px;color:#666;min-height:20px}
.spinner{display:inline-block;width:12px;height:12px;border:2px solid #ddd;border-top-color:#666;border-radius:50%;animation:spin .8s linear infinite;vertical-align:middle;margin-right:4px}
@keyframes spin{to{transform:rotate(360deg)}}
/* history */
.history-row{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px}
.hist-chip{background:#f0f0f0;border:1.5px solid transparent;border-radius:20px;padding:4px 12px;font-size:12px;cursor:pointer;font-family:inherit}
.hist-chip:hover{border-color:#1a1a1a}
.hist-chip.active{background:#1a1a1a;color:#fff}
.hist-chip .hst{font-size:10px;color:#aaa;margin-left:4px}
.hist-chip.active .hst{color:#888}
/* grid */
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:12px;margin-top:4px}
.item-card{background:#fff;border-radius:12px;border:2px solid transparent;cursor:pointer;overflow:hidden;transition:border-color .15s;position:relative;display:flex;flex-direction:column}
.item-card:hover{border-color:#ddd}
.item-card.checked{border-color:#1a1a1a}
.item-card input[type=checkbox]{position:absolute;top:8px;left:8px;width:16px;height:16px;cursor:pointer;accent-color:#1a1a1a;z-index:2}
.item-card img{width:100%;aspect-ratio:1;object-fit:cover;background:#f8f8f8}
.item-info{padding:8px 10px 10px;flex:1;display:flex;flex-direction:column;gap:4px}
.item-title{font-size:12px;line-height:1.4;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden;word-break:break-all}
.item-price{font-size:13px;font-weight:700;color:#c62828}
.item-shop{font-size:11px;color:#aaa;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.item-added{font-size:10px;font-weight:700;color:#2e7d32;background:#e8f5e9;border-radius:6px;padding:1px 6px;align-self:flex-start}
.empty{text-align:center;padding:50px 20px;color:#ccc;font-size:14px}
/* action bar */
.action-bar{position:sticky;bottom:0;background:#fff;border-top:1px solid #eee;padding:12px 20px;display:none;align-items:center;gap:12px;z-index:10}
.brand-sel{border:1.5px solid #ddd;border-radius:10px;padding:8px 12px;font-size:13px;font-family:inherit;outline:none;background:#fff;cursor:pointer}
.brand-sel:focus{border-color:#1a1a1a}
.sel-count{font-size:13px;color:#666;flex:1}
.btn-add{background:#2e7d32;color:#fff;border:none;border-radius:10px;padding:10px 22px;font-size:14px;font-weight:700;cursor:pointer;font-family:inherit}
.btn-add:hover{background:#1b5e20}
.btn-add:disabled{background:#aaa;cursor:default}
.sel-all-row{display:flex;gap:10px;align-items:center;margin-bottom:10px;font-size:13px}
.sel-all-row button{background:none;border:1.5px solid #ddd;border-radius:20px;padding:3px 12px;font-size:12px;cursor:pointer;font-family:inherit}
.sel-all-row button:hover{background:#f5f5f5}
.toast{position:fixed;bottom:80px;left:50%;transform:translateX(-50%);background:#333;color:#fff;padding:9px 20px;border-radius:20px;font-size:14px;z-index:999;opacity:0;transition:opacity .3s;pointer-events:none}
.toast.show{opacity:1}
.pf-badge{font-size:11px;font-weight:700;padding:2px 7px;border-radius:7px}
.pf-1688{background:#fff0f0;color:#c62828}
.pf-taobao{background:#fff4e5;color:#e65100}
</style>
</head>
<body>

<div class="header">
  <a href="/admin/products?key={{ key }}">← 商品搬運</a>
  <div class="header-title">店鋪選品掃描 <span style="font-size:12px;font-weight:400;color:#666">Phase 1</span></div>
</div>

<div class="wrap">

  <div class="card">
    <h3>貼上店鋪 / 分類頁連結</h3>
    <div class="url-row">
      <input type="text" id="urlInput" placeholder="https://shop.1688.com/... 或 https://shop.taobao.com/...">
      <button class="btn-primary" id="scanBtn" onclick="startScan()">開始掃描</button>
    </div>
    <div class="status-bar" id="statusBar"></div>
    <div id="historyWrap" style="margin-top:14px;display:none">
      <div style="font-size:12px;color:#aaa;margin-bottom:6px">最近掃描</div>
      <div class="history-row" id="historyRow"></div>
    </div>
  </div>

  <div class="card" id="itemsCard" style="display:none">
    <div class="sel-all-row">
      <span id="scanTitle" style="flex:1;font-size:13px;color:#666"></span>
      <button onclick="toggleAll(true)">全選</button>
      <button onclick="toggleAll(false)">全不選</button>
    </div>
    <div class="grid" id="itemGrid"></div>
  </div>

</div>

<div class="action-bar" id="actionBar">
  <select class="brand-sel" id="brandSel">
    <option value="">不指定品牌</option>
    <option value="jsimple">JSIMPLE — 高架床</option>
    <option value="lander">朗德燈具 — 燈具</option>
    <option value="filterbreath">濾呼吸 — 濾網</option>
  </select>
  <span class="sel-count" id="selCount">已選 0 筆</span>
  <button class="btn-add" id="addBtn" onclick="addToQueue()">加入待上架</button>
</div>

<div class="toast" id="toast"></div>

<script>
const KEY = '{{ key }}';
let currentJobId = null;
let pollTimer = null;

function esc(s){ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }

async function api(path, opts){
  const sep = path.includes('?')?'&':'?';
  const r = await fetch(path+sep+'key='+KEY, opts);
  if(!r.ok) throw new Error(await r.text());
  return r.json();
}

function showToast(msg){
  const t=document.getElementById('toast');
  t.textContent=msg; t.classList.add('show');
  setTimeout(()=>t.classList.remove('show'),2800);
}

async function loadHistory(){
  try{
    const d = await api('/api/store-scan');
    const jobs = d.jobs||[];
    if(!jobs.length) return;
    const row = document.getElementById('historyRow');
    row.innerHTML = jobs.slice(0,8).map(j=>`
      <span class="hist-chip" onclick="loadJob(${j.id})" data-jid="${j.id}">
        <span class="pf-badge pf-${esc(j.platform)}">${esc(j.platform)}</span>
        ${esc((j.url||'').split('/').slice(-2).join('/').slice(-30))}
        <span class="hst">${j.item_count||0}筆</span>
      </span>
    `).join('');
    document.getElementById('historyWrap').style.display='';
  }catch(e){}
}

async function loadJob(jobId){
  currentJobId = jobId;
  document.querySelectorAll('.hist-chip').forEach(el=>el.classList.toggle('active',+el.dataset.jid===jobId));
  document.getElementById('statusBar').innerHTML='<span class="spinner"></span> 載入中...';
  try{
    const job = await api('/api/store-scan/'+jobId);
    updateUI(job);
    if(job.status==='pending'||job.status==='scanning') startPoll();
  }catch(e){
    document.getElementById('statusBar').textContent='載入失敗：'+e.message;
  }
}

async function startScan(){
  const url = document.getElementById('urlInput').value.trim();
  if(!url){alert('請輸入店鋪/分類頁連結');return;}
  document.getElementById('scanBtn').disabled=true;
  document.getElementById('statusBar').innerHTML='<span class="spinner"></span> 建立掃描任務...';
  document.getElementById('itemsCard').style.display='none';
  document.getElementById('actionBar').style.display='none';
  clearTimeout(pollTimer);
  try{
    const r = await api('/api/store-scan',{method:'POST',body:JSON.stringify({url}),headers:{'Content-Type':'application/json'}});
    currentJobId = r.id;
    await loadHistory();
    startPoll();
  }catch(e){
    document.getElementById('statusBar').textContent='錯誤：'+e.message;
    document.getElementById('scanBtn').disabled=false;
  }
}

function startPoll(){
  clearTimeout(pollTimer);
  pollTimer = setTimeout(async()=>{
    try{
      const job = await api('/api/store-scan/'+currentJobId);
      updateUI(job);
      if(job.status==='pending'||job.status==='scanning') startPoll();
      else document.getElementById('scanBtn').disabled=false;
    }catch(e){ startPoll(); }
  },2500);
}

function updateUI(job){
  const sb = document.getElementById('statusBar');
  if(job.status==='pending'){
    sb.innerHTML='<span class="spinner"></span> 等待 local_worker 處理...（請確認 worker 正在執行）';
  }else if(job.status==='scanning'){
    sb.innerHTML='<span class="spinner"></span> 掃描中...';
  }else if(job.status==='done'){
    sb.textContent='掃描完成，共 '+job.item_count+' 筆商品';
    renderItems(job.items||[], job);
  }else if(job.status==='error'){
    sb.textContent='錯誤：'+(job.error_msg||'未知');
  }
}

function renderItems(items, job){
  const card = document.getElementById('itemsCard');
  const grid = document.getElementById('itemGrid');
  const scanTitle = document.getElementById('scanTitle');
  const bar = document.getElementById('actionBar');
  card.style.display='';
  if(job) scanTitle.innerHTML='<span class="pf-badge pf-'+esc(job.platform)+'">'+esc(job.platform)+'</span> '+job.item_count+' 筆商品';
  if(!items.length){
    grid.innerHTML='<div class="empty">未抓到商品。<br>可能需要手動登入，或頁面格式不支援。<br><br>請用 python local_worker.py --login 先登入，再重試。</div>';
    bar.style.display='none';
    return;
  }
  grid.innerHTML = items.map(it=>`
    <label class="item-card${it.added_to_queue?' checked':''}" onclick="updateCount()">
      <input type="checkbox" value="${it.id}" class="item-cb"${it.added_to_queue?' disabled checked':''}>
      <img src="${esc(it.image)}" onerror="this.style.background='#f0f0f0';this.style.display='block'" alt="">
      <div class="item-info">
        <div class="item-title">${esc(it.title)}</div>
        ${it.price?'<div class="item-price">￥'+esc(it.price)+'</div>':''}
        ${it.shop_name?'<div class="item-shop">'+esc(it.shop_name)+'</div>':''}
        ${it.added_to_queue?'<div class="item-added">已加入</div>':''}
      </div>
    </label>
  `).join('');
  bar.style.display='flex';
  updateCount();
}

function updateCount(){
  const n = document.querySelectorAll('.item-cb:not(:disabled):checked').length;
  document.getElementById('selCount').textContent='已選 '+n+' 筆';
}

function toggleAll(v){
  document.querySelectorAll('.item-cb:not(:disabled)').forEach(el=>{el.checked=v;});
  document.querySelectorAll('.item-card').forEach(el=>{ const cb=el.querySelector('.item-cb'); if(cb&&!cb.disabled) el.classList.toggle('checked',v); });
  updateCount();
}

async function addToQueue(){
  const checked=[...document.querySelectorAll('.item-cb:not(:disabled):checked')].map(el=>parseInt(el.value));
  if(!checked.length){alert('請勾選商品');return;}
  const brand=document.getElementById('brandSel').value;
  document.getElementById('addBtn').disabled=true;
  try{
    const r=await api('/api/store-scan/to-queue',{method:'POST',body:JSON.stringify({item_ids:checked,brand}),headers:{'Content-Type':'application/json'}});
    showToast('已加入 '+r.added+' 筆到商品搬運中心');
    const job=await api('/api/store-scan/'+currentJobId);
    renderItems(job.items||[],job);
  }catch(e){
    alert('錯誤：'+e.message);
  }finally{
    document.getElementById('addBtn').disabled=false;
  }
}

loadHistory();
</script>
</body>
</html>"""


PRODUCTS_HTML = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI 商品搬運中心</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,sans-serif;background:#f0f1f5;color:#333}
/* ── App shell：側邊欄 + 主區域 ───────────────────────── */
.app-shell{display:flex;min-height:100vh}
.sidebar{width:220px;flex-shrink:0;background:#1a1a2e;color:#cfd2e6;display:flex;flex-direction:column;padding:16px 0}
.sidebar-logo{display:flex;align-items:center;gap:10px;padding:6px 18px 18px;border-bottom:1px solid rgba(255,255,255,.08);margin-bottom:10px}
.sidebar-logo .logo-mark{width:32px;height:32px;border-radius:9px;background:#7c5cff;display:flex;align-items:center;justify-content:center;font-weight:800;color:#fff;font-size:14px;flex-shrink:0}
.sidebar-logo .logo-text{line-height:1.25}
.sidebar-logo .logo-text b{display:block;font-size:13px;color:#fff;font-weight:700}
.sidebar-logo .logo-text span{font-size:10px;color:#9b9fc0}
.nav-item{display:flex;align-items:center;gap:10px;padding:10px 18px;font-size:13px;color:#cfd2e6;text-decoration:none;cursor:pointer;border-left:3px solid transparent;white-space:nowrap}
.nav-item:hover{background:rgba(255,255,255,.06);color:#fff}
.nav-item.active{background:rgba(124,92,255,.18);color:#fff;border-left-color:#7c5cff;font-weight:700}
.nav-icon{width:18px;text-align:center;flex-shrink:0}
.sidebar-foot{margin-top:auto;padding:14px 18px;border-top:1px solid rgba(255,255,255,.08);font-size:12px;color:#9b9fc0;display:flex;align-items:center;gap:8px}
.sidebar-foot .avatar{width:26px;height:26px;border-radius:50%;background:#7c5cff;color:#fff;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700}
.main-area{flex:1;min-width:0;display:flex;flex-direction:column}
.topbar{background:#fff;padding:14px 24px;display:flex;align-items:center;gap:14px;border-bottom:1px solid #eceef2}
.topbar-title{font-size:18px;font-weight:800;color:#1a1a2e}
.topbar-help{font-size:12px;color:#9b9fc0;cursor:pointer}
.topbar-help:hover{color:#7c5cff}
.topbar-actions{margin-left:auto;display:flex;gap:8px;position:relative}
.btn-ghost{background:#f3f1ff;color:#5b3df0;border:none;border-radius:9px;padding:8px 16px;font-size:13px;font-weight:700;cursor:pointer;font-family:-apple-system,sans-serif;text-decoration:none;display:inline-flex;align-items:center}
.btn-ghost:hover{background:#e8e3ff}
.btn-dark{background:#1a1a2e;color:#fff;border:none;border-radius:9px;padding:8px 16px;font-size:13px;font-weight:700;cursor:pointer;font-family:-apple-system,sans-serif}
.btn-dark:hover{background:#2d2d44}
.export-flyout{position:absolute;top:42px;right:0;background:#fff;border-radius:10px;box-shadow:0 8px 24px rgba(0,0,0,.15);padding:8px;display:none;flex-direction:column;gap:4px;min-width:140px;z-index:50}
.export-flyout.show{display:flex}
.export-flyout a{color:#333;text-decoration:none;font-size:13px;padding:7px 10px;border-radius:7px}
.export-flyout a:hover{background:#f3f1ff}
.content-wrap{padding:20px 24px;flex:1;min-width:0}
.export-btn{background:#2d2d2d;color:#fff;text-decoration:none;border-radius:8px;padding:5px 12px;font-size:12px;font-weight:700;white-space:nowrap;display:inline-block}
.export-btn:hover{background:#444}
/* 統計卡 */
.stat-row{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin-bottom:20px}
.stat-card{background:#fff;border-radius:14px;padding:14px 16px;box-shadow:0 1px 4px rgba(0,0,0,.06);display:flex;flex-direction:column;gap:4px}
.stat-num{font-size:22px;font-weight:800;color:#1a1a2e}
.stat-label{font-size:12px;color:#9b9fc0;font-weight:600}
/* 兩欄：列表 + 詳情 */
.content-grid{display:grid;grid-template-columns:420px 1fr;gap:18px;align-items:start}
@media (max-width:1100px){.content-grid{grid-template-columns:1fr}}
.list-col{min-width:0}
.detail-col{min-width:0;background:#fff;border-radius:16px;box-shadow:0 1px 4px rgba(0,0,0,.06);overflow:hidden;position:sticky;top:20px}
.detail-empty{padding:80px 20px;text-align:center;color:#bbb;font-size:14px}
.detail-hd{padding:18px 22px;border-bottom:1px solid #f0f0f0;display:flex;align-items:center;gap:12px}
.detail-hd-title{font-size:16px;font-weight:800;flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.detail-hd-meta{font-size:12px;color:#9b9fc0;margin-top:3px}
.detail-actions{display:flex;gap:6px}
.icon-btn{background:#f5f5f8;border:none;border-radius:8px;padding:6px 10px;font-size:12px;cursor:pointer;color:#555;font-family:-apple-system,sans-serif}
.icon-btn:hover{background:#ece9ff;color:#5b3df0}
.icon-btn.danger:hover{background:#fdecea;color:#e53935}
/* input card */
.card{background:#fff;border-radius:14px;padding:20px;margin-bottom:16px;box-shadow:0 1px 4px rgba(0,0,0,.08)}
.card h3{font-size:14px;font-weight:700;color:#666;margin-bottom:12px;letter-spacing:.3px}
.url-row{display:flex;gap:10px}
.url-row input{flex:1;border:1.5px solid #ddd;border-radius:10px;padding:10px 14px;font-size:14px;outline:none;font-family:-apple-system,sans-serif}
.url-row input:focus{border-color:#1a1a1a}
.btn-primary{background:#7c5cff;color:#fff;border:none;border-radius:10px;padding:10px 22px;font-size:14px;font-weight:700;cursor:pointer;white-space:nowrap;font-family:-apple-system,sans-serif;width:100%}
.btn-primary:hover{background:#6a4ce8}
.btn-primary:disabled{background:#c9c2f5;cursor:default}
.err-msg{color:#c62828;font-size:13px;margin-top:8px;display:none}
/* filter */
.filter-row{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap}
.fbtn{border:1.5px solid #ddd;background:#fff;border-radius:20px;padding:5px 14px;font-size:13px;cursor:pointer;font-family:-apple-system,sans-serif}
.fbtn.active{background:#1a1a2e;color:#fff;border-color:#1a1a2e}
/* job cards */
.job{background:#fff;border-radius:12px;padding:13px 14px;margin-bottom:9px;box-shadow:0 1px 3px rgba(0,0,0,.07);cursor:pointer;border:1.5px solid transparent;transition:all .15s;display:flex;gap:10px}
.job:hover{border-color:#e0e0e0;box-shadow:0 3px 10px rgba(0,0,0,.1)}
.job.active{border-color:#7c5cff;box-shadow:0 0 0 1px #7c5cff}
.job-thumb{width:56px;height:56px;border-radius:9px;background:#f3f1ff;flex-shrink:0;object-fit:cover;display:flex;align-items:center;justify-content:center;color:#c4bdf5;font-size:20px}
.job-main{flex:1;min-width:0}
.job-top{display:flex;align-items:center;gap:6px}
.badge{font-size:11px;font-weight:700;padding:3px 8px;border-radius:8px;flex-shrink:0}
.pf-1688{background:#fff0f0;color:#c62828}
.pf-taobao{background:#fff4e5;color:#e65100}
.st-pending{background:#f5f5f5;color:#999}
.st-scraping{background:#e3f2fd;color:#1565c0}
.st-rewriting{background:#fff8e1;color:#f57f17}
.st-done{background:#e8f5e9;color:#2e7d32}
.st-error{background:#fdecea;color:#c62828}
.job-title{flex:1;font-size:13px;font-weight:700;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.job-url{font-size:11px;color:#bbb;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;margin-top:3px}
.job-foot{display:flex;align-items:center;gap:8px;margin-top:4px}
.job-time{font-size:11px;color:#ccc}
.job-imgcount{font-size:11px;color:#9b9fc0;background:#f5f5f8;border-radius:6px;padding:1px 6px}
.del-btn{background:none;border:none;color:#ddd;cursor:pointer;font-size:18px;line-height:1;padding:2px 4px;border-radius:6px;flex-shrink:0;align-self:flex-start}
.del-btn:hover{color:#e53935;background:#fdecea}
.empty{text-align:center;padding:50px 20px;color:#ccc;font-size:15px}
.spinner{display:inline-block;width:12px;height:12px;border:2px solid #ddd;border-top-color:#666;border-radius:50%;animation:spin .8s linear infinite;vertical-align:middle;margin-right:4px}
@keyframes spin{to{transform:rotate(360deg)}}
.detail-body{padding:20px}
.section{margin-bottom:20px}
.slabel{font-size:11px;font-weight:700;color:#aaa;letter-spacing:.5px;text-transform:uppercase;margin-bottom:6px}
.rbox{background:#f8f8f8;border-radius:10px;padding:14px;font-size:14px;line-height:1.7;white-space:pre-wrap;word-break:break-all;position:relative;padding-right:70px}
.copy-btn{position:absolute;top:8px;right:8px;background:#1a1a1a;color:#fff;border:none;border-radius:8px;padding:4px 10px;font-size:12px;cursor:pointer;opacity:.7;font-family:-apple-system,sans-serif}
.copy-btn:hover{opacity:1}
.copy-btn.ok{background:#2e7d32}
.kw-row{display:flex;gap:6px;flex-wrap:wrap}
.kw{background:#e3f2fd;color:#1565c0;border-radius:12px;padding:4px 12px;font-size:13px;font-weight:600}
.imgs-row{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px}
.imgs-row img{width:80px;height:80px;object-fit:cover;border-radius:8px;border:1px solid #eee}
.divider{border:none;border-top:1px solid #f0f0f0;margin:16px 0}
.err-box{background:#fdecea;color:#c62828;border-radius:10px;padding:12px 14px;font-size:13px;margin-bottom:14px}
.raw-toggle{background:none;border:none;color:#bbb;font-size:12px;cursor:pointer;font-family:-apple-system,sans-serif;text-decoration:underline;padding:0}
/* 手動補圖 */
.img-form{margin-top:14px}
.img-form textarea{width:100%;border:1.5px solid #ddd;border-radius:10px;padding:10px;font-size:13px;height:90px;resize:vertical;font-family:-apple-system,sans-serif;outline:none}
.img-form textarea:focus{border-color:#1a1a1a}
/* 選圖 UI */
.img-zone-hd{display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-bottom:6px;padding:6px 10px;background:#fafafa;border-radius:8px;border:1px solid #f0f0f0}
.img-zone-hd .slabel{margin-bottom:0;flex:none}
.sel-btn{background:#f0f0f0;color:#555;border:none;border-radius:6px;padding:3px 9px;font-size:11px;cursor:pointer;font-family:-apple-system,sans-serif}
.sel-btn:hover{background:#e0e0e0}
.img-grid{display:flex;gap:8px;flex-wrap:wrap;margin-top:6px}
.img-thumb{position:relative;cursor:pointer;flex-shrink:0;width:128px;text-align:center}
.img-thumb input[type=checkbox]{position:absolute;top:4px;left:4px;width:16px;height:16px;cursor:pointer;z-index:2;accent-color:#1a73e8}
.img-thumb img{width:120px;height:120px;object-fit:cover;border-radius:6px;border:2px solid #eee;transition:border-color .15s;display:block;margin:0 auto}
.img-thumb.checked img{border-color:#1a73e8;box-shadow:0 0 0 1px #1a73e8}
/* category colours */
.img-thumb.cat-sku input[type=checkbox]{accent-color:#f57c00}
.img-thumb.cat-sku.checked img{border-color:#ff9800;box-shadow:0 0 0 1px #ff9800}
.img-thumb.cat-detail input[type=checkbox]{accent-color:#7b1fa2}
.img-thumb.cat-review input[type=checkbox]{accent-color:#795548}
.img-thumb.cat-review.checked img{border-color:#795548;box-shadow:0 0 0 1px #795548}
.img-thumb.cat-detail.checked img{border-color:#9c27b0;box-shadow:0 0 0 1px #9c27b0}
.img-size{font-size:9px;color:#999;text-align:center;margin:3px 0 1px;min-height:13px;line-height:1.3}
.img-label{font-size:9px;color:#666;text-align:center;max-width:120px;overflow:hidden;white-space:nowrap;text-overflow:ellipsis;margin-bottom:2px}
.thumb-actions{display:flex;justify-content:center;gap:3px;margin-top:3px}
.cat-collapse-btn{margin-left:auto!important;background:#f5f5f5;color:#888;font-size:10px}
.cat-collapse-btn:hover{background:#e0e0e0;color:#333}
.thumb-act{background:#f0f0f0;border:none;border-radius:4px;padding:2px 6px;font-size:11px;cursor:pointer;color:#555;text-decoration:none;display:inline-block;line-height:1.5}
.thumb-act:hover{background:#ddd;color:#333}
.sel-action-bar{position:sticky;bottom:0;background:#fff;border-top:1px solid #f0f0f0;padding:12px 0 4px;display:flex;gap:8px;align-items:center;margin-top:12px}
.sel-count{font-size:12px;color:#888;flex:1}
.btn-confirm{background:#1a1a1a;color:#fff;border:none;border-radius:9px;padding:8px 18px;font-size:13px;font-weight:700;cursor:pointer;font-family:-apple-system,sans-serif}
.btn-confirm:hover{background:#333}
.btn-zip{background:#e3f2fd;color:#1565c0;border:none;border-radius:9px;padding:8px 14px;font-size:13px;font-weight:600;cursor:pointer;font-family:-apple-system,sans-serif}
.btn-zip:hover{background:#bbdefb}
.btn-translate{background:#e8f5e9;color:#2e7d32;border:none;border-radius:9px;padding:8px 14px;font-size:13px;font-weight:600;cursor:pointer;font-family:-apple-system,sans-serif}
.btn-translate:hover{background:#c8e6c9}
.btn-translate:disabled{background:#f5f5f5;color:#bbb;cursor:default}
/* Lightbox */
.lb-overlay{position:fixed;inset:0;background:rgba(0,0,0,.88);z-index:9999;display:flex;align-items:center;justify-content:center}
.lb-content{max-width:90vw;max-height:90vh;display:flex;flex-direction:column;align-items:center}
.lb-content img{max-width:88vw;max-height:80vh;object-fit:contain;border-radius:8px}
.lb-info{color:#fff;font-size:13px;margin-top:10px;display:flex;gap:16px;align-items:center}
.lb-label{font-weight:600;max-width:300px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.lb-count{color:#aaa;font-size:12px}
.lb-nav{position:fixed;top:50%;transform:translateY(-50%);background:rgba(255,255,255,.15);border:none;color:#fff;font-size:40px;cursor:pointer;padding:10px 18px;border-radius:8px;line-height:1;z-index:10000}
.lb-nav:hover{background:rgba(255,255,255,.3)}
.lb-prev{left:12px}
.lb-next{right:12px}
.lb-close{position:fixed;top:14px;right:18px;background:none;border:none;color:#fff;font-size:28px;cursor:pointer;z-index:10000;line-height:1}
.lb-close:hover{color:#ddd}
.sel-btn-green{background:#e8f5e9;color:#2e7d32}
.sel-btn-green:hover{background:#c8e6c9}
.sel-btn-white{background:#f3e5f5;color:#6a1b9a}
.sel-btn-white:hover{background:#e1bee7}
.upload-tr-sec{background:#fff;border-radius:12px;padding:14px 16px;margin-top:10px;border:1.5px dashed #b39ddb}
.upload-tr-sec .slabel{font-size:12px;font-weight:700;color:#6a1b9a;margin-bottom:10px}
.btn-upload-lbl{display:inline-block;background:#ede7f6;color:#4527a0;border-radius:8px;padding:6px 14px;font-size:13px;font-weight:600;cursor:pointer;border:1.5px solid #b39ddb}
.btn-upload-lbl:hover{background:#d1c4e9}
.tr-type-sec{margin-top:10px;padding:10px 0 4px;border-top:1px solid #f0f0f0}
.tr-type-label{font-size:12px;font-weight:700;color:#555;margin-bottom:6px}
.translated-sec{background:#f0fdf4;border:1px solid #86efac;border-radius:10px;padding:12px;margin-top:10px}
.translated-sec .slabel{color:#16a34a}
.btn-sm{background:#333;color:#fff;border:none;border-radius:8px;padding:6px 14px;font-size:13px;cursor:pointer;font-family:-apple-system,sans-serif;margin-top:6px}
.btn-sm:hover{background:#555}
.hidden{display:none}
.toast{position:fixed;bottom:28px;left:50%;transform:translateX(-50%);background:#333;color:#fff;padding:9px 20px;border-radius:20px;font-size:14px;z-index:999;opacity:0;transition:opacity .3s;pointer-events:none}
.toast.show{opacity:1}
/* 編輯模式 */
.edit-ta{width:100%;border:1.5px solid #ddd;border-radius:10px;padding:10px 12px;font-size:14px;line-height:1.6;resize:vertical;font-family:-apple-system,sans-serif;outline:none;background:#fff}
.edit-ta:focus{border-color:#1a1a1a}
.edit-ta.name{height:52px}
.edit-ta.desc{height:200px}
.edit-ta.kw{height:52px}
.edit-bar{display:flex;gap:8px;margin-top:10px}
.btn-save{background:#1a1a1a;color:#fff;border:none;border-radius:9px;padding:8px 18px;font-size:13px;font-weight:700;cursor:pointer;font-family:-apple-system,sans-serif}
.btn-save:hover{background:#333}
.btn-cancel{background:#f5f5f5;color:#555;border:none;border-radius:9px;padding:8px 16px;font-size:13px;cursor:pointer;font-family:-apple-system,sans-serif}
.btn-cancel:hover{background:#eee}
.btn-edit{background:#f0f0f0;color:#333;border:none;border-radius:9px;padding:6px 14px;font-size:13px;cursor:pointer;font-family:-apple-system,sans-serif;margin-left:auto}
.btn-edit:hover{background:#e0e0e0}
/* 批次輸入 */
.url-ta{width:100%;border:1.5px solid #ddd;border-radius:10px;padding:10px 14px;font-size:13px;height:90px;resize:vertical;font-family:-apple-system,sans-serif;outline:none}
.url-ta:focus{border-color:#1a1a1a}
.brand-sel{border:1.5px solid #ddd;border-radius:10px;padding:8px 12px;font-size:13px;font-family:-apple-system,sans-serif;outline:none;background:#fff;color:#333;cursor:pointer;width:100%}
.brand-sel:focus{border-color:#1a1a1a}
.brand-badge{font-size:10px;font-weight:700;padding:2px 7px;border-radius:7px;background:#f3e5f5;color:#7b1fa2;flex-shrink:0}
.batch-hint{font-size:12px;color:#aaa;margin-top:6px}
.progress-bar-wrap{background:#f0f0f0;border-radius:10px;height:6px;margin-top:10px;overflow:hidden;display:none}
.progress-bar{background:#1a1a1a;height:100%;border-radius:10px;transition:width .3s}
/* Tabs */
.tab-nav{display:flex;gap:2px;padding:0 20px;border-bottom:1px solid #f0f0f0;overflow-x:auto;background:#fff;position:sticky;top:57px;z-index:1}
.tab-btn{background:none;border:none;padding:12px 14px;font-size:13px;font-weight:600;color:#999;cursor:pointer;white-space:nowrap;border-bottom:2px solid transparent;font-family:-apple-system,sans-serif}
.tab-btn:hover{color:#333}
.tab-btn.active{color:#1a1a1a;border-bottom-color:#1a1a1a}
.tab-panel{display:none}
.tab-panel.active{display:block}
/* 左右對照 */
.compare-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}
@media (max-width:560px){.compare-grid{grid-template-columns:1fr}}
.compare-col-hd{font-size:12px;font-weight:700;color:#999;letter-spacing:.4px;text-transform:uppercase;margin-bottom:10px;padding-bottom:8px;border-bottom:1px solid #f0f0f0}
.compare-field{margin-bottom:14px}
.compare-field label{display:block;font-size:11px;color:#aaa;margin-bottom:4px}
.compare-field .static-val{background:#f8f8f8;border-radius:8px;padding:8px 10px;font-size:13px;line-height:1.5;word-break:break-all;min-height:18px}
.compare-field input[type=text],.compare-field select{width:100%;border:1.5px solid #ddd;border-radius:8px;padding:7px 10px;font-size:13px;font-family:-apple-system,sans-serif;outline:none}
.compare-field input:focus,.compare-field select:focus{border-color:#1a1a1a}
.status-pill{display:inline-block;font-size:11px;font-weight:700;padding:3px 10px;border-radius:10px;background:#fff8e1;color:#f57f17}
/* 文案欄位 */
.copy-field{margin-bottom:18px}
.copy-field-hd{display:flex;align-items:center;justify-content:space-between;margin-bottom:6px}
.copy-field-hd .slabel{margin-bottom:0}
.copy-field textarea,.copy-field input[type=text]{width:100%;border:1.5px solid #ddd;border-radius:10px;padding:9px 12px;font-size:13px;line-height:1.6;font-family:-apple-system,sans-serif;outline:none;resize:vertical}
.copy-field textarea:focus,.copy-field input:focus{border-color:#1a1a1a}
.btn-regen{background:#fff3e0;color:#e65100;border:none;border-radius:9px;padding:8px 16px;font-size:13px;font-weight:600;cursor:pointer;font-family:-apple-system,sans-serif}
.btn-regen:hover{background:#ffe0b2}
.btn-regen:disabled{background:#f5f5f5;color:#bbb;cursor:default}
/* FAQ */
.faq-item{background:#f8f8f8;border-radius:10px;padding:12px;margin-bottom:10px;position:relative}
.faq-item input[type=text]{width:100%;border:1.5px solid #ddd;border-radius:8px;padding:7px 10px;font-size:13px;margin-bottom:6px;font-family:-apple-system,sans-serif;outline:none;background:#fff}
.faq-item textarea{width:100%;border:1.5px solid #ddd;border-radius:8px;padding:7px 10px;font-size:13px;font-family:-apple-system,sans-serif;outline:none;background:#fff;resize:vertical;height:50px}
.faq-del{position:absolute;top:8px;right:8px;background:none;border:none;color:#bbb;font-size:16px;cursor:pointer;line-height:1}
.faq-del:hover{color:#e53935}
.btn-add-faq{background:#f0f0f0;color:#555;border:none;border-radius:9px;padding:7px 14px;font-size:13px;cursor:pointer;font-family:-apple-system,sans-serif}
.btn-add-faq:hover{background:#e0e0e0}
.main-badge{position:absolute;bottom:4px;left:4px;background:#1a1a1a;color:#fff;font-size:9px;font-weight:700;padding:1px 6px;border-radius:5px;z-index:2}
</style>
</head>
<body>

<div class="app-shell">
  <div class="sidebar">
    <div class="sidebar-logo">
      <div class="logo-mark">J</div>
      <div class="logo-text"><b>J.SIMPLE</b><span>AI 商品工廠</span></div>
    </div>
    <div class="nav-item" data-nav="dashboard" onclick="navStub('總覽 Dashboard')"><span class="nav-icon">📊</span>總覽 Dashboard</div>
    <div class="nav-item active" data-nav="products"><span class="nav-icon">🚚</span>AI 商品搬運中心</div>
    <div class="nav-item" data-nav="manage" onclick="navStub('商品管理')"><span class="nav-icon">📦</span>商品管理</div>
    <div class="nav-item" data-nav="batch" onclick="navStub('批次搬運')"><span class="nav-icon">🗂</span>批次搬運</div>
    <a class="nav-item" href="/admin/brand-settings?key={{ key }}"><span class="nav-icon">🏷</span>品牌設定</a>
    <div class="nav-item" data-nav="tpl" onclick="navStub('文案模板')"><span class="nav-icon">📝</span>文案模板</div>
    <div class="nav-item" data-nav="export-log" onclick="navStub('出口記錄')"><span class="nav-icon">📤</span>出口記錄</div>
    <div class="sidebar-foot"><div class="avatar">J</div>JSIMPLE Admin</div>
  </div>

  <div class="main-area">
    <div class="topbar">
      <div class="topbar-title">AI 商品搬運中心</div>
      <div class="topbar-help" onclick="toast('貼上 1688 / 淘寶商品連結，AI 會自動爬取並改寫文案，完成後可在右側挑圖、編輯文案、確認上架資料。')">ⓘ 如何使用</div>
      <div class="topbar-actions">
        <a class="btn-ghost" href="/admin/products/store-scan?key={{ key }}">店鋪選品</a>
        <button class="btn-ghost" onclick="toggleExportFlyout()">匯出記錄</button>
        <div class="export-flyout" id="exportFlyout">
          <a href="/api/products/export?format=xlsx&key={{ key }}">⬇ 匯出 Excel</a>
          <a href="/api/products/export?format=csv&key={{ key }}">⬇ 匯出 CSV</a>
        </div>
        <button class="btn-dark" onclick="document.getElementById('urlInput').scrollIntoView({behavior:'smooth'});document.getElementById('urlInput').focus()">＋ 新增搬運任務</button>
      </div>
    </div>

    <div class="content-wrap">
      <div class="stat-row" id="statRow">
        <div class="stat-card"><div class="stat-num" id="stat_today">0</div><div class="stat-label">今日處理</div></div>
        <div class="stat-card"><div class="stat-num" id="stat_pickimg">0</div><div class="stat-label">待挑圖</div></div>
        <div class="stat-card"><div class="stat-num" id="stat_review">0</div><div class="stat-label">待審核</div></div>
        <div class="stat-card"><div class="stat-num" id="stat_publish">0</div><div class="stat-label">待上架</div></div>
        <div class="stat-card"><div class="stat-num" id="stat_done">0</div><div class="stat-label">已完成</div></div>
        <div class="stat-card"><div class="stat-num" id="stat_failed">0</div><div class="stat-label">失敗</div></div>
      </div>

      <div class="content-grid">
        <div class="list-col">
          <div class="card">
            <h3>新增商品任務</h3>
            <textarea class="url-ta" id="urlInput" placeholder="https://detail.1688.com/offer/xxx.html&#10;https://item.taobao.com/item.htm?id=xxx&#10;（一行一個連結）"></textarea>
            <div style="margin-top:8px">
              <select id="brandSel" class="brand-sel">
                <option value="">不指定品牌（通用風格）</option>
                <option value="jsimple">JSIMPLE — 高架床 / 家具</option>
                <option value="lander">朗德燈具 — 燈具 / 照明</option>
                <option value="filterbreath">濾呼吸 — 空氣濾網</option>
                <option value="chengguang">澄光窗簾 — 窗簾 / 遮光簾</option>
              </select>
            </div>
            <div class="batch-hint" id="batchHint"></div>
            <div class="progress-bar-wrap" id="progressWrap"><div class="progress-bar" id="progressBar" style="width:0%"></div></div>
            <div style="margin-top:10px;display:flex;gap:10px;align-items:center">
              <button class="btn-primary" id="addBtn" onclick="submitUrl()">開始搬運</button>
            </div>
            <div class="err-msg" id="addErr"></div>
          </div>

          <div class="filter-row">
            <button class="fbtn active" onclick="setFilter('all',this)">全部</button>
            <button class="fbtn" onclick="setFilter('done',this)">完成</button>
            <button class="fbtn" onclick="setFilter('error',this)">失敗</button>
            <button class="fbtn" onclick="setFilter('processing',this)">進行中</button>
          </div>

          <div id="jobList"></div>
          <div id="emptyMsg" class="empty hidden">還沒有任務，貼上連結開始搬運</div>
        </div>

        <div class="detail-col" id="detailCol">
          <div class="detail-empty" id="detailEmpty">👈 點選左側商品查看詳情</div>
        </div>
      </div>
    </div>
  </div>
</div>

<div class="toast" id="toast"></div>

<script>
const KEY = "{{ key }}";
let jobs = [], filter = "all", pollTimer = null, openId = null, _selImgs = new Set(), _lbImgs = [], _lbIdx = 0, _curJobId = 0, _curMainImage = "", _curTab = "tab1";

const api = (url, opts={}) => fetch(url+(url.includes("?")?"&":"?")+"key="+KEY, {headers:{"Content-Type":"application/json"},...opts}).then(r=>r.json());

async function loadJobs(){
  try{
    const d = await api("/api/products");
    jobs = d.jobs||[];
    render();
    schedulePoll();
  }catch(e){console.error(e)}
}

function schedulePoll(){
  clearTimeout(pollTimer);
  const active = jobs.some(j=>["pending","scraping","rewriting"].includes(j.status));
  pollTimer = setTimeout(loadJobs, active ? 2500 : 8000);
}

function updateStats(){
  const todayStr = new Date().toDateString();
  const today = jobs.filter(j=>j.created_at && new Date(j.created_at*1000).toDateString()===todayStr).length;
  const pickImg = jobs.filter(j=>j.status==="done" && (!j.img_status || j.img_status==="pending_images")).length;
  const review = jobs.filter(j=>j.listing_status==="待審核").length;
  const publish = jobs.filter(j=>j.listing_status==="待上架").length;
  const done = jobs.filter(j=>j.status==="done").length;
  const failed = jobs.filter(j=>j.status==="error").length;
  document.getElementById("stat_today").textContent = today;
  document.getElementById("stat_pickimg").textContent = pickImg;
  document.getElementById("stat_review").textContent = review;
  document.getElementById("stat_publish").textContent = publish;
  document.getElementById("stat_done").textContent = done;
  document.getElementById("stat_failed").textContent = failed;
}

function render(){
  const list = document.getElementById("jobList");
  const empty = document.getElementById("emptyMsg");
  updateStats();
  const filtered = jobs.filter(j=>{
    if(filter==="all") return true;
    if(filter==="processing") return ["pending","scraping","rewriting"].includes(j.status);
    return j.status===filter;
  });
  if(!filtered.length){ list.innerHTML=""; empty.classList.remove("hidden"); return; }
  empty.classList.add("hidden");
  list.innerHTML = filtered.map(jobCard).join("");
}

function jobCard(j){
  const pfLabel = {1688:"1688",taobao:"淘寶"}[j.platform]||j.platform;
  const stLabel = {pending:"等待 Worker",scraping:"爬取中",rewriting:"改寫中",done:"完成",error:"失敗"}[j.status]||j.status;
  const isActive = ["pending","scraping","rewriting"].includes(j.status);
  const spin = isActive ? '<span class="spinner"></span>' : "";
  const brandLabel = {jsimple:"JSIMPLE",lander:"朗德",filterbreath:"濾呼吸",chengguang:"澄光"}[j.brand]||"";
  const brandBadge = brandLabel ? `<span class="brand-badge">${brandLabel}</span>` : "";
  const title = esc(j.ai_name||j.raw_title||j.url);
  const urlShort = j.url.length>50 ? j.url.slice(0,50)+"…" : j.url;
  const ts = j.created_at ? new Date(j.created_at*1000).toLocaleString("zh-TW",{month:"2-digit",day:"2-digit",hour:"2-digit",minute:"2-digit"}) : "";
  const thumbSrc = j.main_image || (j.raw_images&&j.raw_images[0]) || "";
  const thumb = thumbSrc ? `<img class="job-thumb" src="${esc(thumbSrc)}" loading="lazy" onerror="this.replaceWith(Object.assign(document.createElement('div'),{className:'job-thumb',textContent:'🖼'}))">` : `<div class="job-thumb">🖼</div>`;
  const imgCount = (j.raw_images||[]).length;
  return `<div class="job${_curJobId===j.id?' active':''}" onclick="openJob(${j.id})">
    ${thumb}
    <div class="job-main">
      <div class="job-top">
        <span class="badge pf-${j.platform}">${pfLabel}</span>
        <span class="badge st-${j.status}">${spin}${stLabel}</span>
        ${brandBadge}
        <button class="del-btn" onclick="delJob(event,${j.id})" title="刪除">×</button>
      </div>
      <div class="job-title">${title}</div>
      <div class="job-url">${esc(urlShort)}</div>
      <div class="job-foot">
        ${ts?`<span class="job-time">${ts}</span>`:""}
        ${imgCount?`<span class="job-imgcount">已選 ${imgCount} 張</span>`:""}
      </div>
    </div>
  </div>`;
}

function setFilter(f,btn){
  filter=f;
  document.querySelectorAll(".fbtn").forEach(b=>b.classList.remove("active"));
  btn.classList.add("active");
  render();
}

async function submitUrl(){
  const ta=document.getElementById("urlInput");
  const errEl=document.getElementById("addErr");
  const btn=document.getElementById("addBtn");
  const hint=document.getElementById("batchHint");
  const pWrap=document.getElementById("progressWrap");
  const pBar=document.getElementById("progressBar");
  errEl.style.display="none";
  const urls=ta.value.split("\\n").map(s=>s.trim()).filter(Boolean);
  if(!urls.length){toast("請貼上商品連結");errEl.textContent="請貼上商品連結";errEl.style.display="block";return;}
  btn.disabled=true;
  let done=0,failed=0;
  pWrap.style.display="block";
  for(const url of urls){
    btn.textContent=`搬運中 ${done+failed+1}/${urls.length}...`;
    hint.textContent=`已提交 ${done} 筆${failed?`，失敗 ${failed} 筆`:""}`;
    pBar.style.width=((done+failed)/urls.length*100)+"%";
    try{
      const brand=document.getElementById("brandSel").value;
      const r=await api("/api/products",{method:"POST",body:JSON.stringify({url,brand})});
      if(r.error){failed++;toast(`連結格式錯誤：${r.error}`);}
      else done++;
    }catch(e){failed++;toast("網路錯誤，請稍後再試");}
  }
  pBar.style.width="100%";
  hint.textContent=`完成：${done} 筆${failed?`，失敗：${failed} 筆`:""}`;
  if(done>0){ta.value="";toast(`成功提交 ${done} 筆任務`);}
  btn.disabled=false; btn.textContent="開始搬運";
  loadJobs();
  setTimeout(()=>{pWrap.style.display="none"; hint.textContent="";},4000);
}

function navStub(name){
  toast(name+" 敬請期待");
}
function toggleExportFlyout(){
  document.getElementById("exportFlyout").classList.toggle("show");
}
document.addEventListener("click", e=>{
  const fly=document.getElementById("exportFlyout");
  if(fly && !fly.contains(e.target) && !e.target.closest("[onclick='toggleExportFlyout()']")) fly.classList.remove("show");
});

async function openJob(id){
  openId=id;
  document.querySelectorAll(".job").forEach(el=>el.classList.remove("active"));
  const card=[...document.querySelectorAll(".job")].find(el=>el.getAttribute("onclick")===`openJob(${id})`);
  if(card) card.classList.add("active");
  const detail=document.getElementById("detailCol");
  detail.innerHTML='<div style="text-align:center;padding:60px"><span class="spinner" style="width:18px;height:18px"></span></div>';
  try{
    const j=await api("/api/products/"+id);
    renderDetail(j);
  }catch(e){detail.innerHTML='<div class="err-box" style="margin:20px">載入失敗</div>';}
}

const BRAND_LABELS = {jsimple:"JSIMPLE",lander:"朗德燈具",filterbreath:"濾呼吸",chengguang:"澄光窗簾"};
let _curJob = null;

function detailHd(j){
  const stLabel = {pending:"等待 Worker",scraping:"爬取中",rewriting:"改寫中",done:"完成",error:"失敗"}[j.status]||j.status;
  return `<div class="detail-hd">
    <div style="min-width:0;flex:1">
      <div class="detail-hd-title">${esc(j.ai_name||j.raw_title||"商品詳情")}</div>
      <div class="detail-hd-meta"><span class="badge st-${j.status}">${stLabel}</span> &nbsp;<a href="${esc(j.url)}" target="_blank" style="color:#9b9fc0">來源連結 ↗</a></div>
    </div>
    <div class="detail-actions">
      <button class="icon-btn" onclick="openJob(${j.id})" title="重新整理">↻</button>
      <button class="icon-btn danger" onclick="delJob(event,${j.id})" title="刪除">🗑</button>
    </div>
  </div>`;
}

function renderDetail(j){
  if(j.status==="error"){
    let h=detailHd(j)+`<div class="detail-body"><div class="err-box">${esc(j.error_msg||"未知錯誤")}</div>`;
    if(j.raw_title) h+=`<div class="section"><div class="slabel">原始標題</div><div class="rbox">${esc(j.raw_title)}</div></div>`;
    h+=`</div>`;
    document.getElementById("detailCol").innerHTML=h; return;
  }
  if(["pending","scraping","rewriting"].includes(j.status)){
    const msg={pending:"等待爬取...",scraping:"正在爬取商品資料...",rewriting:"AI 正在改寫文案，請稍候..."};
    document.getElementById("detailCol").innerHTML = detailHd(j) + `<div style="text-align:center;padding:30px;color:#888"><span class="spinner" style="width:16px;height:16px;border-top-color:#555"></span> ${msg[j.status]}</div>`;
    if(openId===j.id) setTimeout(()=>{if(openId===j.id)openJob(j.id);},2000);
    return;
  }
  _curJob = j;
  _curJobId = j.id;
  _curMainImage = j.main_image || "";
  const tabNav = `<div class="tab-nav">
    <button class="tab-btn" data-tab="tab1" onclick="switchModalTab('tab1')">商品資訊</button>
    <button class="tab-btn" data-tab="tab2" onclick="switchModalTab('tab2')">圖片中心</button>
    <button class="tab-btn" data-tab="tab3" onclick="switchModalTab('tab3')">AI 商品分析</button>
    <button class="tab-btn" data-tab="tab4" onclick="switchModalTab('tab4')">SEO / FAQ</button>
    <button class="tab-btn" data-tab="tab5" onclick="switchModalTab('tab5')">上架資料</button>
  </div>`;
  const body = `<div class="detail-body">
    <div class="tab-panel" id="panel_tab1">${tab1Html(j)}</div>
    <div class="tab-panel" id="panel_tab2">${tab2Html(j)}</div>
    <div class="tab-panel" id="panel_tab3">${tab3Html(j)}</div>
    <div class="tab-panel" id="panel_tab4">${tab4Html(j)}</div>
    <div class="tab-panel" id="panel_tab5">${tab5Html(j)}</div>
  </div>`;
  document.getElementById("detailCol").innerHTML = detailHd(j) + tabNav + body;
  switchModalTab(_curTab);
}

function switchModalTab(name){
  _curTab = name;
  document.querySelectorAll(".tab-btn").forEach(b=>b.classList.toggle("active", b.dataset.tab===name));
  document.querySelectorAll(".tab-panel").forEach(p=>p.classList.toggle("active", p.id==="panel_"+name));
}

function tab1Html(j){
  const specs = (j.raw_extra && j.raw_extra.specs) ? j.raw_extra.specs : null;
  const specsHtml = specs ? (Array.isArray(specs) ? specs.map(s=>esc(typeof s==='object'?(s.name||s.k||'')+'：'+(s.value||s.v||''):String(s))).join('<br>') : Object.entries(specs).map(([k,v])=>`${esc(k)}：${esc(String(v))}`).join('<br>')) : '（無規格資料）';
  const skuPrices = (j.raw_extra && j.raw_extra.sku_prices) || [];
  const skuPricesHtml = skuPrices.length ? skuPrices.map(p=>`${esc(p.label||'')}：¥${esc(String(p.price||''))}`).join('<br>') : '（無多規格價格資料）';
  const brandOpts = Object.entries(BRAND_LABELS).map(([k,v])=>`<option value="${k}"${j.brand===k?" selected":""}>${v}</option>`).join("");
  return `<div class="compare-grid">
    <div>
      <div class="compare-col-hd">來源商品</div>
      <div class="compare-field"><label>原始標題</label><div class="static-val">${esc(j.raw_title||"（無）")}</div></div>
      <div class="compare-field"><label>原始價格</label><div class="static-val">${esc(j.raw_price||"（無）")}</div></div>
      <div class="compare-field"><label>原始規格</label><div class="static-val">${specsHtml}</div></div>
      <div class="compare-field"><label>來源網址</label><div class="static-val" style="word-break:break-all"><a href="${esc(j.url)}" target="_blank" style="color:#1a73e8">${esc(j.url)}</a></div></div>
    </div>
    <div>
      <div class="compare-col-hd">台灣版商品</div>
      <div class="compare-field"><label>商品名稱</label><input type="text" id="t1_name" value="${esc(j.ai_name||"")}"></div>
      <div class="compare-field"><label>品牌</label><select id="t1_brand">${brandOpts}</select></div>
      <div class="compare-field"><label>建議售價</label><div style="display:flex;gap:8px"><input type="text" id="t1_price_min" placeholder="下限" value="${esc(j.price_min||"")}"><input type="text" id="t1_price_max" placeholder="上限" value="${esc(j.price_max||"")}"></div>
        <div style="display:flex;gap:6px;margin-top:6px;align-items:center">
          <input type="text" id="t1_mult" placeholder="乘數，例如 4.3" style="width:110px;border:1.5px solid #ddd;border-radius:8px;padding:7px 10px;font-size:13px">
          <button type="button" class="sel-btn" onclick="applyMultiplier(${j.id})">用原始價格 × 乘數 換算</button>
        </div>
      </div>
      <div class="compare-field"><label>多規格價格</label><div class="static-val" style="max-height:220px;overflow-y:auto">${skuPricesHtml}</div></div>
      <div class="compare-field"><label>商品分類</label><input type="text" id="t1_category" value="${esc(j.category||"")}"></div>
      <div class="compare-field"><label>狀態</label><select id="t1_status">${["草稿","待審核","待上架","上架中","已下架"].map(s=>`<option${j.listing_status===s?" selected":""}>${s}</option>`).join("")}</select></div>
      <button class="btn-save" onclick="saveTab1(${j.id})">儲存</button>
    </div>
  </div>`;
}

function applyMultiplier(id){
  const mult = parseFloat(document.getElementById("t1_mult").value);
  if(!mult || mult<=0){toast("請輸入有效的乘數");return;}
  const raw = (_curJob && _curJob.raw_price) || "";
  const skuPrices = (_curJob && _curJob.raw_extra && _curJob.raw_extra.sku_prices) || [];
  let nums = skuPrices.map(p=>parseFloat(p.price)).filter(n=>!isNaN(n) && n>0);
  if(!nums.length) nums = (raw.match(/[\\d.]+/g)||[]).map(Number).filter(n=>n>0);
  if(!nums.length){toast("沒有原始價格可供換算");return;}
  const lo = Math.round(Math.min(...nums)*mult);
  const hi = Math.round(Math.max(...nums)*mult);
  document.getElementById("t1_price_min").value = lo;
  document.getElementById("t1_price_max").value = hi;
  toast(`已套用乘數 ${mult}：NT$ ${lo}${hi!==lo?" ~ "+hi:""}（記得按儲存）`);
}

async function saveTab1(id){
  const body = {
    ai_name: document.getElementById("t1_name").value.trim(),
    brand: document.getElementById("t1_brand").value,
    price_min: document.getElementById("t1_price_min").value.trim(),
    price_max: document.getElementById("t1_price_max").value.trim(),
    category: document.getElementById("t1_category").value.trim(),
    listing_status: document.getElementById("t1_status").value,
  };
  const r = await api("/api/products/"+id,{method:"PUT",body:JSON.stringify(body)});
  if(r.ok){toast("已儲存");openJob(id);} else toast("儲存失敗");
}

function tab2Html(j){
  let h="";
  const pi = j.product_images || {};
  const mainImgs   = pi.main_images   || [];
  const detailImgs = pi.detail_images || [];
  const skuImgs    = pi.sku_images    || [];
  const reviewImgs = pi.review_images || [];
  const videoUrls  = pi.video_urls    || [];
  _selImgs = new Set(j.raw_images || []);
  _lbImgs = [];
  const _addLb = (srcs, labels, cat) => srcs.forEach((s,i) => _lbImgs.push({src:s, label:labels&&labels[i]?labels[i]:cat, cat}));

  // ── 1. 原圖：依分類分區顯示（主圖/SKU/詳情/評價），各自可全選/取消 ──
  const mainSrcs = mainImgs.length ? mainImgs.map(i=>typeof i==='object'?i.src:i) : (j.raw_images||[]);
  const origCount = mainSrcs.length + skuImgs.length + detailImgs.length + reviewImgs.length;
  h += `<div class="compare-col-hd">原圖（${origCount} 張）</div>`;
  if(mainSrcs.length){_addLb(mainSrcs,[],"主圖");h+=imgCatHtml("main","主圖",mainSrcs,[]);}
  if(skuImgs.length){
    const srcs=skuImgs.map(i=>typeof i==='object'?i.src:i);
    const labels=skuImgs.map(i=>typeof i==='object'?(i.label||''):'');
    _addLb(srcs,labels,"SKU");
    h+=imgCatHtml("sku","SKU 規格圖",srcs,labels);
  }
  if(detailImgs.length){
    const srcs=detailImgs.map(i=>typeof i==='object'?i.src:i);
    _addLb(srcs,[],"詳情");
    h+=imgCatHtml("detail","詳情圖",srcs,[]);
  }
  if(reviewImgs.length){
    const srcs=reviewImgs.map(i=>typeof i==='object'?i.src:i);
    _addLb(srcs,[],"評價圖");
    h+=imgCatHtml("review","買家評價圖",srcs,[]);
  }
  if(videoUrls.length){
    h+=`<div class="section"><div class="slabel">影片（${videoUrls.length} 個）</div><div>${videoUrls.map(v=>`<a href="${esc(v)}" target="_blank" style="font-size:12px;display:block;margin:2px 0;color:#1a73e8;word-break:break-all">${esc(v.slice(0,80))}</a>`).join("")}</div></div>`;
  }
  h+=`<hr class="divider">`;

  // ── 2. 翻譯圖 ──
  h += `<div class="compare-col-hd">翻譯圖</div>`;
  const trImgs=j.translated_images||[];
  if(trImgs.length){
    h+=`<div class="translated-sec"><div class="slabel">翻譯完成（${trImgs.length} 張）</div>`
      +`<div class="img-grid">${trImgs.map((src,i)=>`<div class="img-thumb"><img src="${esc(src)}" loading="lazy" onerror="this.style.display='none'"><div class="thumb-actions"><button class="thumb-act" onclick="downloadImg('${esc(src)}')">⬇</button><button class="thumb-act" onclick="event.stopPropagation();cp(this,'${esc(src)}')">📋</button></div></div>`).join("")}</div>`
      +`<button class="sel-btn" style="margin-top:8px" onclick="useTranslated(${j.id})">✓ 以翻譯圖作為輸出</button>`
      +`</div>`;
  }
  const piTr = j.product_images || {};
  const trTypeMap = {main:'主圖', detail:'詳情圖', sku:'SKU圖'};
  ['main','detail','sku'].forEach(t => {
    const tImgs = piTr['tr_'+t+'_images'] || [];
    if (!tImgs.length) return;
    h += `<div class="tr-type-sec"><div class="tr-type-label">翻譯圖（${trTypeMap[t]}，${tImgs.length} 張）</div><div class="img-grid">${tImgs.map((src,i)=>`<div class="img-thumb"><img src="${esc(src)}" loading="lazy" onerror="this.style.display='none'"><div class="thumb-actions"><button class="thumb-act" onclick="downloadImg('${esc(src)}')">⬇</button><button class="thumb-act" onclick="event.stopPropagation();cp(this,'${esc(src)}')">📋</button></div></div>`).join("")}</div></div>`;
  });
  if(!trImgs.length && !['main','detail','sku'].some(t=>(piTr['tr_'+t+'_images']||[]).length)){
    h += `<div style="color:#bbb;font-size:13px;padding:6px 0">尚無翻譯圖片</div>`;
  }
  h += `<div class="upload-tr-sec">
    <div class="slabel">上傳已翻譯圖片（客優雲翻譯後）</div>
    <div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap">
      <select id="upType_${j.id}" style="border:1.5px solid #b39ddb;border-radius:8px;padding:6px 10px;font-size:13px;color:#4527a0">
        <option value="main">主圖</option>
        <option value="detail">詳情圖</option>
        <option value="sku">SKU圖</option>
      </select>
      <label class="btn-upload-lbl" for="upFiles_${j.id}">選擇圖片</label>
      <input type="file" id="upFiles_${j.id}" multiple accept="image/*" style="display:none" onchange="prevUpload(${j.id})">
      <span id="upCount_${j.id}" style="font-size:12px;color:#666"></span>
      <button class="btn-primary" id="upBtn_${j.id}" onclick="doUpload(${j.id})" disabled style="padding:6px 16px;font-size:13px;width:auto">上傳</button>
    </div>
    <div id="upPreview_${j.id}" class="img-grid" style="margin-top:8px;max-height:180px;overflow-y:auto"></div>
  </div>`;
  h+=`<hr class="divider">`;

  // ── 3. 白底圖 ──
  h += `<div class="compare-col-hd">白底圖</div>`;
  if(j.processed_images&&j.processed_images.length){
    const zipUrl=`/api/products/${j.id}/images/zip?key=${KEY}`;
    h+=`<div class="section"><div class="slabel" style="display:flex;align-items:center;gap:8px">已處理（白底，${j.processed_images.length} 張）<a href="${zipUrl}" class="export-btn" style="font-size:11px">⬇ ZIP</a></div><div class="img-grid">${j.processed_images.map(img=>`<div class="img-thumb"><img src="${esc(img)}" loading="lazy" onerror="this.style.display='none'"></div>`).join("")}</div></div>`;
  } else {
    h += `<div style="color:#bbb;font-size:13px;padding:6px 0">尚無白底圖，可在下方選取原圖後生成</div>`;
  }

  h+=`<div class="sel-action-bar">
    <span class="sel-count" id="selCount">已選 ${_selImgs.size} 張</span>
    <button class="sel-btn" onclick="toggleAllCats(true)">全選全部</button>
    <button class="sel-btn" onclick="toggleAllCats(false)">取消全部</button>
    <button class="btn-translate" id="btnTr_${j.id}" onclick="translateSelected(${j.id})">文A 翻譯選取</button>
    <button class="sel-btn sel-btn-white" onclick="whitebgSelected(${j.id})">⬜ 生成白底圖</button>
    <button class="btn-zip" onclick="downloadZipSelected(${j.id})">⬇ ZIP 下載</button>
    <button class="btn-confirm" onclick="confirmSelect(${j.id})">確認選圖</button>
  </div>`;
  return h;
}

function toggleAllCats(checked){
  ['main','sku','detail','review'].forEach(catId=>toggleAllInCat(catId,checked));
}

async function whitebgSelected(id){
  const urls=[..._selImgs];
  if(!urls.length){toast('請先勾選圖片');return;}
  toast('送出 '+urls.length+' 張，處理中...');
  try{
    const r=await api('/api/products/'+id+'/whitebg-selected',{method:'POST',body:JSON.stringify({urls}),headers:{'Content-Type':'application/json'}});
    if(r.ok){toast('白底圖處理中（'+r.count+' 張）...');pollWhitebg(id,0);}
    else{toast('失敗：'+(r.error||''));}
  }catch(e){toast('錯誤：'+e.message);}
}

function tab3Html(j){
  const fixNl = (v) => String(v||"").replace(/\\r\\n|\\n/g, "\\n");
  const field = (label, id, val, multiline) => `<div class="copy-field">
    <div class="copy-field-hd"><div class="slabel">${label}</div><button class="copy-btn" style="position:static" onclick='cp(this,document.getElementById("${id}").value)'>複製</button></div>
    ${multiline?`<textarea id="${id}" style="height:${multiline}px">${esc(fixNl(val))}</textarea>`:`<input type="text" id="${id}" value="${esc(val||"")}">`}
  </div>`;

  // ── Pipeline 狀態卡 ────────────────────────────────────────────
  const log = j.pipeline_log || "";
  const hasFail = log.includes("PIPELINE_FAIL:");
  let badge, badgeNote;
  if (j.pipeline_used) {
    badge = `<span style="background:#e8f5e9;color:#2e7d32;border-radius:6px;padding:2px 10px;font-size:12px;font-weight:600">Pipeline 成功</span>`;
    badgeNote = `<span style="font-size:12px;color:#888">5 階段分析完成</span>`;
  } else if (hasFail) {
    const failLine = log.split('\\n').find(l=>l.includes('PIPELINE_FAIL:')) || '';
    const reason = failLine.replace('PIPELINE_FAIL:','').trim().slice(0,70);
    badge = `<span style="background:#fff3e0;color:#e65100;border-radius:6px;padding:2px 10px;font-size:12px;font-weight:600">Fallback 舊版</span>`;
    badgeNote = `<span style="font-size:12px;color:#e65100" title="${esc(reason)}">${esc(reason.slice(0,50))}${reason.length>50?'…':''}</span>`;
  } else {
    badge = `<span style="background:#f5f5f5;color:#999;border-radius:6px;padding:2px 10px;font-size:12px">舊版生成</span>`;
    badgeNote = '';
  }
  let logHtml = '';
  if (log.trim()) {
    const lines = log.trim().split('\\n').filter(l=>l.trim());
    const logLines = lines.map(line => {
      let color = '#555';
      if (/完成|Pipeline 全部/.test(line)) color = '#2e7d32';
      else if (/失敗|FAIL/i.test(line))   color = '#c62828';
      else if (/開始|平行/.test(line))     color = '#1565c0';
      return `<div style="padding:1px 0;color:${color};font-size:11px;font-family:monospace">${esc(line)}</div>`;
    }).join('');
    logHtml = `<details style="margin-top:6px"><summary style="font-size:11px;color:#aaa;cursor:pointer">執行 log（${lines.length} 步）</summary>
      <div style="background:#f8f8f8;border-radius:6px;padding:6px 8px;margin-top:4px;line-height:1.7">${logLines}</div></details>`;
  }
  const jsonCard = (title, obj) => {
    if (!obj || !Object.keys(obj).length) return '';
    const rows = Object.entries(obj).map(([k,v]) => {
      const val = Array.isArray(v) ? v.join('、') : (typeof v==='boolean'?(v?'是':'否'):String(v));
      return `<tr><td style="padding:3px 8px;color:#888;white-space:nowrap;vertical-align:top;width:130px">${esc(k)}</td><td style="padding:3px 8px">${esc(val)}</td></tr>`;
    }).join('');
    return `<details style="margin-bottom:7px;border:1px solid #eee;border-radius:8px;overflow:hidden">
      <summary style="padding:6px 12px;cursor:pointer;background:#fafafa;font-size:12px;font-weight:600;color:#555">${title}</summary>
      <table style="width:100%;font-size:12px;border-collapse:collapse">${rows}</table></details>`;
  };

  let h = `<div id="pipelineStatusBlock_${j.id}" style="background:#fafafa;border:1px solid #eee;border-radius:10px;padding:10px 14px;margin-bottom:14px">
    <div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap">
      <span style="font-size:12px;color:#888;font-weight:600;flex-shrink:0">生成方式</span>
      ${badge} ${badgeNote}
      <div style="margin-left:auto;display:flex;gap:6px;flex-shrink:0">
        <button id="regenOld_${j.id}" onclick="regenerateCopy(${j.id},false)" style="background:#f5f5f5;border:none;border-radius:7px;padding:4px 11px;font-size:12px;cursor:pointer;color:#555">↻ 舊版</button>
        <button id="regenPipe_${j.id}" onclick="regenerateCopy(${j.id},true)" style="background:#e8f0fe;color:#1a73e8;border:none;border-radius:7px;padding:4px 11px;font-size:12px;cursor:pointer">✦ Pipeline</button>
      </div>
    </div>
    ${logHtml}
  </div>`;

  h += jsonCard("商品分析（Stage 1）", j.analysis_json);
  h += jsonCard("搜尋意圖（Stage 3）", j.search_intent_json);
  h += jsonCard("競品分析（Stage 4）", j.competitor_json);

  if (j.analysis_json && Object.keys(j.analysis_json).length) {
    h += `<div style="border-top:1px solid #eee;margin:10px 0"></div>`;
  }
  h += field("商品名稱","t3_name",j.ai_name);
  h += field("商品描述","t3_desc",j.ai_desc,180);
  h += field("商品特色","t3_features",j.features,100);
  h += field("SEO 關鍵字（逗號分隔）","t3_kw",j.ai_keywords);
  h += `<div class="edit-bar">
    <button class="btn-save" onclick="saveTab3(${j.id})">儲存文案</button>
  </div>`;
  return h;
}

async function saveTab3(id){
  const body = {
    ai_name: document.getElementById("t3_name").value.trim(),
    ai_desc: document.getElementById("t3_desc").value.trim(),
    features: document.getElementById("t3_features").value.trim(),
    ai_keywords: document.getElementById("t3_kw").value.trim(),
  };
  const r = await api("/api/products/"+id,{method:"PUT",body:JSON.stringify(body)});
  if(r.ok){toast("已儲存");openJob(id);} else toast("儲存失敗");
}

async function regenerateCopy(id, usePipeline){
  const btnOld  = document.getElementById(`regenOld_${id}`);
  const btnPipe = document.getElementById(`regenPipe_${id}`);
  if(btnOld){btnOld.disabled=true;btnOld.textContent='...';}
  if(btnPipe){btnPipe.disabled=true;btnPipe.textContent='...';}

  const body  = usePipeline !== undefined ? JSON.stringify({use_pipeline: usePipeline}) : null;
  const label = usePipeline ? "Pipeline" : "舊版";
  const r = await api("/api/products/"+id+"/regenerate-copy",{method:"POST",body});
  if(!r.ok){
    toast("送出失敗："+(r.error||""));
    if(btnOld){btnOld.disabled=false;btnOld.textContent='↻ 舊版';}
    if(btnPipe){btnPipe.disabled=false;btnPipe.textContent='✦ Pipeline';}
    return;
  }
  const statusBlock = document.getElementById(`pipelineStatusBlock_${id}`);
  if(statusBlock) statusBlock.innerHTML = `<div style="display:flex;align-items:center;gap:8px;color:#666;font-size:13px"><span class="spinner"></span> ${esc(label)} 生成中，請稍候...</div>`;
  pollRegenerate(id, 0, usePipeline);
}

async function pollRegenerate(id, tries, usePipeline){
  if(tries > 30){toast("生成超時（超過 60 秒），請重試");if(openId===id)openJob(id);return;}
  const j = await api("/api/products/"+id);
  if(j.status === "done"){
    const label = usePipeline ? "Pipeline 分析完成" : "文案已更新";
    toast(label + " — 畫面已更新");
    if(openId===id){await openJob(id);switchModalTab("tab3");}
  } else if(j.status === "error"){
    toast("生成失敗：" + (j.error_msg||"未知錯誤"));
    if(openId===id){await openJob(id);switchModalTab("tab3");}
  } else {
    setTimeout(()=>pollRegenerate(id, tries+1, usePipeline), 2000);
  }
}

function tab4Html(j){
  const faq = (j.faq && j.faq.length) ? j.faq.slice() : [];
  while(faq.length<3) faq.push({q:"",a:""});
  const faqRows = faq.map((f,i)=>`<div class="faq-item" id="faqRow_${i}">
    <button class="faq-del" onclick="removeFaqRow(${i})">×</button>
    <input type="text" placeholder="問題" class="faq-q" value="${esc(f.q||"")}">
    <textarea placeholder="回答" class="faq-a">${esc(f.a||"")}</textarea>
  </div>`).join("");
  return `<div class="copy-field">
    <div class="copy-field-hd"><div class="slabel">SEO 描述</div><button class="copy-btn" style="position:static" onclick='cp(this,document.getElementById("t4_seodesc").value)'>複製</button></div>
    <textarea id="t4_seodesc" style="height:70px">${esc(j.seo_desc||"")}</textarea>
  </div>
  <div class="slabel" style="margin-top:10px">FAQ</div>
  <div id="faqList">${faqRows}</div>
  <button class="btn-add-faq" onclick="addFaqRow()">+ 新增 FAQ</button>
  <div class="edit-bar"><button class="btn-save" onclick="saveTab4(${j.id})">儲存</button></div>`;
}

function addFaqRow(){
  const list=document.getElementById("faqList");
  const i=list.children.length;
  const div=document.createElement("div");
  div.className="faq-item"; div.id="faqRow_"+i;
  div.innerHTML=`<button class="faq-del" onclick="removeFaqRow(${i})">×</button><input type="text" placeholder="問題" class="faq-q"><textarea placeholder="回答" class="faq-a"></textarea>`;
  list.appendChild(div);
}
function removeFaqRow(i){
  const el=document.getElementById("faqRow_"+i);
  if(el) el.remove();
}
async function saveTab4(id){
  const seoDesc=document.getElementById("t4_seodesc").value.trim();
  const faq=[...document.querySelectorAll("#faqList .faq-item")].map(row=>({
    q: row.querySelector(".faq-q").value.trim(),
    a: row.querySelector(".faq-a").value.trim(),
  })).filter(f=>f.q||f.a);
  const r=await api("/api/products/"+id,{method:"PUT",body:JSON.stringify({seo_desc:seoDesc,faq})});
  if(r.ok){toast("已儲存");openJob(id);} else toast("儲存失敗");
}

async function saveTab5(id){
  const body = {
    shopee_title: document.getElementById("t5_shopee").value.trim(),
    price_min: document.getElementById("t5_price_min").value.trim(),
    price_max: document.getElementById("t5_price_max").value.trim(),
    website_name: document.getElementById("t5_website").value.trim(),
  };
  const r = await api("/api/products/"+id,{method:"PUT",body:JSON.stringify(body)});
  if(r.ok){toast("已儲存");openJob(id);} else toast("儲存失敗");
}

function tab5Html(j){
  let h = `<div class="section"><div class="slabel">上架狀態</div><span class="status-pill">${esc(j.listing_status||"草稿")}</span></div>`;
  h += `<div class="compare-col-hd">蝦皮資料</div>`;
  h += `<div class="copy-field"><div class="slabel">蝦皮標題</div><input type="text" id="t5_shopee" value="${esc(j.shopee_title||"")}" placeholder="可手動填寫，或在 AI 文案分頁按重新生成"></div>`;
  h += `<div class="copy-field"><div class="slabel">建議售價</div><div style="display:flex;gap:8px;align-items:center">NT$ <input type="text" id="t5_price_min" style="width:90px" value="${esc(j.price_min||"")}"> ~ <input type="text" id="t5_price_max" style="width:90px" value="${esc(j.price_max||"")}"></div></div>`;
  h += `<div class="compare-col-hd" style="margin-top:18px">官網資料</div>`;
  h += `<div class="copy-field"><div class="slabel">官網商品名稱</div><input type="text" id="t5_website" value="${esc(j.website_name||"")}" placeholder="可手動填寫，或在 AI 文案分頁按重新生成"></div>`;
  h += `<div class="edit-bar"><button class="btn-save" onclick="saveTab5(${j.id})">儲存上架資料</button></div>`;
  h += `<div class="compare-col-hd" style="margin-top:18px">匯出</div>`;
  h += `<div class="section" style="display:flex;gap:8px">
    <a class="export-btn" href="/api/products/export?format=xlsx&key=${KEY}">⬇ 匯出 Excel</a>
    <a class="export-btn" href="/api/products/export?format=csv&key=${KEY}">⬇ 匯出 CSV</a>
  </div>
  <div style="font-size:12px;color:#bbb;margin-top:6px">目前匯出尚未串接蝦皮 / 官網實際上架，僅供資料整理使用。</div>`;
  h += `<hr class="divider">`;
  h += `<div class="compare-col-hd">原始資料</div>`;
  h += `<div class="section"><div class="slabel">平台 / 來源網址</div><div class="rbox">${esc(j.platform)} — <a href="${esc(j.url)}" target="_blank" style="color:#1a73e8;word-break:break-all">${esc(j.url)}</a></div></div>`;
  if(j.raw_title) h+=`<div class="section"><div class="slabel">原始標題</div><div class="rbox">${esc(j.raw_title)}</div></div>`;
  if(j.raw_price) h+=`<div class="section"><div class="slabel">原始價格（人民幣）</div><div class="rbox">¥ ${esc(j.raw_price)}</div></div>`;
  const re2 = j.raw_extra || {};
  if(re2.sku_props && re2.sku_props.length){
    let sp = re2.sku_props.map(p=>`<b>${esc(p.name)}</b>：${(p.values||[]).map(v=>esc(v.name)).join('、')}`).join('<br>');
    h+=`<div class="section"><div class="slabel">規格選項</div><div class="rbox" style="line-height:1.8">${sp}</div></div>`;
  }
  if(re2.sku_prices && re2.sku_prices.length){
    const prices2 = re2.sku_prices.map(s=>`¥${esc(s.price)}`);
    const uniq2 = [...new Set(prices2)];
    h+=`<div class="section"><div class="slabel">多規格價格（共 ${re2.sku_prices.length} 筆）</div><div class="rbox">${uniq2.slice(0,20).join('　')}</div></div>`;
  }
  if(j.raw_desc)  h+=`<div class="section"><div class="slabel">原始描述</div><div class="rbox" style="max-height:220px;overflow-y:auto">${esc(j.raw_desc.slice(0,1500))}${j.raw_desc.length>1500?"…":""}</div></div>`;
  return h;
}

function imgCatHtml(catId, label, srcs, labels){
  const catCls={main:'',sku:' cat-sku',detail:' cat-detail',review:' cat-review'}[catId]||'';
  const thumbs=srcs.map((src,idx)=>{
    const checked=_selImgs.has(src);
    const lbl=labels[idx]||'';
    const isMain=_curMainImage&&src===_curMainImage;
    return `<div class="img-thumb${catCls}${checked?" checked":""}" onclick="imgThumbClick(event,this,'${catId}_${idx}')">`
      +`<input type="checkbox" id="ck_${catId}_${idx}" data-url="${esc(src)}"${checked?" checked":""}>`
      +`<img src="${esc(src)}" loading="lazy" onerror="this.style.display='none'" onload="imgSizeLoad(this)" title="${esc(src)}">`
      +`<div class="img-size"></div>`
      +(lbl?`<div class="img-label">${esc(lbl)}</div>`:'')
      +(isMain?`<div class="main-badge">主圖</div>`:'')
      +`<div class="thumb-actions">`
      +`<button class="thumb-act" onclick="event.stopPropagation();downloadImg('${esc(src)}')" title="下載圖片">⬇</button>`
      +`<button class="thumb-act" onclick="event.stopPropagation();cp(this,'${esc(src)}')" title="複製URL">📋</button>`
      +`<button class="thumb-act" onclick="event.stopPropagation();openLightbox('${esc(src)}')" title="放大檢視">⛶</button>`
      +`<button class="thumb-act" onclick="event.stopPropagation();setAsMain('${esc(src)}')" title="設為主圖">★</button>`
      +`</div></div>`;
  }).join("");
  const autoCollapse = (catId==='detail'||catId==='review') && srcs.length>8;
  return `<div class="section" id="cat_${catId}">`
    +`<div class="img-zone-hd">`
    +`<div class="slabel">${label}（${srcs.length} 張）</div>`
    +`<button class="sel-btn" onclick="toggleAllInCat('${catId}',true)">全選</button>`
    +`<button class="sel-btn" onclick="toggleAllInCat('${catId}',false)">取消</button>`
    +`<button class="sel-btn cat-collapse-btn" id="collapseBtn_${catId}" onclick="toggleCatCollapse('${catId}')">${autoCollapse?'▶ 展開':'▼ 收合'}</button>`
    +`</div>`
    +`<div class="img-grid" id="grid_${catId}" style="${autoCollapse?'display:none':''}">${thumbs}</div>`
    +`</div>`;
}
async function setAsMain(url){
  const r=await api(`/api/products/${_curJobId}/set-main-image`,{method:"POST",body:JSON.stringify({url})});
  if(r.ok){_curMainImage=url;toast("已設為主圖");openJob(_curJobId);}
  else toast("設定失敗");
}
function downloadImg(src){
  const a=document.createElement('a');
  a.href=`/api/products/proxy-img?url=${encodeURIComponent(src)}&key=${KEY}`;
  a.download=src.split('/').pop().replace(/\?.*$/,'').replace(/[^a-zA-Z0-9._-]/g,'_')||'image.jpg';
  document.body.appendChild(a);a.click();document.body.removeChild(a);
}
function toggleCatCollapse(catId){
  const grid=document.getElementById('grid_'+catId);
  const btn=document.getElementById('collapseBtn_'+catId);
  if(!grid)return;
  const collapsed=grid.style.display==='none';
  grid.style.display=collapsed?'':'none';
  if(btn)btn.textContent=collapsed?'▼ 收合':'▶ 展開';
}
function imgSizeLoad(img){
  const w=img.naturalWidth, h=img.naturalHeight;
  const wrap=img.closest('.img-thumb');
  if(w>0 && h>0 && (w<100 || h<100)){wrap.style.display='none';return;}
  const el=wrap.querySelector('.img-size');
  if(el&&w>0) el.textContent=w+' × '+h;
}
function imgThumbClick(e,wrap,ckId){
  if(e.target.type==="checkbox") return;
  const cb=document.getElementById("ck_"+ckId);
  if(!cb) return;
  cb.checked=!cb.checked;
  syncThumb(cb,wrap);
}
function syncThumb(cb,wrap){
  const url=cb.dataset.url;
  if(cb.checked){_selImgs.add(url);wrap.classList.add("checked");}
  else{_selImgs.delete(url);wrap.classList.remove("checked");}
  const el=document.getElementById("selCount");
  if(el) el.textContent=`已選 ${_selImgs.size} 張`;
}
function toggleAllInCat(catId,checked){
  document.querySelectorAll(`#cat_${catId} .img-thumb`).forEach(wrap=>{
    const cb=wrap.querySelector("input[type=checkbox]");
    if(!cb) return;
    cb.checked=checked;
    syncThumb(cb,wrap);
  });
}
async function pollTranslate(id, tries){
  if(tries>24){toast("翻譯超時，請重試");return;}
  const j=await api("/api/products/"+id);
  if(j.translate_status==="done"||((j.translated_images||[]).length>0&&j.translate_status!=="processing")){
    toast("翻譯完成！");
    if(openId===id) openJob(id);
  } else {
    setTimeout(()=>pollTranslate(id,tries+1),5000);
  }
}
async function useTranslated(id){
  const j=await api("/api/products/"+id);
  const urls=j.translated_images||[];
  if(!urls.length){toast("沒有翻譯圖片");return;}
  const r=await api("/api/products/"+id+"/select-images",{method:"POST",body:JSON.stringify({urls})});
  if(r.ok) toast("已設定翻譯圖為輸出圖");
}
function openLightbox(src){
  const idx=_lbImgs.findIndex(i=>i.src===src);
  _lbIdx=idx>=0?idx:0;
  _showLb();
}
function _showLb(){
  const img=_lbImgs[_lbIdx];
  if(!img) return;
  document.getElementById('lbImg').src=img.src;
  document.getElementById('lbLabel').textContent=img.label||'';
  document.getElementById('lbCount').textContent=(_lbIdx+1)+' / '+_lbImgs.length;
  document.getElementById('lbBox').classList.remove('hidden');
}
function lbPrev(){_lbIdx=(_lbIdx-1+_lbImgs.length)%_lbImgs.length;_showLb();}
function lbNext(){_lbIdx=(_lbIdx+1)%_lbImgs.length;_showLb();}
function lbClose(){document.getElementById('lbBox').classList.add('hidden');}
document.addEventListener('keydown',e=>{
  if(document.getElementById('lbBox').classList.contains('hidden')) return;
  if(e.key==='ArrowLeft') lbPrev();
  else if(e.key==='ArrowRight') lbNext();
  else if(e.key==='Escape') lbClose();
});
async function translateSelected(id){
  const urls=[..._selImgs];
  await _doTranslate(id,urls);
}
async function _doTranslate(id,urls){
  if(!urls.length){toast('請先選取要翻譯的圖片');return;}
  const btn=document.getElementById('btnTr_'+id);
  if(btn){btn.disabled=true;btn.textContent='翻譯中...';}
  const r=await api('/api/products/'+id+'/translate-images',{method:'POST',body:JSON.stringify({urls})});
  if(r.ok){
    toast('翻譯開始，約需 1-3 分鐘...');
    setTimeout(()=>pollTranslate(id,0),5000);
  } else {
    toast('翻譯失敗：'+(r.error||''));
    if(btn){btn.disabled=false;btn.textContent='文A 翻譯選取';}
  }
}
async function pollWhitebg(id, tries){
  if(tries>20){toast('白底圖處理超時，請至 Render log 查看錯誤');return;}
  try{
    const j=await api('/api/products/'+id);
    if(j.img_status==='done'){
      toast('白底圖完成！');
      if(openId===id) openJob(id);
    } else if(j.img_status==='failed'){
      toast('白底圖處理失敗，請查 Render log');
    } else {
      setTimeout(()=>pollWhitebg(id,tries+1),4000);
    }
  }catch(e){ setTimeout(()=>pollWhitebg(id,tries+1),4000); }
}
async function confirmSelect(id){
  const urls=[..._selImgs];
  const r=await api("/api/products/"+id+"/select-images",{method:"POST",body:JSON.stringify({urls})});
  if(r.ok) toast(`已確認 ${r.count} 張圖片為輸出圖`);
  else toast("儲存失敗："+r.error);
}
async function downloadZipSelected(id){
  const items=[];
  const cc={};
  ['main','sku','detail','review'].forEach(catId=>{
    const el=document.getElementById('cat_'+catId);
    if(!el) return;
    el.querySelectorAll('.img-thumb input[type=checkbox]:checked').forEach(cb=>{
      cc[catId]=(cc[catId]||0)+1;
      items.push({url:cb.dataset.url,cat:catId,idx:cc[catId]});
    });
  });
  // fallback: _selImgs not in any category
  const seen=new Set(items.map(i=>i.url)); let ex=0;
  [..._selImgs].forEach(url=>{ if(!seen.has(url)){ex++;items.push({url,cat:'img',idx:ex});} });
  if(!items.length){toast("請先勾選圖片");return;}
  toast("打包中，請稍候...");
  try{
    const res=await fetch(`/api/products/${id}/images/zip-selected?key=${KEY}`,{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({items})});
    if(!res.ok){toast("下載失敗");return;}
    const blob=await res.blob();
    const a=document.createElement("a");
    a.href=URL.createObjectURL(blob);
    a.download=`product_${id}_images.zip`;
    a.click();
    URL.revokeObjectURL(a.href);
    toast("下載完成");
  }catch(e){toast("下載失敗");}
}

function prevUpload(id){
  const inp=document.getElementById('upFiles_'+id);
  const preview=document.getElementById('upPreview_'+id);
  const count=document.getElementById('upCount_'+id);
  const btn=document.getElementById('upBtn_'+id);
  const files=[...inp.files];
  count.textContent=files.length+' 張已選';
  btn.disabled=files.length===0;
  preview.innerHTML=files.map(f=>{
    const url=URL.createObjectURL(f);
    return `<div class="img-thumb"><img src="${url}" style="object-fit:cover;width:80px;height:80px"></div>`;
  }).join('');
}
async function doUpload(id){
  const inp=document.getElementById('upFiles_'+id);
  const type=document.getElementById('upType_'+id).value;
  const btn=document.getElementById('upBtn_'+id);
  const files=[...inp.files];
  if(!files.length){toast('請選擇圖片');return;}
  btn.disabled=true; btn.textContent='上傳中...';
  const fd=new FormData();
  fd.append('type',type);
  files.forEach(f=>fd.append('files',f));
  try{
    const res=await fetch(`/api/products/${id}/upload-translated?key=${KEY}`,{method:'POST',body:fd});
    const j2=await res.json();
    if(j2.ok){
      toast(`上傳完成，${j2.added} 張`);
      inp.value='';
      document.getElementById('upCount_'+id).textContent='';
      document.getElementById('upPreview_'+id).innerHTML='';
      btn.disabled=true;
      if(openId===id) openJob(id);
    } else { toast('上傳失敗：'+(j2.error||'')); }
  }catch(e){ toast('錯誤：'+e.message); }
  finally{ btn.disabled=false; btn.textContent='上傳'; }
}
async function delJob(e,id){
  e.stopPropagation();
  if(!confirm("確定刪除這筆任務？"))return;
  await api("/api/products/"+id,{method:"DELETE"});
  jobs=jobs.filter(j=>j.id!==id);
  if(openId===id){
    openId=null;
    document.getElementById("detailCol").innerHTML='<div class="detail-empty" id="detailEmpty">👈 點選左側商品查看詳情</div>';
  }
  render();
  toast("已刪除");
}

function cp(btn,text){
  navigator.clipboard.writeText(text).then(()=>{
    const orig=btn.textContent;
    btn.textContent="已複製";btn.classList.add("ok");
    setTimeout(()=>{btn.textContent=orig;btn.classList.remove("ok");},1500);
    toast("已複製");
  });
}
function toast(msg){
  const t=document.getElementById("toast");
  t.textContent=msg;t.classList.add("show");
  setTimeout(()=>t.classList.remove("show"),2000);
}
function esc(s){return String(s||"").replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");}

document.getElementById("urlInput").addEventListener("keydown",e=>{if(e.key==="Enter")submitUrl();});
loadJobs();
</script>
<div id="lbBox" class="lb-overlay hidden" onclick="lbClose()">
  <button class="lb-close" onclick="lbClose()">✕</button>
  <button class="lb-nav lb-prev" onclick="event.stopPropagation();lbPrev()">&#8249;</button>
  <div class="lb-content" onclick="event.stopPropagation()">
    <img id="lbImg" src="" alt="">
    <div class="lb-info"><span class="lb-label" id="lbLabel"></span><span class="lb-count" id="lbCount"></span></div>
  </div>
  <button class="lb-nav lb-next" onclick="event.stopPropagation();lbNext()">&#8250;</button>
</div>
</body></html>"""


# ═══════════════════════════════════════════════════════════
# Routes
# ═══════════════════════════════════════════════════════════

# ── 後台路由 ──────────────────────────────────────────────────

@products_bp.route("/admin/products")
def admin_products():
    ok, key = check_auth()
    if not ok:
        return render_template_string(LOGIN_HTML, next="/admin/products", error=None)
    return render_template_string(PRODUCTS_HTML, key=key)

@products_bp.route("/admin/products/store-scan")
def admin_store_scan():
    ok, key = check_auth()
    if not ok:
        return render_template_string(LOGIN_HTML, next="/admin/products/store-scan", error=None)
    return render_template_string(STORE_SCAN_HTML, key=key)

BRAND_SETTINGS_HTML = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>品牌設定</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,sans-serif;background:#f5f5f5;color:#333}
.header{background:#1a1a1a;color:#fff;padding:14px 20px;display:flex;align-items:center;gap:14px}
.header a{color:#888;text-decoration:none;font-size:14px}
.header a:hover{color:#fff}
.header-title{font-size:17px;font-weight:700;flex:1}
.wrap{max-width:1100px;margin:0 auto;padding:24px 16px}
.brand-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:18px}
.card{background:#fff;border-radius:16px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,.08);transition:box-shadow .15s}
.card.disabled{opacity:.55}
.card-head{display:flex;align-items:center;gap:10px;margin-bottom:16px}
.brand-key{font-size:11px;font-weight:700;background:#f0f0f0;color:#888;border-radius:6px;padding:2px 8px;letter-spacing:.5px}
.brand-name{font-size:16px;font-weight:700;flex:1}
.toggle-wrap{display:flex;align-items:center;gap:6px;font-size:11px;color:#999;font-weight:600}
.switch{position:relative;display:inline-block;width:36px;height:20px}
.switch input{opacity:0;width:0;height:0}
.slider{position:absolute;cursor:pointer;inset:0;background:#ddd;border-radius:20px;transition:.2s}
.slider:before{position:absolute;content:"";height:16px;width:16px;left:2px;bottom:2px;background:#fff;border-radius:50%;transition:.2s}
input:checked+.slider{background:#1a1a1a}
input:checked+.slider:before{transform:translateX(16px)}
label{display:block;font-size:12px;color:#666;font-weight:600;margin-bottom:4px;margin-top:12px}
input[type=text],textarea{width:100%;border:1.5px solid #ddd;border-radius:8px;padding:9px 12px;font-size:13px;font-family:-apple-system,sans-serif;outline:none;resize:vertical}
input[type=text]:focus,textarea:focus{border-color:#1a1a1a}
.btn-save{background:#1a1a1a;color:#fff;border:none;border-radius:8px;padding:9px 22px;font-size:14px;font-weight:700;cursor:pointer;margin-top:14px}
.btn-save:hover{background:#333}
.btn-save:disabled{background:#aaa;cursor:default}
.toast{position:fixed;bottom:28px;left:50%;transform:translateX(-50%);background:#222;color:#fff;padding:10px 22px;border-radius:20px;font-size:14px;opacity:0;transition:opacity .3s;pointer-events:none;z-index:999}
.toast.show{opacity:1}
</style>
</head>
<body>
<div class="header">
  <a href="/admin/products?key={{ key }}">← 商品搬運</a>
  <div class="header-title">品牌設定</div>
</div>
<div class="wrap">
<div class="brand-grid">
{% for p in profiles %}
<div class="card{{ '' if p.enabled else ' disabled' }}" id="card_{{ p.brand_key }}">
  <div class="card-head">
    <span class="brand-key">{{ p.brand_key }}</span>
    <span class="brand-name">{{ p.name }}</span>
    <div class="toggle-wrap">
      啟用
      <label class="switch">
        <input type="checkbox" id="enabled_{{ p.brand_key }}" {{ 'checked' if p.enabled else '' }} onchange="toggleEnabled('{{ p.brand_key }}')">
        <span class="slider"></span>
      </label>
    </div>
  </div>
  <label>品牌定位 / 商品分類</label>
  <input type="text" id="category_{{ p.brand_key }}" value="{{ p.category }}">
  <label>文案語氣</label>
  <textarea id="tone_{{ p.brand_key }}" rows="2">{{ p.tone }}</textarea>
  <label>文案風格</label>
  <textarea id="style_{{ p.brand_key }}" rows="3">{{ p.style }}</textarea>
  <label>SEO 關鍵字方向</label>
  <textarea id="seo_direction_{{ p.brand_key }}" rows="2">{{ p.seo_direction }}</textarea>
  <label>AI 背景提示詞（中英文均可，留空則自動依商品分析）</label>
  <textarea id="image_style_{{ p.brand_key }}" rows="2" placeholder="e.g. 現代北歐臥室，木質地板，柔和自然光">{{ p.image_style }}</textarea>
  <label style="display:flex;align-items:center;gap:8px;cursor:pointer;font-weight:normal;margin-top:6px">
    <input type="checkbox" id="bg_composite_{{ p.brand_key }}" {% if p.bg_composite %}checked{% endif %}>
    <span>安全合成模式（產品像素 100% 不變，只換背景）</span>
  </label>
  <label>自訂 Prompt（選填）</label>
  <textarea id="custom_prompt_{{ p.brand_key }}" rows="3">{{ p.custom_prompt }}</textarea>
  <input type="hidden" id="name_{{ p.brand_key }}" value="{{ p.name }}">
  <button class="btn-save" onclick="save('{{ p.brand_key }}', this)">儲存</button>
</div>
{% endfor %}
</div>
</div>
<div class="toast" id="toast"></div>
<script>
const KEY = '{{ key }}';
function toast(msg){
  const el = document.getElementById('toast');
  el.textContent = msg; el.classList.add('show');
  setTimeout(()=>el.classList.remove('show'), 2200);
}
function fields(bk){
  return {
    name:          document.getElementById('name_'+bk).value.trim(),
    category:      document.getElementById('category_'+bk).value.trim(),
    style:         document.getElementById('style_'+bk).value.trim(),
    tone:          document.getElementById('tone_'+bk).value.trim(),
    custom_prompt: document.getElementById('custom_prompt_'+bk).value.trim(),
    image_style:   document.getElementById('image_style_'+bk).value.trim(),
    seo_direction: document.getElementById('seo_direction_'+bk).value.trim(),
    bg_composite:  document.getElementById('bg_composite_'+bk).checked,
    enabled:       document.getElementById('enabled_'+bk).checked,
  };
}
async function save(bk, btn){
  btn.disabled = true;
  try{
    const r = await fetch('/api/brand-profiles/'+bk+'?key='+KEY, {
      method:'PUT', headers:{'Content-Type':'application/json'},
      body: JSON.stringify(fields(bk))
    });
    const j = await r.json();
    if(j.ok) toast('✓ '+bk+' 已儲存');
    else toast('儲存失敗');
  }catch(e){ toast('錯誤：'+e.message); }
  finally{ btn.disabled = false; }
}
async function toggleEnabled(bk){
  document.getElementById('card_'+bk).classList.toggle('disabled', !document.getElementById('enabled_'+bk).checked);
  try{
    await fetch('/api/brand-profiles/'+bk+'?key='+KEY, {
      method:'PUT', headers:{'Content-Type':'application/json'},
      body: JSON.stringify(fields(bk))
    });
    toast(document.getElementById('enabled_'+bk).checked ? '已啟用' : '已停用');
  }catch(e){ toast('錯誤：'+e.message); }
}
</script>
</body>
</html>"""


IMAGE_BG_HTML = """<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>AI 背景生成</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f5f7;color:#1a1a1a;min-height:100vh}
.topbar{background:#fff;border-bottom:1px solid #e8e8e8;padding:0 24px;height:54px;display:flex;align-items:center;gap:12px}
.topbar a{color:#666;text-decoration:none;font-size:14px}.topbar a:hover{color:#000}
.topbar .sep{color:#ccc}.topbar h1{font-size:15px;font-weight:600}
.page{display:grid;grid-template-columns:340px 1fr;height:calc(100vh - 54px);overflow:hidden}
.left{background:#fff;border-right:1px solid #eee;padding:20px;overflow-y:auto;display:flex;flex-direction:column;gap:14px}
.panel-title{font-size:12px;font-weight:700;color:#999;letter-spacing:.8px;text-transform:uppercase;margin-bottom:2px}
.drop-zone{border:2px dashed #d8d8d8;border-radius:12px;padding:24px 16px;text-align:center;cursor:pointer;transition:.2s;background:#fafafa;position:relative;min-height:130px;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:5px}
.drop-zone input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.drop-zone:hover,.drop-zone.drag{border-color:#1a1a1a;background:#f0f0f0}
.drop-zone .dz-icon{font-size:26px}.drop-zone .dz-hint{font-size:11px;color:#aaa;line-height:1.4}
.drop-zone .dz-name{font-size:11px;font-weight:600;color:#333;word-break:break-all}
.thumb-row{display:flex;align-items:center;gap:10px;background:#f7f7f7;border-radius:8px;padding:8px;display:none}
.thumb-row img{width:52px;height:52px;border-radius:6px;object-fit:contain;border:1px solid #eee;background:repeating-conic-gradient(#e8e8e8 0% 25%,#fff 0% 50%) 0 0/12px 12px}
.thumb-row .tinfo{font-size:11px;color:#666;flex:1;overflow:hidden}
.thumb-row .tinfo .tname{font-weight:600;color:#333;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.fl{display:block;font-size:12px;font-weight:500;color:#666;margin-bottom:3px}
select,textarea{width:100%;border:1px solid #e0e0e0;border-radius:8px;padding:7px 10px;font-size:13px;font-family:inherit;outline:none;transition:border .2s;resize:vertical}
select:focus,textarea:focus{border-color:#1a1a1a}
.toggle{display:flex;align-items:center;gap:8px;padding:9px 10px;background:#f7f7f7;border-radius:8px;cursor:pointer;user-select:none}
.toggle input[type=checkbox]{width:15px;height:15px;cursor:pointer;accent-color:#1a1a1a;flex-shrink:0}
.toggle span{font-size:12px;color:#444;line-height:1.4}
.btn-primary{width:100%;padding:11px;border:none;border-radius:10px;background:#1a1a1a;color:#fff;font-size:14px;font-weight:600;cursor:pointer;transition:.15s}
.btn-primary:hover{background:#333}.btn-primary:disabled{background:#bbb;cursor:not-allowed}
.err{background:#fff3f3;border:1px solid #ffd0d0;border-radius:8px;padding:9px;font-size:12px;color:#c00;display:none}
.right{overflow-y:auto;padding:24px;display:flex;flex-direction:column;gap:22px}
.empty-state{flex:1;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:12px;color:#ccc;text-align:center;min-height:300px}
.empty-state .ei{font-size:48px;line-height:1}.empty-state p{font-size:14px;line-height:1.6}
.section-card{background:#fff;border-radius:14px;border:1px solid #eee;padding:18px}
.section-hd{font-size:13px;font-weight:600;display:flex;align-items:center;gap:8px;margin-bottom:14px}
.badge{display:inline-block;font-size:10px;padding:2px 7px;border-radius:5px;font-weight:700}
.badge-free{background:#e8f5e9;color:#2e7d32}.badge-paid{background:#fff3e0;color:#e65100}
.cutout-row{display:flex;gap:14px;align-items:center;margin-bottom:18px}
.cutout-thumb{width:100px;height:100px;border-radius:10px;object-fit:contain;border:1.5px dashed #ddd;background:repeating-conic-gradient(#e8e8e8 0% 25%,#fff 0% 50%) 0 0/14px 14px;flex-shrink:0}
.cutout-desc{font-size:12px;color:#777;line-height:1.7}
.swatches{display:grid;grid-template-columns:repeat(5,1fr);gap:8px}
.swatch{border-radius:9px;aspect-ratio:1;cursor:pointer;border:2.5px solid transparent;transition:.12s;position:relative;overflow:hidden}
.swatch:hover{transform:scale(1.07);border-color:#aaa}.swatch.active{border-color:#1a1a1a;box-shadow:0 0 0 2px rgba(0,0,0,.6)}
.swatch .sl{position:absolute;bottom:0;left:0;right:0;font-size:9px;text-align:center;padding:2px;background:rgba(0,0,0,.32);color:#fff}
.preview-area{margin-top:14px;display:none}
.preview-wrap{border-radius:12px;overflow:hidden;background:#e8e8e8;position:relative;min-height:80px}
.preview-wrap img{width:100%;display:block;max-height:400px;object-fit:contain}
.preview-loading{position:absolute;inset:0;background:rgba(255,255,255,.82);display:flex;align-items:center;justify-content:center;flex-direction:column;gap:8px}
.spin{width:32px;height:32px;border:3px solid #eee;border-top-color:#1a1a1a;border-radius:50%;animation:spin .7s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
.dl-row{display:flex;gap:10px;margin-top:10px}
.btn-dl{flex:1;padding:9px;background:#1a1a1a;color:#fff;border:none;border-radius:8px;font-size:12px;font-weight:600;cursor:pointer;text-decoration:none;display:flex;align-items:center;justify-content:center;gap:5px}
.btn-dl:hover{background:#333}
.btn-outline{flex:1;padding:9px;background:#fff;color:#1a1a1a;border:1.5px solid #d0d0d0;border-radius:8px;font-size:12px;font-weight:600;cursor:pointer}
.btn-outline:hover{border-color:#1a1a1a}
.ai-row{display:flex;gap:8px;align-items:flex-end;margin-bottom:10px}
.ai-row textarea{flex:1;height:68px}
.btn-ai{padding:9px 16px;border:none;border-radius:8px;background:#1a1a1a;color:#fff;font-size:13px;font-weight:600;cursor:pointer;flex-shrink:0;transition:.15s}
.btn-ai:hover{background:#333}.btn-ai:disabled{background:#bbb;cursor:not-allowed}
.compare{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.compare-img .cl{font-size:11px;color:#999;margin-bottom:5px}
.compare-img img{width:100%;border-radius:10px;border:1px solid #eee;object-fit:contain;max-height:320px}
@media(max-width:720px){.page{grid-template-columns:1fr;height:auto;overflow:visible}.left{height:auto}.compare{grid-template-columns:1fr}}
</style>
</head>
<body>
<div class="topbar">
  <a href="/admin?key={{ key }}">← 後台</a><span class="sep">/</span>
  <h1>AI 背景生成</h1>
</div>
<div class="page">

  <!-- ═══ Left panel ═══ -->
  <div class="left">
    <div class="panel-title">Step 1 — 上傳商品圖</div>

    <div class="drop-zone" id="dropZone">
      <input type="file" id="fileInput" accept="image/*" onchange="onFile(this)">
      <div class="dz-icon">🖼️</div>
      <div class="dz-hint">點擊或拖曳圖片<br>JPG / PNG / WebP</div>
      <div class="dz-name" id="fname"></div>
    </div>
    <div class="thumb-row" id="thumbRow">
      <img id="origThumb" src="">
      <div class="tinfo"><div class="tname" id="thumbName"></div><div>已選取，按下方按鈕開始去背</div></div>
    </div>

    <hr style="border:none;border-top:1px solid #f0f0f0">

    <div class="panel-title">選填 — 套用品牌設定</div>
    <select id="brandSel" onchange="loadBrand(this.value)">
      <option value="">— 不套用品牌 —</option>
      {% for p in profiles %}
      <option value="{{ p.brand_key }}" data-prompt="{{ p.image_style or '' }}" data-composite="{{ 'true' if p.bg_composite else 'false' }}">{{ p.name }}</option>
      {% endfor %}
    </select>

    <div>
      <label class="fl">AI 提示詞（中英文均可）</label>
      <textarea id="prompt" rows="2" placeholder="例：現代木質客廳 自然光&#10;modern nordic bedroom"></textarea>
    </div>

    <label class="toggle">
      <input type="checkbox" id="composite">
      <span>安全合成模式（產品像素 100% 不變）</span>
    </label>
    <label class="toggle">
      <input type="checkbox" id="skipRmbg">
      <span>圖片已去背，跳過 remove.bg</span>
    </label>

    <button class="btn-primary" id="rmbgBtn" onclick="doRemoveBg()" disabled>Step 1 — 去背 ＆ 載入</button>
    <div class="err" id="errBox"></div>
  </div>

  <!-- ═══ Right panel ═══ -->
  <div class="right" id="rightPanel">
    <div class="empty-state" id="emptyState">
      <div class="ei">✨</div>
      <p>上傳商品圖片<br>按「去背 ＆ 載入」開始</p>
    </div>

    <!-- Step 2: 快速場景 -->
    <div class="section-card" id="step2Card" style="display:none">
      <div class="section-hd">Step 2 — 快速場景 <span class="badge badge-free">免費 · 即時</span></div>
      <div class="cutout-row">
        <img class="cutout-thumb" id="cutoutImg" src="">
        <div class="cutout-desc">
          去背完成 ✅<br>點擊下方色塊立即套用場景。<br>
          <span style="color:#999;font-size:11px">· 快速場景：免費合成<br>· AI 場景：下方輸入提示詞生成</span>
        </div>
      </div>
      <div class="swatches">
        <div class="swatch" style="background:#fff;border:1px solid #ddd" data-bg="white"    onclick="qc(this,'white')"  ><div class="sl">純白</div></div>
        <div class="swatch" style="background:#f2f2f2"                    data-bg="grey"     onclick="qc(this,'grey')"   ><div class="sl">淺灰</div></div>
        <div class="swatch" style="background:#f5f0e8"                    data-bg="cream"    onclick="qc(this,'cream')"  ><div class="sl">奶油</div></div>
        <div class="swatch" style="background:#e8f2f8"                    data-bg="blue"     onclick="qc(this,'blue')"   ><div class="sl">天空</div></div>
        <div class="swatch" style="background:#1e1e1e"                    data-bg="dark"     onclick="qc(this,'dark')"   ><div class="sl" style="background:rgba(255,255,255,.2)">深色</div></div>
        <div class="swatch" style="background:linear-gradient(135deg,#ffecd2,#fcb69f)" data-bg="warm"   onclick="qc(this,'warm')"  ><div class="sl">暖橙</div></div>
        <div class="swatch" style="background:linear-gradient(135deg,#a8edea,#fed6e3)" data-bg="cool"   onclick="qc(this,'cool')"  ><div class="sl">清涼</div></div>
        <div class="swatch" style="background:linear-gradient(135deg,#d4fc79,#96e6a1)" data-bg="green"  onclick="qc(this,'green')" ><div class="sl">清新</div></div>
        <div class="swatch" style="background:linear-gradient(135deg,#667eea,#764ba2)" data-bg="purple" onclick="qc(this,'purple')"><div class="sl">質感</div></div>
        <div class="swatch" style="background:linear-gradient(135deg,#2c3e50,#4ca1af)" data-bg="ocean"  onclick="qc(this,'ocean')" ><div class="sl">深海</div></div>
      </div>
      <div class="preview-area" id="previewArea">
        <div class="preview-wrap" id="previewWrap">
          <img id="previewImg" src="">
          <div class="preview-loading" id="prevLoad" style="display:none"><div class="spin"></div><span style="font-size:11px;color:#666">合成中…</span></div>
        </div>
        <div class="dl-row">
          <a class="btn-dl" id="quickDl" download="result.jpg">⬇ 下載</a>
          <button class="btn-outline" onclick="clearPreview()">換場景</button>
        </div>
      </div>
    </div>

    <!-- Step 3: AI 生成 -->
    <div class="section-card" id="aiCard" style="display:none">
      <div class="section-hd">Step 3 — AI 場景生成 <span class="badge badge-paid">每次計費</span></div>
      <div class="ai-row">
        <textarea id="aiPrompt" placeholder="例：現代北歐臥室，木地板，窗邊自然光&#10;clean studio with white marble floor"></textarea>
        <button class="btn-ai" id="aiBtn" onclick="doAi()">生成</button>
      </div>
      <div class="err" id="aiErr"></div>
      <div id="aiResult" style="display:none">
        <div class="compare">
          <div class="compare-img"><div class="cl">去背原圖</div><img id="aiOrigImg"></div>
          <div class="compare-img"><div class="cl">AI 場景</div><img id="aiResultImg"></div>
        </div>
        <div class="dl-row">
          <a class="btn-dl" id="aiDl" download="ai_result.jpg">⬇ 下載 AI 結果</a>
          <button class="btn-outline" onclick="resetAi()">重新生成</button>
        </div>
      </div>
    </div>
  </div>
</div>

<script>
const KEY='{{ key }}';
let selectedFile=null,transparentUrl=null;

// ─ drop zone ─
const dz=document.getElementById('dropZone');
dz.addEventListener('dragover',e=>{e.preventDefault();dz.classList.add('drag');});
dz.addEventListener('dragleave',()=>dz.classList.remove('drag'));
dz.addEventListener('drop',e=>{e.preventDefault();dz.classList.remove('drag');const f=e.dataTransfer.files[0];if(f&&f.type.startsWith('image/'))setFile(f);});
function onFile(i){if(i.files[0])setFile(i.files[0]);}
function setFile(f){
  selectedFile=f;
  document.getElementById('fname').textContent=f.name;
  document.getElementById('thumbName').textContent=f.name;
  document.getElementById('rmbgBtn').disabled=false;
  const r=new FileReader();
  r.onload=e=>{
    document.getElementById('origThumb').src=e.target.result;
    document.getElementById('thumbRow').style.display='flex';
    dz.style.display='none';
  };
  r.readAsDataURL(f);
}

// ─ brand ─
function loadBrand(bk){
  const o=document.querySelector('#brandSel option[value="'+bk+'"]');
  if(!o)return;
  if(o.dataset.prompt){document.getElementById('prompt').value=o.dataset.prompt;document.getElementById('aiPrompt').value=o.dataset.prompt;}
  document.getElementById('composite').checked=(o.dataset.composite==='true');
}

// ─ util ─
function showErr(id,m){const e=document.getElementById(id);e.textContent=m;e.style.display='block';}
function hideErr(id){document.getElementById(id).style.display='none';}

// ─ Step 1: remove bg ─
async function doRemoveBg(){
  if(!selectedFile)return;
  hideErr('errBox');
  const btn=document.getElementById('rmbgBtn');
  btn.disabled=true;btn.textContent='去背中…';
  document.getElementById('emptyState').style.display='none';
  document.getElementById('step2Card').style.display='none';
  document.getElementById('aiCard').style.display='none';
  const fd=new FormData();
  fd.append('image',selectedFile);
  fd.append('skip_removebg',document.getElementById('skipRmbg').checked?'1':'0');
  try{
    const r=await fetch('/api/image-bg/removebg-only?key='+KEY,{method:'POST',body:fd});
    const j=await r.json();
    if(!j.ok){showErr('errBox',j.error||'去背失敗');document.getElementById('emptyState').style.display='flex';}
    else{
      transparentUrl=j.transparent_url;
      document.getElementById('cutoutImg').src=j.transparent_url;
      document.getElementById('aiPrompt').value=document.getElementById('prompt').value;
      document.getElementById('step2Card').style.display='block';
      document.getElementById('aiCard').style.display='block';
    }
  }catch(e){showErr('errBox','連線錯誤: '+e.message);document.getElementById('emptyState').style.display='flex';}
  btn.disabled=false;btn.textContent='Step 1 — 去背 ＆ 載入';
}

// ─ Step 2: quick composite ─
let qcBusy=false;
async function qc(el,bgType){
  if(!transparentUrl||qcBusy)return;
  document.querySelectorAll('.swatch').forEach(s=>s.classList.remove('active'));
  el.classList.add('active');
  const pa=document.getElementById('previewArea');
  pa.style.display='block';
  document.getElementById('prevLoad').style.display='flex';
  qcBusy=true;
  try{
    const r=await fetch('/api/image-bg/quick-composite?key='+KEY,{
      method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({transparent_url:transparentUrl,bg_type:bgType})
    });
    const j=await r.json();
    if(j.ok){
      document.getElementById('previewImg').src=j.url;
      document.getElementById('quickDl').href=j.url;
    }
  }catch(e){}
  document.getElementById('prevLoad').style.display='none';
  qcBusy=false;
}
function clearPreview(){
  document.querySelectorAll('.swatch').forEach(s=>s.classList.remove('active'));
  document.getElementById('previewArea').style.display='none';
}

// ─ Step 3: AI generate ─
async function doAi(){
  if(!transparentUrl)return;
  const p=document.getElementById('aiPrompt').value.trim()||document.getElementById('prompt').value.trim();
  const btn=document.getElementById('aiBtn');
  btn.disabled=true;btn.textContent='生成中…';
  hideErr('aiErr');
  document.getElementById('aiResult').style.display='none';
  try{
    const r=await fetch('/api/image-bg/ai-generate?key='+KEY,{
      method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({transparent_url:transparentUrl,prompt:p,composite:document.getElementById('composite').checked})
    });
    const j=await r.json();
    if(j.ok){
      document.getElementById('aiOrigImg').src=transparentUrl;
      document.getElementById('aiResultImg').src=j.url;
      document.getElementById('aiDl').href=j.url;
      document.getElementById('aiResult').style.display='block';
    }else showErr('aiErr',j.error||'AI 生成失敗');
  }catch(e){showErr('aiErr','連線錯誤: '+e.message);}
  btn.disabled=false;btn.textContent='生成';
}
function resetAi(){document.getElementById('aiResult').style.display='none';hideErr('aiErr');}
</script>
</body>
</html>"""


@products_bp.route("/admin/image-bg")
def admin_image_bg():
    ok, key = check_auth()
    if not ok:
        return render_template_string(LOGIN_HTML, next="/admin/image-bg", error=None)
    profiles = _bp_all()
    return render_template_string(IMAGE_BG_HTML, key=key, profiles=profiles)


@products_bp.route("/api/image-bg/removebg-only", methods=["POST"])
def api_image_bg_removebg():
    import sys, io
    ok, _ = auth_required()
    if not ok:
        return jsonify({"ok": False, "error": "unauthorized"}), 403
    f = request.files.get("image")
    if not f:
        return jsonify({"ok": False, "error": "未上傳圖片"})
    skip = request.form.get("skip_removebg", "0") == "1"
    try:
        img_bytes = f.read()
        if skip:
            transparent = img_bytes
        else:
            transparent = _removebg_api(img_bytes)
            if not transparent:
                return jsonify({"ok": False, "error": "remove.bg 去背失敗，請確認 API 金鑰"})
        ts = int(time.time())
        fname = f"image-bg/transparent_{ts}.png"
        pub_url, err = upload_image_to_supabase(fname, transparent, "image/png")
        if not pub_url:
            return jsonify({"ok": False, "error": f"上傳去背圖失敗：{err}"})
        return jsonify({"ok": True, "transparent_url": pub_url})
    except Exception as e:
        print(f"[removebg-only] {e}", file=sys.stderr)
        return jsonify({"ok": False, "error": str(e)})


# preset bg definitions for quick-composite
_BG_PRESETS = {
    "white":  {"solid": (255, 255, 255)},
    "grey":   {"solid": (242, 242, 242)},
    "cream":  {"solid": (245, 240, 232)},
    "blue":   {"solid": (232, 242, 248)},
    "dark":   {"solid": (30, 30, 30)},
    "warm":   {"gradient": [(255, 236, 210), (252, 182, 159)]},
    "cool":   {"gradient": [(168, 237, 234), (254, 214, 227)]},
    "green":  {"gradient": [(212, 252, 121), (150, 230, 161)]},
    "purple": {"gradient": [(102, 126, 234), (118, 75, 162)]},
    "ocean":  {"gradient": [(44, 62, 80), (76, 161, 175)]},
}

def _make_preset_bg(bg_type, size=1024):
    from PIL import Image
    preset = _BG_PRESETS.get(bg_type, {"solid": (255, 255, 255)})
    if "solid" in preset:
        bg = Image.new("RGB", (size, size), preset["solid"])
    else:
        c1, c2 = preset["gradient"]
        bg = Image.new("RGB", (size, size))
        from PIL import ImageDraw
        draw = ImageDraw.Draw(bg)
        for i in range(size):
            t = i / size
            r = int(c1[0] + (c2[0] - c1[0]) * t)
            g = int(c1[1] + (c2[1] - c1[1]) * t)
            b = int(c1[2] + (c2[2] - c1[2]) * t)
            draw.line([(0, i), (size, i)], fill=(r, g, b))
    return bg


@products_bp.route("/api/image-bg/quick-composite", methods=["POST"])
def api_image_bg_quick():
    import sys, io, requests as _req
    ok, _ = auth_required()
    if not ok:
        return jsonify({"ok": False, "error": "unauthorized"}), 403
    data = request.get_json(force=True)
    transparent_url = data.get("transparent_url", "")
    bg_type = data.get("bg_type", "white")
    if not transparent_url:
        return jsonify({"ok": False, "error": "缺少 transparent_url"})
    try:
        from PIL import Image
        import io as _io
        r = _req.get(transparent_url, timeout=20)
        fg = Image.open(_io.BytesIO(r.content)).convert("RGBA")
        size = max(fg.size)
        size = max(size, 800)
        bg = _make_preset_bg(bg_type, size).convert("RGBA")
        # centre product on background
        offset = ((size - fg.width) // 2, (size - fg.height) // 2)
        bg.paste(fg, offset, fg)
        out = _io.BytesIO()
        bg.convert("RGB").save(out, format="JPEG", quality=92)
        out.seek(0)
        ts = int(time.time())
        fname = f"image-bg/quick_{bg_type}_{ts}.jpg"
        pub_url, err = upload_image_to_supabase(fname, out.read(), "image/jpeg")
        if not pub_url:
            return jsonify({"ok": False, "error": f"上傳失敗：{err}"})
        return jsonify({"ok": True, "url": pub_url})
    except Exception as e:
        print(f"[quick-composite] {e}", file=sys.stderr)
        return jsonify({"ok": False, "error": str(e)})


@products_bp.route("/api/image-bg/ai-generate", methods=["POST"])
def api_image_bg_ai():
    import sys, io, requests as _req
    ok, _ = auth_required()
    if not ok:
        return jsonify({"ok": False, "error": "unauthorized"}), 403
    data = request.get_json(force=True)
    transparent_url = data.get("transparent_url", "")
    prompt_raw = data.get("prompt", "").strip()
    composite = data.get("composite", False)
    if not transparent_url:
        return jsonify({"ok": False, "error": "缺少 transparent_url"})
    if not OPENAI_API_KEY:
        return jsonify({"ok": False, "error": "未設定 OPENAI_API_KEY"})
    try:
        r = _req.get(transparent_url, timeout=20)
        transparent = r.content
        # translate / auto-generate prompt
        if prompt_raw:
            prompt_en = _translate_prompt_to_en(prompt_raw)
        else:
            prompt_en = _auto_bg_prompt("product", "general")
        result = None
        if composite:
            result = _gpt_image2_composite_bg(transparent, prompt_en)
        else:
            result = _gpt_image2_bg(transparent, prompt_en)
        if not result:
            return jsonify({"ok": False, "error": "AI 生成失敗，請稍後再試"})
        ts = int(time.time())
        fname = f"image-bg/ai_{ts}.jpg"
        pub_url, err = upload_image_to_supabase(fname, result, "image/jpeg")
        if not pub_url:
            return jsonify({"ok": False, "error": f"上傳失敗：{err}"})
        return jsonify({"ok": True, "url": pub_url})
    except Exception as e:
        print(f"[ai-generate] {e}", file=sys.stderr)
        return jsonify({"ok": False, "error": str(e)})


@products_bp.route("/admin/brand-settings")
def admin_brand_settings():
    ok, key = check_auth()
    if not ok:
        return render_template_string(LOGIN_HTML, next="/admin/brand-settings", error=None)
    profiles = _bp_all()
    return render_template_string(BRAND_SETTINGS_HTML, key=key, profiles=profiles)

@products_bp.route("/api/brand-profiles", methods=["GET"])
def api_brand_profiles_list():
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    return jsonify({"profiles": _bp_all()})

@products_bp.route("/api/brand-profiles/<brand_key>", methods=["PUT"])
def api_brand_profiles_save(brand_key):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    success = _bp_save(
        brand_key,
        data.get("name", ""),
        data.get("category", ""),
        data.get("style", ""),
        data.get("tone", ""),
        data.get("custom_prompt", ""),
        data.get("image_style", ""),
        data.get("seo_direction", ""),
        bool(data.get("enabled", True)),
        bool(data.get("bg_composite", False)),
    )
    return jsonify({"ok": success})

# ── API 路由 ──────────────────────────────────────────────────

@products_bp.route("/api/products", methods=["POST"])
def api_products_add():
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    if not url:
        return jsonify({"error": "請提供商品連結"}), 400
    platform = _detect_platform(url)
    if platform == "unknown":
        return jsonify({"error": "目前只支援 1688、淘寶、天貓 連結"}), 400
    brand = (data.get("brand") or "").strip()
    job_id = _pj_insert(url, platform, brand)
    if not job_id:
        return jsonify({"error": "建立任務失敗，請確認資料庫連線"}), 500
    return jsonify({"ok": True, "id": job_id, "platform": platform, "brand": brand})

@products_bp.route("/api/products", methods=["GET"])
def api_products_list():
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    return jsonify({"jobs": _pj_list(50)})

@products_bp.route("/api/products/<int:job_id>", methods=["GET"])
def api_products_get(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    job = _pj_get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    return jsonify(job)

@products_bp.route("/api/products/<int:job_id>", methods=["DELETE"])
def api_products_delete(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    _pj_delete(job_id)
    return jsonify({"ok": True})

@products_bp.route("/api/products/<int:job_id>/images", methods=["POST"])
def api_products_images(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    urls = data.get("urls", [])
    cleaned = _clean_images(urls)
    _pj_update(job_id, raw_images=json.dumps(cleaned, ensure_ascii=False))
    return jsonify({"ok": True, "count": len(cleaned)})

@products_bp.route("/api/products/<int:job_id>", methods=["PUT"])
def api_products_update(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    allowed = {"ai_name", "ai_desc", "ai_keywords", "category", "price_min", "price_max",
               "shopee_title", "website_name", "features", "seo_desc", "listing_status"}
    fields = {k: v for k, v in data.items() if k in allowed}
    if "faq" in data:
        fields["faq"] = json.dumps(data["faq"], ensure_ascii=False)
    if not fields:
        return jsonify({"error": "no valid fields"}), 400
    _pj_update(job_id, **fields)
    return jsonify({"ok": True})

@products_bp.route("/api/products/<int:job_id>/regenerate-copy", methods=["POST"])
def api_products_regenerate_copy(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    job = _pj_get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    if not job.get("raw_title") and not job.get("raw_desc"):
        return jsonify({"error": "尚無原始商品資料，無法生成文案"}), 400
    data = request.get_json(silent=True) or {}
    up_param = request.args.get("use_pipeline") or data.get("use_pipeline")
    use_pipeline = None
    if up_param is not None:
        use_pipeline = str(up_param).lower() in ("1", "true", "yes")
    threading.Thread(target=_run_ai_rewrite_for_job, args=(job_id, use_pipeline), daemon=True).start()
    return jsonify({"ok": True, "pipeline": use_pipeline})

@products_bp.route("/api/products/<int:job_id>/set-main-image", methods=["POST"])
def api_products_set_main_image(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    if not url:
        return jsonify({"error": "缺少圖片網址"}), 400
    _pj_update(job_id, main_image=url)
    return jsonify({"ok": True})

@products_bp.route("/api/products/<int:job_id>/process-images", methods=["POST"])
def api_products_process_images(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    job = _pj_get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    data = request.get_json(silent=True) or {}
    selected_urls = data.get("urls", [])
    target_imgs = selected_urls if selected_urls else job.get("raw_images", [])
    if not target_imgs:
        return jsonify({"error": "沒有圖片可處理"}), 400
    update_fields = {"img_status": "pending_images"}
    if selected_urls:
        update_fields["raw_images"] = json.dumps(selected_urls, ensure_ascii=False)
    _pj_update(job_id, **update_fields)
    return jsonify({"ok": True})

@products_bp.route("/api/products/<int:job_id>/select-images", methods=["POST"])
def api_products_select_images(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    job = _pj_get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    data = request.get_json(force=True)
    urls = [u.strip() for u in (data.get("urls") or []) if u.strip()]
    _pj_update(job_id, raw_images=json.dumps(urls, ensure_ascii=False))
    return jsonify({"ok": True, "count": len(urls)})


@products_bp.route("/api/products/<int:job_id>/images/zip-selected", methods=["POST"])
def api_products_images_zip_selected(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    job = _pj_get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    data = request.get_json(force=True)
    items = data.get("items")
    if not items:
        old_urls = data.get("urls", [])
        items = [{"url": u, "cat": "img", "idx": i+1} for i, u in enumerate(old_urls)]
    if not items:
        return jsonify({"error": "沒有選取圖片"}), 400
    import io, zipfile
    buf = io.BytesIO()
    downloaded = 0
    cat_counter = {}
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for item in items:
            img_url = item.get("url", "")
            cat     = item.get("cat", "img")
            if not img_url: continue
            cat_counter[cat] = cat_counter.get(cat, 0) + 1
            idx = cat_counter[cat]
            try:
                req = urllib.request.Request(img_url, headers={"User-Agent": "Mozilla/5.0", "Referer": "https://www.1688.com/"})
                with urllib.request.urlopen(req, timeout=15) as r:
                    img_data = r.read()
                # 統一轉 JPG（處理 WEBP / PNG / 任何格式）
                try:
                    from PIL import Image as _PILImg
                    _img = _PILImg.open(io.BytesIO(img_data)).convert("RGB")
                    _out = io.BytesIO()
                    _img.save(_out, format="JPEG", quality=88, optimize=True)
                    img_data = _out.getvalue()
                except Exception:
                    pass  # PIL 失敗就保留原始 bytes
                zf.writestr(f"{cat}_{idx:02d}.jpg", img_data)
                downloaded += 1
            except Exception:
                pass
    if downloaded == 0:
        return jsonify({"error": "圖片下載失敗"}), 500
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="application/zip",
        headers={"Content-Disposition": f'attachment; filename="product_{job_id}_images.zip"'})




def _get_cjk_font(size=22):
    """取得支援中文的字型，找不到就下載 NotoSans"""
    from PIL import ImageFont
    import os, urllib.request as _ureq
    paths = [
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJKsc-Regular.otf",
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/arphic/uming.ttc",
        "/tmp/NotoSansCJK.ttc",
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                return ImageFont.truetype(p, size)
            except Exception:
                pass
    cache = "/tmp/NotoSansCJK.ttc"
    if not os.path.exists(cache):
        try:
            _ureq.urlretrieve(
                "https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTF/TraditionalChinese/NotoSansCJKtc-Regular.otf",
                cache
            )
            return ImageFont.truetype(cache, size)
        except Exception:
            pass
    return ImageFont.load_default()


def _sample_bg_info(img_pil, x1, y1, x2, y2):
    """返回 (variance, avg_brightness, edge_color)。純 Pillow，不依賴 numpy。
    variance  = RGB 各 channel stddev 的 RMS（等效 np.std 行為）
    brightness= 灰階平均亮度
    edge_color= bbox 外圍 3px 採樣均色（避開文字像素）
    """
    from PIL import ImageStat
    W, H = img_pil.size
    region = img_pil.crop((x1, y1, x2, y2))
    rw, rh = region.size
    if rw <= 0 or rh <= 0:
        return 0.0, 200.0, (255, 255, 255)

    stat_rgb = ImageStat.Stat(region)
    variance = (sum(v * v for v in stat_rgb.stddev) / len(stat_rgb.stddev)) ** 0.5

    stat_l = ImageStat.Stat(region.convert("L"))
    avg_brightness = stat_l.mean[0]

    # 背景色：從 bbox 外圍 3px 採樣，避免文字像素干擾
    ox1, oy1 = max(0, x1 - 3), max(0, y1 - 3)
    ox2, oy2 = min(W, x2 + 3), min(H, y2 + 3)
    stat_outer = ImageStat.Stat(img_pil.crop((ox1, oy1, ox2, oy2)))
    edge_color = tuple(int(v) for v in stat_outer.mean[:3])

    return variance, avg_brightness, edge_color


def _fit_and_draw(draw, xy, text, font, text_color=(15, 15, 15)):
    """貼上翻譯文字，支援自訂文字顏色"""
    draw.text(xy, text, fill=text_color, font=font)


def _translate_images_job(job_id, img_urls):
    """
    圖片翻譯 v2：結構化 OCR → 語言過濾 → 依 variance 分流 Pillow/Stability → 動態字型
    只翻簡體中文，日文/英文/數字一律跳過。
    variance < 20 → pillow | 20-50 → pillow+pad | >= 50 → stability
    brightness < 128 → 白字 | >= 128 → 深色字
    """
    import io, base64, re as _re3, traceback
    try:
        import requests as _req
        from PIL import Image, ImageDraw
    except ImportError as e:
        print(f"[translate] import error: {e}")
        _pj_update(job_id, translate_status="failed")
        return

    STABILITY_KEY = os.getenv("STABILITY_API_KEY", "")
    if not STABILITY_KEY:
        print("[translate] STABILITY_API_KEY not set — complex blocks will fallback to Pillow")

    OCR_PROMPT = (
        '請偵測圖片中所有文字區塊，直接輸出 JSON（不要 markdown）。\n'
        '格式：{"blocks":[{"text":"原文","language":"zh-CN/zh-TW/ja/en/number/mixed",'
        '"should_translate":true/false,"translated_text":"繁體（false時留空字串）",'
        '"bbox":[x1,y1,x2,y2],"font_role":"title/subtitle/body/label/number",'
        '"background_type":"white/solid/complex"}]}\n'
        '規則（嚴格遵守）：\n'
        '1. language分類：zh-CN=簡體中文漢字，zh-TW=繁體中文，ja=日文（含ひらがな/カタカナ），en=英文，number=純數字百分比\n'
        '2. should_translate只有純zh-CN才true，其他全部false\n'
        '3. translated_text：簡體→台灣繁體，混合文字只翻中文部分，英文/數字原樣保留\n'
        '4. bbox：[左%,上%,右%,下%]，圖片寬高各為100\n'
        '5. font_role：title=主標題大字，subtitle=副標題，body=說明文字，label=小標籤，number=數字\n'
        '6. background_type：white=白色背景，solid=純色背景，complex=照片/漸層/複雜背景\n'
        '範例（此圖色卡）：\n'
        '- 「海盐蓝」→zh-CN,true,「海鹽藍」,title,white\n'
        '- 「クリームホワイト」→ja,false,\"\",label,white\n'
        '- 「Shading：80%」→en,false,\"\",number,white\n'
        '- 「遮光率：80%」→zh-CN,true,「遮光率：80%」,label,white（數字保留不翻）\n'
        '只輸出JSON，不要說明文字。'
    )

    def _pct2px(pct, dim):
        return max(0, min(dim, int(pct / 100 * dim)))

    def _sample_bg(base_img, x1, y1, x2, y2):
        W, H = base_img.size
        samples = []
        for px, py in [(max(0,x1-8),max(0,y1-8)), (min(W-1,x2+8),max(0,y1-8)),
                       (max(0,x1-8),min(H-1,y2+8)), (min(W-1,x2+8),min(H-1,y2+8))]:
            try:
                s = base_img.getpixel((px, py))
                samples.append(s[:3] if len(s) > 3 else s)
            except Exception:
                pass
        if not samples:
            return (255, 255, 255)
        return tuple(sum(s[i] for s in samples) // len(samples) for i in range(3))

    def _fit_and_draw(draw, text, x1, y1, x2, y2, role, text_color=(15, 15, 15)):
        from PIL import ImageFont
        box_w, box_h = x2 - x1, y2 - y1
        start_sz = int(box_h * (0.80 if role == "title" else 0.75))
        start_sz = max(10, min(start_sz, 120))
        font, chosen_sz = None, start_sz
        for sz in range(start_sz, 7, -2):
            try:
                f = _get_cjk_font(sz)
                try:
                    bb = f.getbbox(text)
                    tw = bb[2] - bb[0]
                except Exception:
                    tw = len(text) * sz * 0.65
                if tw <= box_w * 1.05:
                    font, chosen_sz = f, sz
                    break
            except Exception:
                pass
        if font is None:
            font = _get_cjk_font(10)
        try:
            bb = font.getbbox(text)
            tw, th = bb[2]-bb[0], bb[3]-bb[1]
        except Exception:
            tw, th = len(text)*chosen_sz, chosen_sz
        # Horizontal alignment
        if role in ("title", "subtitle"):
            tx = x1 + max(0, (box_w - tw) // 2)
        else:
            tx = x1 + 4
        ty = y1 + max(0, (box_h - th) // 2)
        draw.text((tx, ty), text, fill=text_color, font=font)

    # ── main loop ─────────────────────────────────────────────
    translated_urls = []
    stats = dict(total=0, translated=0, skip_ja=0, skip_en=0, skip_num=0,
                 skip_tw=0, pillow=0, stab=0, stab_fail=0)

    for img_idx, url in enumerate(img_urls):
        print(f"[translate] {img_idx+1}/{len(img_urls)} {url[:70]}")
        try:
            # 1. Download
            req = urllib.request.Request(url, headers={"User-Agent":"Mozilla/5.0","Referer":"https://www.1688.com/"})
            with urllib.request.urlopen(req, timeout=20) as r:
                img_bytes = r.read()
            media_type = "image/jpeg"
            if img_bytes[:8] == b'\x89PNG\r\n\x1a\n': media_type = "image/png"
            elif img_bytes[:4] == b'RIFF': media_type = "image/webp"

            # 2. Claude Vision OCR
            img_b64 = base64.standard_b64encode(img_bytes).decode()
            claude_resp = urllib.request.urlopen(
                urllib.request.Request(
                    "https://api.anthropic.com/v1/messages",
                    data=json.dumps({
                        "model": "claude-haiku-4-5-20251001",
                        "max_tokens": 3000,
                        "messages": [{"role":"user","content":[
                            {"type":"image","source":{"type":"base64","media_type":media_type,"data":img_b64}},
                            {"type":"text","text":OCR_PROMPT}
                        ]}]
                    }).encode(),
                    headers={"x-api-key":ANTHROPIC_API_KEY,"anthropic-version":"2023-06-01","content-type":"application/json"},
                    method="POST"
                ), timeout=45
            )
            raw_text = json.loads(claude_resp.read())["content"][0]["text"].strip()
            m = _re3.search(r'\{[\s\S]*\}', raw_text)
            if not m:
                print("  [OCR] no JSON"); translated_urls.append(url); continue

            blocks = json.loads(m.group()).get("blocks", [])
            stats['total'] += len(blocks)
            print(f"  [OCR] {len(blocks)} 區塊")

            # 3. Filter
            to_do = []
            for b in blocks:
                lang = b.get("language","")
                if not b.get("should_translate") or not b.get("translated_text","") or lang != "zh-CN":
                    if lang=="ja": stats['skip_ja']+=1
                    elif lang=="en": stats['skip_en']+=1
                    elif lang=="number": stats['skip_num']+=1
                    elif lang=="zh-TW": stats['skip_tw']+=1
                    continue
                stats['translated']+=1
                to_do.append(b)

            if not to_do:
                print("  [OCR] 無需翻譯"); translated_urls.append(url); continue

            # 4. Convert bbox % → px，同時計算 variance / brightness
            from PIL import Image
            img_pil = Image.open(io.BytesIO(img_bytes)).convert("RGB")
            W, H = img_pil.size
            for b in to_do:
                bx = b.get("bbox", [0, 0, 100, 100])
                b['_px'] = (_pct2px(bx[0],W), _pct2px(bx[1],H), _pct2px(bx[2],W), _pct2px(bx[3],H))

            # 5. 依 variance 分類（不盲信 Claude background_type）
            # var < 20 → pillow | 20-50 → pillow+pad | >= 50 → stability
            simple  = []
            complex_ = []
            for b in to_do:
                x1, y1, x2, y2 = b['_px']
                variance, brightness, edge_color = _sample_bg_info(img_pil, x1, y1, x2, y2)
                b['_variance'] = variance
                b['_brightness'] = brightness
                b['_edge_color'] = edge_color
                b['_text_color'] = (255, 255, 255) if brightness < 128 else (15, 15, 15)
                color_label = "white" if brightness < 128 else "dark"
                bg_hint = b.get("background_type", "white")

                if variance < 20:
                    b['_pad'] = 2; simple.append(b)
                    print(f"    [{b.get('text','')[:12]}] var={variance:.1f} bright={brightness:.0f} ({bg_hint}) → pillow | {color_label}")
                elif variance < 50:
                    b['_pad'] = 6; simple.append(b)
                    print(f"    [{b.get('text','')[:12]}] var={variance:.1f} bright={brightness:.0f} ({bg_hint}) → pillow+pad | {color_label}")
                else:
                    complex_.append(b)
                    print(f"    [{b.get('text','')[:12]}] var={variance:.1f} bright={brightness:.0f} ({bg_hint}) → stability | {color_label}")

            stability_saved = len(simple)
            print(f"  [classify] pillow={len(simple)} stability={len(complex_)} | 本張省下 {stability_saved}/{len(to_do)} 次 Stability 呼叫")

            result_img = img_pil.copy()

            # 6a. Pillow cover (white/solid/simple)
            # pad = 2~6px（來自 variance 分類），不用 bbox 百分比避免巨大方塊
            if simple:
                draw = ImageDraw.Draw(result_img)
                for b in simple:
                    x1,y1,x2,y2 = b['_px']
                    role = b.get("font_role","body")
                    pad = b.get('_pad', 4)
                    rx1,ry1 = max(0,x1-pad), max(0,y1-pad)
                    rx2,ry2 = min(W,x2+pad), min(H,y2+pad)
                    bg = b.get('_edge_color') or _sample_bg(img_pil, rx1,ry1,rx2,ry2)
                    draw.rectangle([rx1,ry1,rx2,ry2], fill=bg)
                    _fit_and_draw(draw, b.get("translated_text",""), rx1,ry1,rx2,ry2, role, b['_text_color'])
                    stats['pillow']+=1
                    clabel = "white" if b.get('_brightness', 200) < 128 else "dark"
                    print(f"    Pillow [{role}]: '{b.get('text','')}' → '{b.get('translated_text','')}' | {clabel}")

            # 6b. 無 STABILITY_KEY：complex 也走 Pillow fallback
            if complex_ and not STABILITY_KEY:
                draw_fb = ImageDraw.Draw(result_img)
                for b in complex_:
                    x1,y1,x2,y2 = b['_px']
                    role = b.get("font_role","body")
                    pad = 6
                    rx1,ry1 = max(0,x1-pad),max(0,y1-pad)
                    rx2,ry2 = min(W,x2+pad),min(H,y2+pad)
                    bg = b.get('_edge_color') or _sample_bg(img_pil,rx1,ry1,rx2,ry2)
                    draw_fb.rectangle([rx1,ry1,rx2,ry2],fill=bg)
                    _fit_and_draw(draw_fb, b.get("translated_text",""), rx1,ry1,rx2,ry2, role, b['_text_color'])
                    stats['pillow']+=1
                complex_ = []
                print(f"  [Stability] skipped (no key) — fell back to Pillow for all blocks")

            # 6c. Stability inpaint (complex photo backgrounds only)
            if complex_:
                from PIL import ImageDraw as _ID2
                mask = Image.new("L",(W,H),0)
                dm = _ID2.Draw(mask)
                for b in complex_:
                    x1,y1,x2,y2 = b['_px']
                    mp = 8  # mask padding: 略大一點讓 inpaint 邊緣乾淨
                    dm.rectangle([max(0,x1-mp),max(0,y1-mp),min(W,x2+mp),min(H,y2+mp)],fill=255)

                ib = io.BytesIO(); result_img.save(ib,"PNG"); ib.seek(0)
                mb = io.BytesIO(); mask.convert("RGB").save(mb,"PNG"); mb.seek(0)
                inpaint_ok = False
                try:
                    sr = _req.post(
                        "https://api.stability.ai/v2beta/stable-image/edit/erase",
                        headers={"Authorization":f"Bearer {STABILITY_KEY}","Accept":"image/*"},
                        files={"image":("i.png",ib.getvalue(),"image/png"),
                               "mask":("m.png",mb.getvalue(),"image/png")},
                        data={"output_format":"png"}, timeout=60
                    )
                    if sr.status_code == 200:
                        result_img = Image.open(io.BytesIO(sr.content)).convert("RGB")
                        stats['stab']+=len(complex_)
                        inpaint_ok = True
                        print(f"    Stability: {len(complex_)} 區塊")
                    else:
                        print(f"    Stability {sr.status_code} → Pillow fallback")
                        stats['stab_fail']+=len(complex_)
                except Exception as se:
                    print(f"    Stability error: {se} → Pillow fallback")
                    stats['stab_fail']+=len(complex_)

                if not inpaint_ok:
                    draw_fb = ImageDraw.Draw(result_img)
                    for b in complex_:
                        x1,y1,x2,y2 = b['_px']
                        bg = b.get('_edge_color') or _sample_bg(img_pil,x1,y1,x2,y2)
                        draw_fb.rectangle([x1,y1,x2,y2],fill=bg)
                    stats['pillow']+=len(complex_)

                # Overlay text for complex（直接用原始 bbox，Stability 已清除原文）
                draw_c = ImageDraw.Draw(result_img)
                for b in complex_:
                    x1,y1,x2,y2 = b['_px']
                    role = b.get("font_role","body")
                    _fit_and_draw(draw_c, b.get("translated_text",""),
                                  x1,y1,x2,y2, role,
                                  b.get('_text_color', (15,15,15)))

            # 7. Debug stats
            print(f"  [stats] OCR:{stats['total']} 翻:{stats['translated']} 跳日:{stats['skip_ja']} 跳英:{stats['skip_en']} 跳數:{stats['skip_num']} 跳繁:{stats['skip_tw']} Pillow:{stats['pillow']} Stab:{stats['stab']} Stab失敗:{stats['stab_fail']}")

            # 8. Upload
            out = io.BytesIO()
            result_img.save(out,"JPEG",quality=93)
            fname = f"translated_{job_id}_{img_idx+1}.jpg"
            turl,_ = upload_image_to_supabase(fname, out.getvalue(), "image/jpeg")
            if not turl:
                turl,_ = upload_image_to_github(fname, out.getvalue())
            translated_urls.append(turl or url)
            print(f"  [done] {(turl or url)[:70]}")

        except Exception as e:
            print(f"  [ERROR] {e}")
            traceback.print_exc()
            translated_urls.append(url)

    _pj_update(job_id,
               translated_images=json.dumps(translated_urls, ensure_ascii=False),
               translate_status="done")
    print(f"[translate] 完成 {len(translated_urls)}/{len(img_urls)} 張，stats={stats}")

@products_bp.route("/api/products/<int:job_id>/translate-images", methods=["POST"])
def api_translate_images(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    if not os.getenv("STABILITY_API_KEY"):
        return jsonify({"error": "STABILITY_API_KEY 未設定"}), 500
    job = _pj_get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    data = request.get_json(force=True)
    urls = [u.strip() for u in (data.get("urls") or []) if u.strip()]
    if not urls:
        return jsonify({"error": "請選取要翻譯的圖片"}), 400
    _pj_update(job_id, translate_status="processing")
    import threading
    threading.Thread(target=_translate_images_job, args=(job_id, urls), daemon=True).start()
    return jsonify({"ok": True, "count": len(urls)})


@products_bp.route("/api/products/<int:job_id>/upload-translated", methods=["POST"])
def api_upload_translated(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    job = _pj_get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    img_type = request.form.get("type", "main")
    if img_type not in ("main", "detail", "sku"):
        img_type = "main"
    files = request.files.getlist("files")
    if not files:
        return jsonify({"error": "沒有上傳圖片"}), 400
    uploaded_urls = []
    for i, f in enumerate(files[:20]):
        img_bytes = f.read()
        if not img_bytes: continue
        fname_orig = f.filename or f"img_{i+1}.jpg"
        ext = fname_orig.rsplit(".", 1)[-1].lower() if "." in fname_orig else "jpg"
        if ext not in ("jpg", "jpeg", "png", "webp"): ext = "jpg"
        filename = f"products/{job_id}_tr_{img_type}_{int(time.time())}_{i+1}.{ext}"
        pub_url, _ = upload_image_to_supabase(filename, img_bytes, f.content_type or "image/jpeg")
        if pub_url:
            uploaded_urls.append(pub_url)
    if not uploaded_urls:
        return jsonify({"error": "上傳失敗，請檢查 Supabase 設定"}), 500
    pi = job.get("product_images") or {}
    tr_key = f"tr_{img_type}_images"
    pi[tr_key] = (pi.get(tr_key) or []) + uploaded_urls
    _pj_update(job_id, product_images=json.dumps(pi, ensure_ascii=False))
    return jsonify({"ok": True, "added": len(uploaded_urls), "urls": uploaded_urls, "type": img_type})


def _download_image_robust(url):
    """Download image with multiple strategies for CDN hotlink protection."""
    import sys
    if not url:
        return None
    if url.startswith("//"):
        url = "https:" + url
    headers_list = [
        {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
         "Referer": "https://www.1688.com/", "Accept": "image/webp,image/apng,image/*,*/*"},
        {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
         "Referer": "https://detail.tmall.com/", "Accept": "image/*,*/*"},
        {"User-Agent": "Mozilla/5.0", "Accept": "*/*"},
    ]
    try:
        import requests as _req
        for hdrs in headers_list:
            try:
                r = _req.get(url, headers=hdrs, timeout=20, stream=True)
                if r.status_code == 200:
                    return r.content
            except Exception as e:
                print(f"[WhiteBG] download attempt failed: {e}", file=sys.stderr)
    except ImportError:
        pass
    # fallback urllib
    try:
        req = urllib.request.Request(url, headers=headers_list[0])
        with urllib.request.urlopen(req, timeout=20) as r:
            return r.read()
    except Exception as e:
        print(f"[WhiteBG] urllib fallback failed: {e}", file=sys.stderr)
    return None


def _whitebg_selected_job(job_id, img_urls):
    import sys
    print(f"[WhiteBG] job={job_id} urls={len(img_urls)}", file=sys.stderr)
    job = _pj_get(job_id)
    if not job:
        print(f"[WhiteBG] job {job_id} not found", file=sys.stderr)
        return
    existing = job.get("processed_images", [])
    existing_set = set(existing)
    new_processed = list(existing)
    for i, url in enumerate(img_urls[:20]):
        print(f"[WhiteBG] [{i+1}/{len(img_urls)}] downloading {url[:80]}", file=sys.stderr)
        img_bytes = _download_image_robust(url)
        if not img_bytes:
            print(f"[WhiteBG] download failed for img {i+1}", file=sys.stderr)
            continue
        print(f"[WhiteBG] downloaded {len(img_bytes)} bytes, processing...", file=sys.stderr)
        result = _process_to_white_bg(img_bytes)
        if not result:
            print(f"[WhiteBG] PIL processing failed for img {i+1}", file=sys.stderr)
            continue
        filename = f"products/{job_id}_sel_{int(time.time())}_{i}.jpg"
        pub_url, err = upload_image_to_supabase(filename, result, "image/jpeg")
        if err:
            print(f"[WhiteBG] Supabase upload failed: {err}", file=sys.stderr)
        if pub_url and pub_url not in existing_set:
            new_processed.append(pub_url)
            existing_set.add(pub_url)
            print(f"[WhiteBG] uploaded: {pub_url[:60]}", file=sys.stderr)
    added = len(new_processed) - len(existing)
    print(f"[WhiteBG] done, added {added} new images", file=sys.stderr)
    _pj_update(job_id,
               processed_images=json.dumps(new_processed, ensure_ascii=False),
               img_status="done" if new_processed else "failed")

@products_bp.route("/api/products/<int:job_id>/whitebg-selected", methods=["POST"])
def api_whitebg_selected(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    job = _pj_get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    data = request.get_json(force=True)
    urls = [u.strip() for u in (data.get("urls") or []) if u.strip()]
    if not urls:
        return jsonify({"error": "請選取圖片"}), 400
    import threading
    threading.Thread(target=_whitebg_selected_job, args=(job_id, urls), daemon=True).start()
    return jsonify({"ok": True, "count": len(urls)})


# ── 本機 Worker API ───────────────────────────────────────────

# ── 匯出 / 下載 ────────────────────────────────────────────────

def _pj_list_done():
    if not DATABASE_URL:
        return []
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id,url,platform,raw_title,raw_price,ai_name,ai_desc,ai_keywords,raw_images,processed_images,created_at FROM product_jobs WHERE status='done' ORDER BY created_at DESC"
        )
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id":r[0],"url":r[1],"platform":r[2],"raw_title":r[3],"raw_price":r[4],
                 "ai_name":r[5],"ai_desc":r[6],"ai_keywords":r[7],
                 "raw_images":json.loads(r[8] or "[]"),
                 "processed_images":json.loads(r[9] or "[]"),
                 "created_at":r[10]} for r in rows]
    except Exception:
        return []

def _export_csv(jobs):
    import csv, io
    from flask import Response
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["平台","AI商品名稱","AI商品描述","SEO關鍵字","原始標題","原始價格","來源URL","圖片URL_1","圖片URL_2","圖片URL_3"])
    for j in jobs:
        imgs = j.get("processed_images") or j.get("raw_images",[])
        w.writerow([
            j["platform"], j["ai_name"], j["ai_desc"], j["ai_keywords"],
            j["raw_title"], j["raw_price"], j["url"],
            imgs[0] if len(imgs)>0 else "",
            imgs[1] if len(imgs)>1 else "",
            imgs[2] if len(imgs)>2 else "",
        ])
    output = buf.getvalue().encode("utf-8-sig")
    return Response(output, mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=products.csv"})

def _export_xlsx(jobs):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import openpyxl.utils
    except ImportError:
        return jsonify({"error": "openpyxl 未安裝"}), 500
    import io
    from flask import Response
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品資料"
    headers = ["平台","AI商品名稱","AI商品描述","SEO關鍵字","原始標題","原始價格","來源URL","圖片URL_1","圖片URL_2","圖片URL_3"]
    hfill = PatternFill("solid", fgColor="1a1a1a")
    hfont = Font(color="FFFFFF", bold=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    col_widths = [10, 35, 60, 30, 35, 12, 55, 55, 55, 55]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for ri, j in enumerate(jobs, 2):
        imgs = j.get("processed_images") or j.get("raw_images",[])
        values = [
            j["platform"], j["ai_name"], j["ai_desc"], j["ai_keywords"],
            j["raw_title"], j["raw_price"], j["url"],
            imgs[0] if len(imgs)>0 else "",
            imgs[1] if len(imgs)>1 else "",
            imgs[2] if len(imgs)>2 else "",
        ]
        for ci, v in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[ri].height = 60
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products.xlsx"})

@products_bp.route("/api/products/export")
def api_products_export():
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    fmt = request.args.get("format", "xlsx")
    jobs = _pj_list_done()
    if not jobs:
        return jsonify({"error": "沒有已完成的商品"}), 400
    if fmt == "csv":
        return _export_csv(jobs)
    return _export_xlsx(jobs)

@products_bp.route("/api/products/<int:job_id>/images/zip")
def api_products_images_zip(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    job = _pj_get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    img_type = request.args.get("type", "processed")
    if img_type == "raw":
        imgs = job.get("raw_images", [])
    elif img_type == "sku":
        pi = job.get("product_images", {})
        skus = pi.get("sku_images", [])
        imgs = [i["src"] if isinstance(i, dict) else i for i in skus]
    else:
        imgs = job.get("processed_images") or job.get("raw_images", [])
    if not imgs:
        return jsonify({"error": "沒有圖片可下載"}), 400
    import io, zipfile, re as _re
    from flask import Response
    buf = io.BytesIO()
    downloaded = 0
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for i, img_url in enumerate(imgs[:10]):
            try:
                req = urllib.request.Request(img_url, headers={"User-Agent":"Mozilla/5.0","Referer":"https://www.1688.com/"})
                with urllib.request.urlopen(req, timeout=12) as r:
                    data = r.read()
                ext = img_url.split(".")[-1].split("?")[0].lower()
                if ext not in ("jpg","jpeg","png","webp"): ext = "jpg"
                zf.writestr(f"img_{i+1:02d}.{ext}", data)
                downloaded += 1
            except Exception:
                pass
    if downloaded == 0:
        return jsonify({"error": "圖片下載失敗"}), 500
    buf.seek(0)
    safe = _re.sub(r'[^\w]', '_', (job.get("ai_name") or "product")[:20])
    return Response(buf.getvalue(), mimetype="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{safe}_images.zip"'})

@products_bp.route("/api/products/proxy-img")
def api_proxy_img():
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    url = request.args.get("url", "").strip()
    if not url:
        return jsonify({"error": "missing url"}), 400
    allowed = ["alicdn.com", "tbcdn.cn", "aliimg.com", "taobao.com", "1688.com", "supabase.co", "githubusercontent.com"]
    if not any(d in url for d in allowed):
        return jsonify({"error": "domain not allowed"}), 403
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Referer": "https://www.1688.com/"
        })
        with urllib.request.urlopen(req, timeout=15) as r:
            data = r.read()
            ct = r.headers.get("Content-Type", "image/jpeg")
    except Exception as e:
        return jsonify({"error": str(e)}), 502
    filename = url.split("/")[-1].split("?")[0] or "image.jpg"
    from flask import Response
    resp = Response(data, content_type=ct)
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

# ── 本機 Worker API ───────────────────────────────────────────

@products_bp.route("/api/products/pending", methods=["GET"])
def api_products_pending():
    """本機 Worker 輪詢：取得待爬取的任務列表。"""
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    if not DATABASE_URL:
        return jsonify({"jobs": []})
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id, url, platform, raw_images, img_status FROM product_jobs"
            " WHERE status='pending' OR img_status='pending_images'"
            " ORDER BY created_at ASC LIMIT 10"
        )
        rows = cur.fetchall(); cur.close(); conn.close()
        jobs = []
        for r in rows:
            job_id_r, url, platform, raw_imgs_json, img_status = r
            if img_status == "pending_images":
                jobs.append({"id": job_id_r, "url": url, "platform": platform,
                             "mode": "images_only",
                             "raw_images": json.loads(raw_imgs_json or "[]")})
            else:
                jobs.append({"id": job_id_r, "url": url, "platform": platform})
        return jsonify({"jobs": jobs})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@products_bp.route("/api/products/from-extension", methods=["POST"])
def api_products_from_extension():
    """Chrome Extension 直送：不需要 Worker，直接存圖 + 觸發 AI 改寫。"""
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    url      = (data.get("url") or "").strip()
    platform = data.get("platform", "1688")
    brand    = data.get("brand", "")
    title    = data.get("title", "")
    product_imgs = data.get("product_images", {})
    if not url:
        return jsonify({"error": "缺少 url"}), 400
    job_id = _pj_insert(url, platform, brand)
    if not job_id:
        return jsonify({"error": "DB error"}), 500
    main_srcs = [i["src"] if isinstance(i, dict) else i for i in product_imgs.get("main_images", [])]
    # 核心欄位先更新（確保 status 改變、AI thread 能跑）
    _pj_update(job_id,
        status    = "scraping",
        raw_title = title,
        raw_images= json.dumps(main_srcs, ensure_ascii=False),
    )
    # product_images 欄位分開更新（若 column 不存在不影響主流程）
    _pj_update(job_id, product_images=json.dumps(product_imgs, ensure_ascii=False))
    threading.Thread(target=_run_ai_rewrite_for_job, args=(job_id,), daemon=True).start()
    return jsonify({"ok": True, "id": job_id})

@products_bp.route("/api/products/<int:job_id>/scrape-result", methods=["POST"])
def api_products_scrape_result(job_id):
    """本機 Worker 回傳爬取結果，觸發 AI 改寫。"""
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}

    if data.get("error"):
        _pj_update(job_id, status="error", error_msg=f"本機爬取失敗：{data['error']}")
        return jsonify({"ok": True})

    if data.get("mode") == "images_only":
        processed = data.get("processed_images", [])
        _pj_update(job_id,
            processed_images = json.dumps(processed, ensure_ascii=False),
            img_status       = "done" if processed else "failed",
        )
        return jsonify({"ok": True})

    processed = data.get("processed_images", [])
    product_imgs = data.get("product_images", {})
    _pj_update(job_id,
        status="scraping",
        raw_title         = data.get("raw_title", ""),
        raw_desc          = data.get("raw_desc", ""),
        raw_images        = json.dumps(data.get("raw_images", []), ensure_ascii=False),
        raw_price         = data.get("raw_price", ""),
        raw_extra         = data.get("raw_extra", "{}"),
        processed_images  = json.dumps(processed, ensure_ascii=False),
        img_status        = "done" if processed else "",
        product_images    = json.dumps(product_imgs, ensure_ascii=False),
    )
    threading.Thread(target=_run_ai_rewrite_for_job, args=(job_id,), daemon=True).start()
    return jsonify({"ok": True})


# ── Store Scan API ────────────────────────────────────────────

@products_bp.route("/api/store-scan", methods=["POST"])
def api_store_scan_create():
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    if not url:
        return jsonify({"error": "請提供店鋪/分類頁連結"}), 400
    platform = _detect_platform(url)
    if platform == "unknown":
        return jsonify({"error": "目前只支援 1688、淘寶、天貓 連結"}), 400
    job_id = _ss_insert(url, platform)
    if not job_id:
        return jsonify({"error": "建立失敗，請確認資料庫連線"}), 500
    return jsonify({"ok": True, "id": job_id, "platform": platform})

@products_bp.route("/api/store-scan", methods=["GET"])
def api_store_scan_list():
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    return jsonify({"jobs": _ss_list(20)})

@products_bp.route("/api/store-scan/pending", methods=["GET"])
def api_store_scan_pending():
    """本機 Worker 輪詢：取得待掃描任務。"""
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    if not DATABASE_URL:
        return jsonify({"jobs": []})
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id, url, platform FROM store_scan_jobs WHERE status='pending' ORDER BY created_at ASC LIMIT 3"
        )
        rows = cur.fetchall(); cur.close(); conn.close()
        return jsonify({"jobs": [{"id": r[0], "url": r[1], "platform": r[2]} for r in rows]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@products_bp.route("/api/store-scan/<int:job_id>", methods=["GET"])
def api_store_scan_get(job_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    job = _ss_get(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    return jsonify(job)

@products_bp.route("/api/store-scan/<int:job_id>/result", methods=["POST"])
def api_store_scan_result(job_id):
    """本機 Worker 回傳掃描結果。"""
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    if data.get("error"):
        _ss_update(job_id, status="error", error_msg=data["error"])
        return jsonify({"ok": True})
    items = data.get("items", [])
    count = _ss_insert_items(job_id, items)
    _ss_update(job_id, status="done", item_count=count)
    return jsonify({"ok": True, "count": count})

@products_bp.route("/api/store-scan/to-queue", methods=["POST"])
def api_store_scan_to_queue():
    """將勾選商品加入 product_jobs。"""
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    item_ids = data.get("item_ids", [])
    brand = (data.get("brand") or "").strip()
    if not item_ids:
        return jsonify({"error": "請勾選商品"}), 400
    if not DATABASE_URL:
        return jsonify({"error": "資料庫未設定"}), 500
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id,url,platform FROM store_scan_items WHERE id=ANY(%s) AND added_to_queue=FALSE",
            (item_ids,)
        )
        rows = cur.fetchall(); cur.close(); conn.close()
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    added = []
    for r in rows:
        item_id, url, platform = r
        job_id = _pj_insert(url, platform, brand)
        if job_id:
            added.append({"item_id": item_id, "job_id": job_id})
    if added:
        _ss_mark_added([a["item_id"] for a in added])
    return jsonify({"ok": True, "added": len(added), "jobs": added})


# ── Product Rules DB helpers ──────────────────────────────────────

def _pr_init():
    """建立 product_rules 資料表（若不存在）。"""
    if not DATABASE_URL:
        return
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS product_rules (
                id                SERIAL PRIMARY KEY,
                brand             TEXT NOT NULL DEFAULT '',
                category          TEXT NOT NULL DEFAULT '',
                exchange_rate     NUMERIC(10,4) NOT NULL DEFAULT 4.5,
                margin_rate       NUMERIC(5,4)  NOT NULL DEFAULT 0.4,
                ad_rate           NUMERIC(5,4)  NOT NULL DEFAULT 0.1,
                sea_shipping_rate NUMERIC(5,4)  NOT NULL DEFAULT 0.05,
                tw_shipping_cost  NUMERIC(10,2) NOT NULL DEFAULT 0,
                package_cost      NUMERIC(10,2) NOT NULL DEFAULT 0,
                round_rule        TEXT NOT NULL DEFAULT 'round_10',
                created_at        FLOAT DEFAULT 0,
                updated_at        FLOAT DEFAULT 0
            )
        """)
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        import sys; print(f"[PR Init] {e}", file=sys.stderr)

def _pr_all():
    if not DATABASE_URL:
        return []
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id,brand,category,exchange_rate,margin_rate,ad_rate,"
            "sea_shipping_rate,tw_shipping_cost,package_cost,round_rule,"
            "created_at,updated_at FROM product_rules ORDER BY id DESC"
        )
        rows = cur.fetchall(); cur.close(); conn.close()
        return [{"id":r[0],"brand":r[1],"category":r[2],
                 "exchange_rate":float(r[3]),"margin_rate":float(r[4]),
                 "ad_rate":float(r[5]),"sea_shipping_rate":float(r[6]),
                 "tw_shipping_cost":float(r[7]),"package_cost":float(r[8]),
                 "round_rule":r[9],"created_at":r[10],"updated_at":r[11]} for r in rows]
    except Exception as e:
        import sys; print(f"[PR All] {e}", file=sys.stderr)
        return []

def _pr_insert(data):
    if not DATABASE_URL:
        return None
    try:
        now = time.time()
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "INSERT INTO product_rules"
            "(brand,category,exchange_rate,margin_rate,ad_rate,"
            "sea_shipping_rate,tw_shipping_cost,package_cost,round_rule,created_at,updated_at)"
            " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
            (data.get("brand",""), data.get("category",""),
             data.get("exchange_rate", 4.5), data.get("margin_rate", 0.4),
             data.get("ad_rate", 0.1), data.get("sea_shipping_rate", 0.05),
             data.get("tw_shipping_cost", 0), data.get("package_cost", 0),
             data.get("round_rule", "round_10"), now, now)
        )
        row_id = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return row_id
    except Exception as e:
        import sys; print(f"[PR Insert] {e}", file=sys.stderr)
        return None

def _pr_update(rule_id, data):
    if not DATABASE_URL:
        return False
    try:
        now = time.time()
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "UPDATE product_rules SET brand=%s,category=%s,exchange_rate=%s,"
            "margin_rate=%s,ad_rate=%s,sea_shipping_rate=%s,tw_shipping_cost=%s,"
            "package_cost=%s,round_rule=%s,updated_at=%s WHERE id=%s",
            (data.get("brand",""), data.get("category",""),
             data.get("exchange_rate", 4.5), data.get("margin_rate", 0.4),
             data.get("ad_rate", 0.1), data.get("sea_shipping_rate", 0.05),
             data.get("tw_shipping_cost", 0), data.get("package_cost", 0),
             data.get("round_rule", "round_10"), now, rule_id)
        )
        conn.commit(); cur.close(); conn.close()
        return True
    except Exception as e:
        import sys; print(f"[PR Update] {e}", file=sys.stderr)
        return False

def _pr_delete(rule_id):
    if not DATABASE_URL:
        return
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute("DELETE FROM product_rules WHERE id=%s", (rule_id,))
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        import sys; print(f"[PR Delete] {e}", file=sys.stderr)

_pr_init()


# ── Product Rules Admin HTML ──────────────────────────────────────

PRODUCT_RULES_HTML = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>商品定價規則中心</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,sans-serif;background:#f5f5f5;color:#333}
.header{background:#1a1a1a;color:#fff;padding:14px 20px;display:flex;align-items:center;gap:14px}
.header a{color:#888;text-decoration:none;font-size:14px}
.header a:hover{color:#fff}
.header-title{font-size:17px;font-weight:700;flex:1}
.wrap{max-width:1100px;margin:0 auto;padding:20px 16px}
.card{background:#fff;border-radius:14px;padding:20px;margin-bottom:16px;box-shadow:0 1px 4px rgba(0,0,0,.08)}
.top-row{display:flex;align-items:center;justify-content:space-between;margin-bottom:16px}
.top-row h3{font-size:15px;font-weight:700}
.btn-new{background:#1a1a1a;color:#fff;border:none;border-radius:10px;padding:9px 20px;font-size:13px;font-weight:700;cursor:pointer;font-family:inherit}
.btn-new:hover{background:#333}
.btn-edit{background:#e3f2fd;color:#1565c0;border:none;border-radius:8px;padding:5px 12px;font-size:12px;font-weight:600;cursor:pointer;font-family:inherit}
.btn-edit:hover{background:#bbdefb}
.btn-del{background:#fce4ec;color:#c62828;border:none;border-radius:8px;padding:5px 12px;font-size:12px;font-weight:600;cursor:pointer;font-family:inherit}
.btn-del:hover{background:#f8bbd9}
table{width:100%;border-collapse:collapse;font-size:13px}
th{text-align:left;padding:9px 10px;background:#f8f8f8;font-weight:700;color:#666;border-bottom:2px solid #eee;white-space:nowrap}
td{padding:8px 10px;border-bottom:1px solid #f0f0f0;vertical-align:middle}
tr:hover td{background:#fafafa}
.rb{display:inline-block;padding:2px 8px;border-radius:12px;font-size:11px;font-weight:700}
.rb-round_10{background:#e8f5e9;color:#2e7d32}
.rb-round_50{background:#e3f2fd;color:#1565c0}
.rb-round_100{background:#fce4ec;color:#c62828}
.rb-charm_9{background:#fff3e0;color:#e65100}
.empty{text-align:center;padding:50px;color:#bbb;font-size:14px}
.hint-card{background:#fffbf0;border:1px solid #ffe082;border-radius:12px;padding:14px 18px;margin-bottom:16px;font-size:12px;color:#795548;line-height:2}
/* Modal */
.overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:999;align-items:center;justify-content:center}
.overlay.open{display:flex}
.modal{background:#fff;border-radius:16px;padding:24px;width:520px;max-width:95vw;max-height:90vh;overflow-y:auto;box-shadow:0 8px 32px rgba(0,0,0,.18)}
.modal h3{font-size:15px;font-weight:700;margin-bottom:18px}
.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}
.full{grid-column:1/-1}
.frow{display:flex;flex-direction:column;gap:4px}
.frow label{font-size:11px;font-weight:700;color:#555}
.frow input,.frow select{border:1.5px solid #ddd;border-radius:8px;padding:8px 10px;font-size:13px;outline:none;font-family:inherit;background:#fff}
.frow input:focus,.frow select:focus{border-color:#1a1a1a}
.fhint{font-size:10px;color:#bbb;margin-top:1px}
.modal-footer{display:flex;gap:8px;justify-content:flex-end;margin-top:18px;padding-top:14px;border-top:1px solid #eee}
.btn-save{background:#1a1a1a;color:#fff;border:none;border-radius:8px;padding:9px 22px;font-size:13px;font-weight:700;cursor:pointer;font-family:inherit}
.btn-cancel{background:#f0f0f0;color:#555;border:none;border-radius:8px;padding:9px 18px;font-size:13px;cursor:pointer;font-family:inherit}
.toast{position:fixed;bottom:24px;left:50%;transform:translateX(-50%);background:#333;color:#fff;padding:9px 20px;border-radius:20px;font-size:13px;z-index:9999;opacity:0;transition:opacity .3s;pointer-events:none}
.toast.show{opacity:1}
</style>
</head>
<body>

<div class="header">
  <a href="/admin/products?key={{ key }}">&#8592; 商品搬運</a>
  <div class="header-title">商品定價規則中心</div>
  <button class="btn-new" onclick="openModal()">&#xff0b; 新增品牌規則</button>
</div>

<div class="wrap">

  <div class="hint-card">
    <strong>進位規則：</strong>
    <span class="rb rb-round_10">round_10</span> 進到10位（179&#8594;190）&nbsp;&nbsp;
    <span class="rb rb-round_50">round_50</span> 進到50（179&#8594;200）&nbsp;&nbsp;
    <span class="rb rb-round_100">round_100</span> 進到百位（179&#8594;200）&nbsp;&nbsp;
    <span class="rb rb-charm_9">charm_9</span> 尾數9（178&#8594;179）
  </div>

  <div class="card">
    <div class="top-row">
      <h3>品牌定價規則</h3>
      <span id="ruleCount" style="font-size:12px;color:#aaa"></span>
    </div>
    <div id="tableWrap"><div class="empty">載入中...</div></div>
  </div>

</div>

<!-- Modal -->
<div class="overlay" id="overlay">
  <div class="modal">
    <h3 id="modalTitle">新增品牌規則</h3>
    <input type="hidden" id="editId" value="">
    <div class="form-grid">
      <div class="frow full">
        <label>品牌名稱 *</label>
        <input type="text" id="fBrand" placeholder="例：JSIMPLE、朗德燈具">
      </div>
      <div class="frow full">
        <label>分類</label>
        <input type="text" id="fCategory" placeholder="例：高架床、燈具">
      </div>
      <div class="frow">
        <label>匯率（RMB &#8594; TWD）</label>
        <input type="number" id="fExchange" step="0.01" min="0" value="4.5">
        <span class="fhint">1 RMB = ? TWD，例：4.5</span>
      </div>
      <div class="frow">
        <label>毛利率 (%)</label>
        <input type="number" id="fMargin" step="0.1" min="0" max="100" value="40">
        <span class="fhint">例：40 代表毛利 40%</span>
      </div>
      <div class="frow">
        <label>廣告成本率 (%)</label>
        <input type="number" id="fAd" step="0.1" min="0" max="100" value="10">
        <span class="fhint">例：10 代表廣告佔售價 10%</span>
      </div>
      <div class="frow">
        <label>海運成本率 (%)</label>
        <input type="number" id="fSea" step="0.1" min="0" max="100" value="5">
        <span class="fhint">例：5 代表海運佔成本 5%</span>
      </div>
      <div class="frow">
        <label>台灣物流成本（TWD/件）</label>
        <input type="number" id="fTwShipping" step="1" min="0" value="80">
        <span class="fhint">固定金額，例：80 元</span>
      </div>
      <div class="frow">
        <label>包材成本（TWD/件）</label>
        <input type="number" id="fPackage" step="1" min="0" value="30">
        <span class="fhint">固定金額，例：30 元</span>
      </div>
      <div class="frow full">
        <label>進位規則</label>
        <select id="fRound">
          <option value="round_10">round_10 — 進到個位數10（179 &#8594; 190）</option>
          <option value="round_50">round_50 — 進到50（179 &#8594; 200）</option>
          <option value="round_100">round_100 — 進到百位（179 &#8594; 200）</option>
          <option value="charm_9">charm_9 — 尾數9（178 &#8594; 179）</option>
        </select>
      </div>
    </div>
    <div class="modal-footer">
      <button class="btn-cancel" onclick="closeModal()">取消</button>
      <button class="btn-save" onclick="saveRule()">儲存</button>
    </div>
  </div>
</div>

<div class="toast" id="toast"></div>

<script>
const KEY = '{{ key }}';
let rules = [];

async function api(path, opts){
  const sep = path.includes('?') ? '&' : '?';
  const r = await fetch(path + sep + 'key=' + KEY, opts);
  if(!r.ok) throw new Error(await r.text());
  return r.json();
}

function toast(msg){
  const t = document.getElementById('toast');
  t.textContent = msg; t.classList.add('show');
  setTimeout(()=>t.classList.remove('show'), 2500);
}

function esc(s){ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }

function pct(v){ const n = parseFloat(v)*100; return (Number.isInteger(n)?n:n.toFixed(1))+'%'; }

const RBL = {round_10:'進10', round_50:'進50', round_100:'進100', charm_9:'尾數9'};

function renderTable(){
  document.getElementById('ruleCount').textContent = rules.length + ' 條規則';
  const wrap = document.getElementById('tableWrap');
  if(!rules.length){
    wrap.innerHTML = '<div class="empty">尚無規則。點右上角「新增品牌規則」開始</div>';
    return;
  }
  const rows = rules.map(r => `<tr>
    <td><strong>${esc(r.brand||'—')}</strong></td>
    <td>${esc(r.category||'—')}</td>
    <td>${parseFloat(r.exchange_rate).toFixed(2)}</td>
    <td>${pct(r.margin_rate)}</td>
    <td>${pct(r.ad_rate)}</td>
    <td>${pct(r.sea_shipping_rate)}</td>
    <td>$${Math.round(r.tw_shipping_cost)}</td>
    <td>$${Math.round(r.package_cost)}</td>
    <td><span class="rb rb-${esc(r.round_rule)}">${esc(RBL[r.round_rule]||r.round_rule)}</span></td>
    <td style="white-space:nowrap">
      <button class="btn-edit" onclick="openEdit(${r.id})">編輯</button>
      <button class="btn-del" onclick="delRule(${r.id})">刪除</button>
    </td>
  </tr>`).join('');
  wrap.innerHTML = `<table>
    <thead><tr>
      <th>品牌</th><th>分類</th><th>匯率</th><th>毛利率</th><th>廣告</th><th>海運</th><th>物流</th><th>包材</th><th>進位</th><th></th>
    </tr></thead>
    <tbody>${rows}</tbody>
  </table>`;
}

async function loadRules(){
  try{
    const d = await api('/api/product-rules');
    rules = d.rules || [];
    renderTable();
  }catch(e){
    document.getElementById('tableWrap').innerHTML = '<div class="empty">載入失敗：'+e.message+'</div>';
  }
}

function resetForm(){
  document.getElementById('editId').value = '';
  document.getElementById('fBrand').value = '';
  document.getElementById('fCategory').value = '';
  document.getElementById('fExchange').value = '4.5';
  document.getElementById('fMargin').value = '40';
  document.getElementById('fAd').value = '10';
  document.getElementById('fSea').value = '5';
  document.getElementById('fTwShipping').value = '80';
  document.getElementById('fPackage').value = '30';
  document.getElementById('fRound').value = 'round_10';
}

function openModal(){
  document.getElementById('modalTitle').textContent = '新增品牌規則';
  resetForm();
  document.getElementById('overlay').classList.add('open');
  document.getElementById('fBrand').focus();
}

function openEdit(id){
  const r = rules.find(x=>x.id===id);
  if(!r) return;
  document.getElementById('modalTitle').textContent = '編輯品牌規則';
  document.getElementById('editId').value = id;
  document.getElementById('fBrand').value = r.brand || '';
  document.getElementById('fCategory').value = r.category || '';
  document.getElementById('fExchange').value = parseFloat(r.exchange_rate).toFixed(2);
  document.getElementById('fMargin').value = (parseFloat(r.margin_rate)*100).toFixed(1).replace(/\\.0$/,'');
  document.getElementById('fAd').value = (parseFloat(r.ad_rate)*100).toFixed(1).replace(/\\.0$/,'');
  document.getElementById('fSea').value = (parseFloat(r.sea_shipping_rate)*100).toFixed(1).replace(/\\.0$/,'');
  document.getElementById('fTwShipping').value = Math.round(r.tw_shipping_cost);
  document.getElementById('fPackage').value = Math.round(r.package_cost);
  document.getElementById('fRound').value = r.round_rule || 'round_10';
  document.getElementById('overlay').classList.add('open');
  document.getElementById('fBrand').focus();
}

function closeModal(){
  document.getElementById('overlay').classList.remove('open');
}

function getFormData(){
  const brand = document.getElementById('fBrand').value.trim();
  if(!brand){ alert('請輸入品牌名稱'); return null; }
  return {
    brand,
    category:         document.getElementById('fCategory').value.trim(),
    exchange_rate:    parseFloat(document.getElementById('fExchange').value) || 4.5,
    margin_rate:      (parseFloat(document.getElementById('fMargin').value) || 40) / 100,
    ad_rate:          (parseFloat(document.getElementById('fAd').value) || 10) / 100,
    sea_shipping_rate:(parseFloat(document.getElementById('fSea').value) || 5) / 100,
    tw_shipping_cost: parseFloat(document.getElementById('fTwShipping').value) || 0,
    package_cost:     parseFloat(document.getElementById('fPackage').value) || 0,
    round_rule:       document.getElementById('fRound').value,
  };
}

async function saveRule(){
  const data = getFormData();
  if(!data) return;
  const editId = document.getElementById('editId').value;
  try{
    if(editId){
      await api('/api/product-rules/'+editId, {method:'PUT', body:JSON.stringify(data), headers:{'Content-Type':'application/json'}});
      toast('已更新');
    }else{
      await api('/api/product-rules', {method:'POST', body:JSON.stringify(data), headers:{'Content-Type':'application/json'}});
      toast('已新增');
    }
    closeModal();
    await loadRules();
  }catch(e){
    alert('儲存失敗：'+e.message);
  }
}

async function delRule(id){
  const r = rules.find(x=>x.id===id);
  if(!confirm('確定刪除「'+(r?r.brand:id)+'」的規則？')) return;
  try{
    await api('/api/product-rules/'+id, {method:'DELETE'});
    toast('已刪除');
    await loadRules();
  }catch(e){
    alert('刪除失敗：'+e.message);
  }
}

document.getElementById('overlay').addEventListener('click', function(e){
  if(e.target === this) closeModal();
});

loadRules();
</script>
</body>
</html>"""


# ── Product Rules Routes ──────────────────────────────────────────

@products_bp.route("/admin/product-rules")
def admin_product_rules():
    ok, key = check_auth()
    if not ok:
        return render_template_string(LOGIN_HTML, next="/admin/product-rules", error="")
    return render_template_string(PRODUCT_RULES_HTML, key=key)

@products_bp.route("/api/product-rules", methods=["GET"])
def api_product_rules_list():
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    return jsonify({"rules": _pr_all()})

@products_bp.route("/api/product-rules", methods=["POST"])
def api_product_rules_create():
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    if not data.get("brand","").strip():
        return jsonify({"error": "brand 必填"}), 400
    row_id = _pr_insert(data)
    if row_id is None:
        return jsonify({"error": "資料庫未設定或寫入失敗"}), 500
    return jsonify({"ok": True, "id": row_id})

@products_bp.route("/api/product-rules/<int:rule_id>", methods=["PUT"])
def api_product_rules_update(rule_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    success = _pr_update(rule_id, data)
    return jsonify({"ok": success})

@products_bp.route("/api/product-rules/<int:rule_id>", methods=["DELETE"])
def api_product_rules_delete(rule_id):
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    _pr_delete(rule_id)
    return jsonify({"ok": True})


# ── Store Import (Batch) ─────────────────────────────────────────

def _ss_migrate():
    """Add brand / category columns to store_scan_jobs."""
    if not DATABASE_URL:
        return
    try:
        conn = _pg_conn(); cur = conn.cursor()
        for sql in [
            "ALTER TABLE store_scan_jobs ADD COLUMN IF NOT EXISTS brand    TEXT DEFAULT ''",
            "ALTER TABLE store_scan_jobs ADD COLUMN IF NOT EXISTS category TEXT DEFAULT ''",
        ]:
            try: cur.execute(sql)
            except Exception: pass
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        import sys; print(f"[SS Migrate] {e}", file=sys.stderr)

_ss_migrate()


def _ss_create(url, platform, brand="", category=""):
    """Insert store_scan_jobs with brand and category; returns job_id."""
    if not DATABASE_URL:
        return None
    try:
        with _db_lock:
            conn = _pg_conn(); cur = conn.cursor()
            cur.execute(
                "INSERT INTO store_scan_jobs "
                "(url,platform,brand,category,status,created_at,updated_at) "
                "VALUES (%s,%s,%s,%s,'pending',%s,%s) RETURNING id",
                (url, platform, brand, category, time.time(), time.time())
            )
            job_id = cur.fetchone()[0]
            conn.commit(); cur.close(); conn.close()
            return job_id
    except Exception as e:
        import sys; print(f"[SS Create] {e}", file=sys.stderr)
        return None


def _parse_price_str(price_str):
    """Extract first number from price string. Returns float or None."""
    import re as _re2
    if not price_str:
        return None
    m = _re2.search(r'[\d.]+', str(price_str))
    if m:
        try: return float(m.group())
        except Exception: pass
    return None


def _ss_get_with_brand(job_id):
    """Extended _ss_get: includes brand/category and maps item fields."""
    if not DATABASE_URL:
        return None
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT id,url,platform,status,item_count,error_msg,created_at,"
            "COALESCE(brand,''),COALESCE(category,'') "
            "FROM store_scan_jobs WHERE id=%s",
            (job_id,)
        )
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close(); return None
        result = {
            "id": row[0], "url": row[1], "platform": row[2],
            "status": row[3], "item_count": row[4],
            "error_msg": row[5], "created_at": row[6],
            "brand": row[7], "category": row[8],
        }
        cur.execute(
            "SELECT id,title,url,image,price,shop_name,platform,scraped_at,added_to_queue "
            "FROM store_scan_items WHERE scan_job_id=%s ORDER BY id ASC",
            (job_id,)
        )
        result["items"] = [
            {
                "id": r[0], "title": r[1],
                "url": r[2], "product_url": r[2],
                "image": r[3], "main_image": r[3],
                "price": r[4], "original_price": _parse_price_str(r[4]),
                "shop_name": r[5], "platform": r[6],
                "scraped_at": r[7], "added_to_queue": bool(r[8]),
            }
            for r in cur.fetchall()
        ]
        cur.close(); conn.close()
        return result
    except Exception as e:
        import sys; print(f"[SS GetBrand] {e}", file=sys.stderr)
        return None


def _pj_import(url, platform, brand="", raw_title="", raw_price="", raw_images=None):
    """Insert product_job pre-filled with scan data (status=pending)."""
    if not DATABASE_URL:
        return None
    try:
        now = time.time()
        imgs_json = json.dumps(raw_images or [], ensure_ascii=False)
        with _db_lock:
            conn = _pg_conn(); cur = conn.cursor()
            cur.execute(
                "INSERT INTO product_jobs "
                "(url,platform,brand,status,raw_title,raw_price,raw_images,created_at,updated_at) "
                "VALUES (%s,%s,%s,'pending',%s,%s,%s,%s,%s) RETURNING id",
                (url, platform, brand, raw_title, raw_price, imgs_json, now, now)
            )
            job_id = cur.fetchone()[0]
            conn.commit(); cur.close(); conn.close()
            return job_id
    except Exception as e:
        import sys; print(f"[PJ Import] {e}", file=sys.stderr)
        return None


# ── Store Import API Routes ───────────────────────────────────────

@products_bp.route("/api/store-scan/create", methods=["POST"])
def api_store_scan_create_v2():
    """建立店鋪掃描任務（含品牌/分類）。"""
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    url = (data.get("url") or "").strip()
    if not url:
        return jsonify({"error": "url 必填"}), 400
    platform = data.get("platform", "")
    if not platform:
        platform = "1688" if "1688.com" in url else "taobao" if "taobao.com" in url else "unknown"
    brand    = (data.get("brand") or "").strip()
    category = (data.get("category") or "").strip()
    job_id = _ss_create(url, platform, brand, category)
    if job_id is None:
        return jsonify({"error": "資料庫未設定或寫入失敗"}), 500
    return jsonify({"ok": True, "id": job_id})


@products_bp.route("/api/store-scan/<int:job_id>/items", methods=["GET"])
def api_store_scan_items_v2(job_id):
    """取得掃描商品清單（含 brand/category，item 欄位映射新格式）。"""
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    job = _ss_get_with_brand(job_id)
    if not job:
        return jsonify({"error": "not found"}), 404
    return jsonify(job)


@products_bp.route("/api/store-scan/import-selected", methods=["POST"])
def api_store_scan_import_selected():
    """將勾選商品建立成 product_jobs（含預填 title/price/image）。"""
    ok, _ = auth_required()
    if not ok:
        return jsonify({"error": "unauthorized"}), 403
    data = request.get_json(silent=True) or {}
    item_ids = data.get("item_ids", [])
    brand    = (data.get("brand") or "").strip()
    if not item_ids:
        return jsonify({"error": "請勾選商品"}), 400
    if not DATABASE_URL:
        return jsonify({"error": "資料庫未設定"}), 500
    try:
        conn = _pg_conn(); cur = conn.cursor()
        cur.execute(
            "SELECT ssi.id, ssi.url, ssi.platform, ssi.title, ssi.price, ssi.image, "
            "COALESCE(ssj.brand,'') "
            "FROM store_scan_items ssi "
            "JOIN store_scan_jobs ssj ON ssj.id = ssi.scan_job_id "
            "WHERE ssi.id=ANY(%s) AND ssi.added_to_queue=FALSE",
            (item_ids,)
        )
        rows = cur.fetchall(); cur.close(); conn.close()
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    added = []
    for r in rows:
        item_id, url, platform, title, price, image, job_brand = r
        eff_brand = brand or job_brand
        raw_images = [image] if image else []
        job_id = _pj_import(url, platform, eff_brand, title or "", price or "", raw_images)
        if job_id:
            added.append({"item_id": item_id, "job_id": job_id})
    if added:
        _ss_mark_added([a["item_id"] for a in added])
    return jsonify({"ok": True, "added": len(added), "jobs": added})


# ── Store Import Admin Page ───────────────────────────────────────

STORE_IMPORT_HTML = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>店鋪批次搬運</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:-apple-system,sans-serif;background:#f5f5f5;color:#333;height:100vh;display:flex;flex-direction:column;overflow:hidden}
.header{background:#1a1a1a;color:#fff;padding:13px 20px;display:flex;align-items:center;gap:14px;flex-shrink:0}
.header a{color:#888;text-decoration:none;font-size:14px}
.header a:hover{color:#fff}
.header-title{font-size:17px;font-weight:700;flex:1}
.layout{display:flex;flex:1;min-height:0}
.panel-l{width:300px;min-width:300px;background:#fff;border-right:1px solid #eee;display:flex;flex-direction:column;overflow:hidden}
.sec{padding:14px 16px;border-bottom:1px solid #f0f0f0;flex-shrink:0}
.sec h4{font-size:11px;font-weight:700;color:#888;margin-bottom:10px;text-transform:uppercase;letter-spacing:.5px}
.frow{margin-bottom:8px}
.frow label{display:block;font-size:11px;font-weight:700;color:#666;margin-bottom:3px}
.frow input,.frow select{width:100%;border:1.5px solid #e0e0e0;border-radius:8px;padding:7px 10px;font-size:13px;outline:none;font-family:inherit;background:#fff}
.frow input:focus,.frow select:focus{border-color:#1a1a1a}
.btn-scan{width:100%;background:#1a1a1a;color:#fff;border:none;border-radius:9px;padding:10px;font-size:13px;font-weight:700;cursor:pointer;font-family:inherit;margin-top:2px}
.btn-scan:hover{background:#333}
.btn-scan:disabled{background:#bbb;cursor:default}
.sbar{margin-top:8px;min-height:16px;font-size:11px;color:#888}
.spinner{display:inline-block;width:9px;height:9px;border:2px solid #ddd;border-top-color:#888;border-radius:50%;animation:spin .8s linear infinite;vertical-align:middle;margin-right:3px}
@keyframes spin{to{transform:rotate(360deg)}}
.task-scroller{flex:1;overflow-y:auto}
.task-item{padding:10px 16px;border-bottom:1px solid #f8f8f8;cursor:pointer;transition:background .1s}
.task-item:hover{background:#f8f8f8}
.task-item.active{background:#e8f4fd;border-left:3px solid #1a1a1a;padding-left:13px}
.ti-url{font-size:11px;font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-bottom:3px}
.ti-meta{display:flex;gap:5px;align-items:center;flex-wrap:wrap}
.pf{font-size:10px;font-weight:700;padding:1px 5px;border-radius:4px}
.pf-1688{background:#fff0f0;color:#c62828}
.pf-taobao{background:#fff4e5;color:#e65100}
.ts{font-size:10px;font-weight:700;padding:1px 6px;border-radius:10px}
.ts-pending,.ts-scanning{background:#fef3cd;color:#856404}
.ts-done{background:#d1fae5;color:#065f46}
.ts-error,.ts-failed{background:#fee2e2;color:#991b1b}
.panel-r{flex:1;display:flex;flex-direction:column;min-width:0}
.ph{padding:12px 18px;background:#fff;border-bottom:1px solid #eee;display:flex;align-items:center;gap:10px;flex-shrink:0}
.ph h3{font-size:14px;font-weight:700;flex:1}
.ph-meta{font-size:12px;color:#aaa}
.pb{flex:1;overflow-y:auto;padding:16px}
.empty-r{display:flex;flex-direction:column;align-items:center;justify-content:center;height:100%;color:#ccc;font-size:13px;gap:8px}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(155px,1fr));gap:10px}
.ic{background:#fff;border-radius:12px;border:2px solid transparent;overflow:hidden;cursor:pointer;transition:border-color .12s;position:relative;display:flex;flex-direction:column}
.ic:hover{border-color:#ddd}
.ic.sel{border-color:#1a1a1a}
.ic.done{opacity:.6}
.ic input[type=checkbox]{position:absolute;top:8px;left:8px;width:15px;height:15px;z-index:2;accent-color:#1a1a1a;cursor:pointer}
.ic img{width:100%;aspect-ratio:1;object-fit:cover;background:#f5f5f5}
.ic-body{padding:7px 9px 9px;flex:1;display:flex;flex-direction:column;gap:3px}
.ic-title{font-size:11px;line-height:1.4;display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden}
.ic-price{font-size:12px;font-weight:700;color:#c62828}
.ic-added{font-size:10px;font-weight:700;padding:1px 5px;border-radius:5px;background:#d1fae5;color:#065f46;align-self:flex-start}
.abar{background:#fff;border-top:1px solid #eee;padding:10px 18px;display:none;align-items:center;gap:10px;flex-shrink:0}
.abar.show{display:flex}
.sel-info{font-size:13px;color:#555;flex:1}
.sb-row{display:flex;gap:6px}
.sbtn{background:#f5f5f5;border:1.5px solid #e0e0e0;border-radius:8px;padding:6px 12px;font-size:12px;cursor:pointer;font-family:inherit}
.sbtn:hover{background:#eee}
.brand-pick{border:1.5px solid #e0e0e0;border-radius:8px;padding:7px 10px;font-size:12px;font-family:inherit;outline:none;background:#fff}
.btn-imp{background:#2e7d32;color:#fff;border:none;border-radius:9px;padding:9px 20px;font-size:13px;font-weight:700;cursor:pointer;font-family:inherit;white-space:nowrap}
.btn-imp:hover{background:#1b5e20}
.btn-imp:disabled{background:#bbb;cursor:default}
.toast{position:fixed;bottom:60px;left:50%;transform:translateX(-50%);background:#222;color:#fff;padding:8px 18px;border-radius:18px;font-size:13px;z-index:9999;opacity:0;transition:opacity .3s;pointer-events:none}
.toast.show{opacity:1}
</style>
</head>
<body>

<div class="header">
  <a href="/admin/products?key={{ key }}">&#8592; 商品搬運</a>
  <div class="header-title">店鋪批次搬運</div>
</div>

<div class="layout">

  <div class="panel-l">
    <div class="sec">
      <h4>新增掃描任務</h4>
      <div class="frow">
        <label>店鋪 / 分類頁網址</label>
        <input type="text" id="urlIn" placeholder="https://shop.1688.com/...">
      </div>
      <div class="frow">
        <label>平台</label>
        <select id="pfSel">
          <option value="1688">1688</option>
          <option value="taobao">淘寶</option>
        </select>
      </div>
      <div class="frow">
        <label>品牌</label>
        <select id="brandSel">
          <option value="">不指定</option>
          <option value="jsimple">JS家具</option>
          <option value="lander">朗德LIGHT+</option>
          <option value="filterbreath">濾呼吸</option>
          <option value="lander_curtain">澄光窗簾</option>
        </select>
      </div>
      <div class="frow">
        <label>分類</label>
        <input type="text" id="catIn" placeholder="例：高架床、窗簾">
      </div>
      <button class="btn-scan" id="scanBtn" onclick="startScan()">開始掃描</button>
      <div class="sbar" id="sbar"></div>
    </div>
    <div class="sec" style="padding-bottom:6px">
      <h4>掃描歷史</h4>
    </div>
    <div class="task-scroller" id="taskList">
      <div style="padding:14px 16px;font-size:12px;color:#ccc">載入中...</div>
    </div>
  </div>

  <div class="panel-r">
    <div class="ph" id="ph" style="display:none">
      <h3 id="phTitle">商品清單</h3>
      <span class="ph-meta" id="phMeta"></span>
    </div>
    <div class="pb" id="pb">
      <div class="empty-r">
        <div style="font-size:36px">&#128722;</div>
        <div>從左側選擇掃描任務，查看商品清單</div>
      </div>
    </div>
    <div class="abar" id="abar">
      <div class="sb-row">
        <button class="sbtn" onclick="selAll(true)">全選</button>
        <button class="sbtn" onclick="selAll(false)">取消</button>
      </div>
      <span class="sel-info" id="selInfo">已選 0 筆</span>
      <select class="brand-pick" id="impBrand">
        <option value="">不指定品牌</option>
        <option value="jsimple">JS家具</option>
        <option value="lander">朗德LIGHT+</option>
        <option value="filterbreath">濾呼吸</option>
        <option value="lander_curtain">澄光窗簾</option>
      </select>
      <button class="btn-imp" id="impBtn" onclick="doImport()">加入商品搬運</button>
    </div>
  </div>

</div>

<div class="toast" id="toast"></div>

<script>
const KEY = '{{ key }}';
let currentJobId = null;
let pollTimer = null;

async function api(path, opts){
  const sep = path.includes('?') ? '&' : '?';
  const r = await fetch(path + sep + 'key=' + KEY, opts);
  if(!r.ok) throw new Error(await r.text());
  return r.json();
}
function toast(msg){
  const t = document.getElementById('toast');
  t.textContent = msg; t.classList.add('show');
  setTimeout(()=>t.classList.remove('show'), 2800);
}
function esc(s){ return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }

async function loadHistory(){
  try{
    const d = await api('/api/store-scan');
    const jobs = d.jobs||[];
    const tl = document.getElementById('taskList');
    if(!jobs.length){
      tl.innerHTML = '<div style="padding:14px 16px;font-size:11px;color:#ccc">尚無記錄</div>';
      return;
    }
    tl.innerHTML = jobs.slice(0,30).map(j=>`
      <div class="task-item${j.id===currentJobId?' active':''}" onclick="loadTask(${j.id})" data-jid="${j.id}">
        <div class="ti-url">${esc((j.url||'').replace(/https?:\\/\\//,'').slice(0,42))}</div>
        <div class="ti-meta">
          <span class="pf pf-${esc(j.platform)}">${esc(j.platform)}</span>
          <span class="ts ts-${esc(j.status)}">${esc(j.status)}</span>
          <span style="font-size:10px;color:#aaa">${j.item_count||0}筆</span>
        </div>
      </div>
    `).join('');
  }catch(e){}
}

async function loadTask(jobId){
  clearTimeout(pollTimer);
  currentJobId = jobId;
  document.querySelectorAll('.task-item').forEach(el=>el.classList.toggle('active', +el.dataset.jid===jobId));
  const ph = document.getElementById('ph');
  ph.style.display = 'flex';
  document.getElementById('phTitle').textContent = '商品清單';
  document.getElementById('phMeta').textContent = '載入中...';
  document.getElementById('pb').innerHTML = '<div style="text-align:center;padding:50px;color:#bbb;font-size:13px">載入中...</div>';
  document.getElementById('abar').classList.remove('show');
  pollTask(jobId);
}

async function pollTask(jobId){
  try{
    const job = await api('/api/store-scan/'+jobId+'/items');
    if(job.status==='pending'||job.status==='scanning'){
      document.getElementById('phMeta').textContent = '掃描中，請稍候...';
      document.getElementById('pb').innerHTML = '<div style="text-align:center;padding:50px;color:#aaa;font-size:13px"><span class="spinner"></span> 等待 local_worker 掃描中...</div>';
      pollTimer = setTimeout(()=>pollTask(jobId), 2500);
      return;
    }
    await loadHistory();
    renderItems(job);
  }catch(e){
    document.getElementById('phMeta').textContent = '載入失敗';
    pollTimer = setTimeout(()=>pollTask(jobId), 3500);
  }
}

function renderItems(job){
  const items = job.items||[];
  document.getElementById('phMeta').textContent = items.length + ' 筆';
  if(job.status==='error'||job.status==='failed'){
    document.getElementById('pb').innerHTML = '<div style="text-align:center;padding:40px;color:#c62828;font-size:13px">掃描失敗：'+esc(job.error_msg||'')+'</div>';
    document.getElementById('abar').classList.remove('show');
    return;
  }
  if(!items.length){
    document.getElementById('pb').innerHTML = '<div style="text-align:center;padding:50px;color:#bbb;font-size:13px">未抓到商品<br><br>請確認 local_worker 正在執行，<br>或此頁面格式已變更。</div>';
    document.getElementById('abar').classList.remove('show');
    return;
  }
  document.getElementById('pb').innerHTML = '<div class="grid">'+items.map(it=>`
    <label class="ic${it.added_to_queue?' done':''}" onclick="updCount()">
      <input type="checkbox" class="icb" value="${it.id}"${it.added_to_queue?' disabled checked':''}>
      <img src="${esc(it.image||it.main_image||'')}" onerror="this.style.background='#eee'" alt="">
      <div class="ic-body">
        <div class="ic-title">${esc(it.title)}</div>
        ${it.price?'<div class="ic-price">&#165;'+esc(it.price)+'</div>':''}
        ${it.added_to_queue?'<span class="ic-added">&#10003; 已加入</span>':''}
      </div>
    </label>
  `).join('')+'</div>';
  document.getElementById('abar').classList.add('show');
  if(job.brand) document.getElementById('impBrand').value = job.brand;
  updCount();
}

function updCount(){
  const n = document.querySelectorAll('.icb:not(:disabled):checked').length;
  document.getElementById('selInfo').textContent = '已選 ' + n + ' 筆';
}
function selAll(v){
  document.querySelectorAll('.icb:not(:disabled)').forEach(el=>{ el.checked=v; });
  document.querySelectorAll('.ic').forEach(el=>{ const cb=el.querySelector('.icb'); if(cb&&!cb.disabled) el.classList.toggle('sel',v); });
  updCount();
}

async function startScan(){
  const url = document.getElementById('urlIn').value.trim();
  if(!url){ alert('請輸入店鋪網址'); return; }
  const platform = document.getElementById('pfSel').value;
  const brand    = document.getElementById('brandSel').value;
  const category = document.getElementById('catIn').value.trim();
  document.getElementById('scanBtn').disabled = true;
  document.getElementById('sbar').innerHTML = '<span class="spinner"></span>建立掃描任務...';
  try{
    const r = await api('/api/store-scan/create',{
      method:'POST', body:JSON.stringify({url,platform,brand,category}),
      headers:{'Content-Type':'application/json'}
    });
    document.getElementById('sbar').textContent = '任務 #'+r.id+' 已建立，等待 worker...';
    await loadHistory();
    await loadTask(r.id);
  }catch(e){
    document.getElementById('sbar').textContent = '錯誤：'+e.message;
  }finally{
    document.getElementById('scanBtn').disabled = false;
  }
}

async function doImport(){
  const ids = [...document.querySelectorAll('.icb:not(:disabled):checked')].map(el=>parseInt(el.value));
  if(!ids.length){ alert('請勾選商品'); return; }
  const brand = document.getElementById('impBrand').value;
  document.getElementById('impBtn').disabled = true;
  try{
    const r = await api('/api/store-scan/import-selected',{
      method:'POST', body:JSON.stringify({item_ids:ids,brand}),
      headers:{'Content-Type':'application/json'}
    });
    toast('已加入 '+r.added+' 筆商品搬運任務');
    if(currentJobId) await loadTask(currentJobId);
  }catch(e){
    alert('加入失敗：'+e.message);
  }finally{
    document.getElementById('impBtn').disabled = false;
  }
}

loadHistory();
</script>
</body>
</html>"""


@products_bp.route("/admin/store-import")
def admin_store_import():
    ok, key = check_auth()
    if not ok:
        return render_template_string(LOGIN_HTML, next="/admin/store-import", error="")
    return render_template_string(STORE_IMPORT_HTML, key=key)

